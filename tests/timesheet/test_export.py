"""T4 — `GET /sites/{site_id}/timesheet/export.xlsx` (spec §3).

Çıktı, T3 matrisinin AYNI türevlerini taşır: kişi satırları + gün sütunları +
kişi toplam saati + adam-günü + günlük alt toplam satırı. Hesap `matrix.py`de
kalır, bu uç yalnız onu çalışma kitabına çevirir — iki farklı toplam mantığı
ekranla Excel'i ayrıştırırdı.

Kapı `timesheet:view`tir (indirme bir OKUMADIR): matrisi göremeyen indiremez,
gören indirebilir. Görünmeyen şantiye 404'tür (var olmayanla ayırt EDİLEMEZ).
"""

import enum
from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest
from httpx import AsyncClient

from app.modules.site_diary.models import WorkerSource
from app.modules.timesheet.export import (
    COLUMN_HEADERS,
    DAY_TOTAL_LABEL,
    EMPTY_VALUE,
    HEADER_ROW,
    HOURS_TOTAL_HEADER,
    INFO_LABELS,
    SHEET_TITLE,
    SOURCE_LABELS,
    XLSX_MEDIA_TYPE,
    source_label,
)
from app.modules.timesheet.models import TimesheetCode
from tests.timesheet.conftest import AY, YIL, gun

pytestmark = pytest.mark.asyncio

_I = TimesheetCode.leave
_T = TimesheetCode.holiday
_G = TimesheetCode.temporary_duty


async def _indir(client: AsyncClient, headers, site_id, **params):
    sorgu = {"year": YIL, "month": AY, **params}
    yanit = await client.get(
        f"/sites/{site_id}/timesheet/export.xlsx", params=sorgu, headers=headers
    )
    assert yanit.status_code == 200, yanit.text
    return yanit


def _sayfa(yanit):
    workbook = openpyxl.load_workbook(BytesIO(yanit.content))
    return workbook[SHEET_TITLE]


def _satir(sheet, row: int) -> list:
    return [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]


def _kisi_satiri(sheet, ad: str) -> list:
    for row in range(HEADER_ROW + 1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == ad:
            return _satir(sheet, row)
    raise AssertionError(f"kişi satırı yok: {ad}")


@pytest.fixture
async def ornek_ay(hucre_fabrikasi, santiye, admin_kullanicisi, mehmet, ali, bolum):
    """İki kişi, ilk üç gün. Mehmet 9 + 12,5 saat + tatil; Ali 9 saat + izin + görev.

    Ali'nin 3. günü `G`dir — alt toplam satırındaki "G" işaretinin kaynağı.
    """
    await hucre_fabrikasi(santiye, mehmet, gun(1), admin_kullanicisi, hours=9, section=bolum)
    await hucre_fabrikasi(
        santiye, mehmet, gun(2), admin_kullanicisi, hours=Decimal("12.5"), section=bolum
    )
    await hucre_fabrikasi(santiye, mehmet, gun(3), admin_kullanicisi, code=_T, section=bolum)
    await hucre_fabrikasi(santiye, ali, gun(1), admin_kullanicisi, hours=9, section=bolum)
    await hucre_fabrikasi(santiye, ali, gun(2), admin_kullanicisi, code=_I, section=bolum)
    await hucre_fabrikasi(santiye, ali, gun(3), admin_kullanicisi, code=_G, section=bolum)


# --- HTTP zarfı (boq/audit export deseni) ---


async def test_xlsx_content_type_ve_dosya_adi(client, admin_headers, santiye, ornek_ay):
    """`Content-Type` xlsx'tir ve dosya adı şantiye kodu + dönemden türer."""
    yanit = await _indir(client, admin_headers, santiye.id)

    assert yanit.headers["content-type"] == XLSX_MEDIA_TYPE
    disposition = yanit.headers["content-disposition"]
    assert disposition.startswith("attachment;")
    assert f"puantaj-{santiye.code}-{YIL}-{AY:02d}.xlsx" in disposition


# --- İçerik: başlık şeridi (ŞP 116-119) ---


async def test_baslik_seridi_matrisin_turevleriyle_ayni(
    client, admin_headers, santiye, proje, bolum, ornek_ay
):
    """Bölüm · işçi sayısı · toplam saat · TÜREV adam-gün.

    Toplam saat 30,5 (9 + 12,5 + 9) -> 30,5/9 = 3,4 adam-gün. `İ`, `T` ve `G`
    saate 0 katar.
    """
    sheet = _sayfa(await _indir(client, admin_headers, santiye.id, section_id=str(bolum.id)))

    bilgi = {
        sheet.cell(row=i, column=1).value: sheet.cell(row=i, column=2).value
        for i in range(1, len(INFO_LABELS) + 1)
    }
    assert bilgi[INFO_LABELS[0]] == proje.name
    assert bilgi[INFO_LABELS[1]] == santiye.name
    assert bilgi[INFO_LABELS[2]] == f"{AY:02d}.{YIL}"
    assert bilgi[INFO_LABELS[3]] == bolum.name
    assert bilgi[INFO_LABELS[4]] == "2"
    assert bilgi[INFO_LABELS[5]] == "30.5"
    assert bilgi[INFO_LABELS[6]] == "3.4"


async def test_bolum_secilmemisse_serit_bolum_iddia_etmez(client, admin_headers, santiye, ornek_ay):
    sheet = _sayfa(await _indir(client, admin_headers, santiye.id))
    assert sheet.cell(row=4, column=1).value == INFO_LABELS[3]
    assert sheet.cell(row=4, column=2).value == EMPTY_VALUE


# --- İçerik: tablo ---


async def test_sutun_basliklari_kisi_alanlari_gunler_ve_toplam(
    client, admin_headers, santiye, ornek_ay
):
    """Sabit kişi sütunları + ayın TÜM günleri (Temmuz = 31) + iki toplam sütunu."""
    sheet = _sayfa(await _indir(client, admin_headers, santiye.id))
    basliklar = _satir(sheet, HEADER_ROW)

    assert basliklar[: len(COLUMN_HEADERS)] == list(COLUMN_HEADERS)
    assert basliklar[len(COLUMN_HEADERS) : len(COLUMN_HEADERS) + 31] == [
        str(day) for day in range(1, 32)
    ]
    assert basliklar[-2:] == [HOURS_TOTAL_HEADER, "Adam-Gün"]


async def test_kisi_satiri_ad_meslek_tur_firma_kod_harfleri_ve_adam_gun(
    client, admin_headers, santiye, ornek_ay
):
    """ŞP 149-150 / 169-170: taşeron firma AYRI sütundur, meslekle birleştirilmez."""
    sheet = _sayfa(await _indir(client, admin_headers, santiye.id))

    mehmet_satir = _kisi_satiri(sheet, "Mehmet Yılmaz")
    assert mehmet_satir[:4] == ["Mehmet Yılmaz", "Kalıpçı Usta", "Şirket", EMPTY_VALUE]
    # 🔴 Çalışılan gün artık HARF değil SAAT basar; kodlu gün harf kalır.
    assert mehmet_satir[4:7] == ["9.0", "12.5", "T"]
    assert mehmet_satir[7] is None  # kaydı olmayan gün BOŞ kalır
    assert mehmet_satir[-2:] == ["21.5", "2.4"]

    ali_satir = _kisi_satiri(sheet, "Ali Kaya")
    assert ali_satir[:4] == ["Ali Kaya", "Demir Ustası", "Taşeron", "Akın İnşaat"]
    assert ali_satir[4:7] == ["9.0", "İ", "G"]
    assert ali_satir[-2:] == ["9.0", "1.0"]


async def test_alt_toplam_satiri_gunluk_sayilar_arti_ve_g_isareti(
    client, admin_headers, santiye, ornek_ay
):
    """Gün sütunu artık SAAT toplar; `G` yalnızca bir işarettir, sayıyı değiştirmez.

    🔴 `+` işareti KALKTI: FM haftalık bir türevdir, bir gün sütununda
    hesaplanamaz.
    """
    sheet = _sayfa(await _indir(client, admin_headers, santiye.id))
    toplam = _kisi_satiri(sheet, DAY_TOTAL_LABEL)

    assert toplam[4] == "18.0"  # 1. gün: 9 + 9
    assert toplam[5] == "12.5"  # 2. gün: Mehmet 12,5 + Ali İzin (0)
    assert toplam[6] == "0.0G"  # 3. gün: Mehmet T, Ali G — 0 saat, `G` işaretli
    assert toplam[-2:] == ["30.5", "3.4"]


async def test_bolum_filtresi_ciktiyi_daraltir(
    client,
    admin_headers,
    santiye,
    admin_kullanicisi,
    hucre_fabrikasi,
    personel_fabrikasi,
    ikinci_bolum,
    ornek_ay,
):
    diger = await personel_fabrikasi("Veli Ak", trade="Sıvacı")
    await hucre_fabrikasi(santiye, diger, gun(1), admin_kullanicisi, hours=9, section=ikinci_bolum)

    sheet = _sayfa(await _indir(client, admin_headers, santiye.id, section_id=str(ikinci_bolum.id)))
    adlar = [
        sheet.cell(row=row, column=1).value for row in range(HEADER_ROW + 1, sheet.max_row + 1)
    ]
    assert adlar == ["Veli Ak", DAY_TOTAL_LABEL]


async def test_bos_donemde_gecerli_dosya_doner(client, admin_headers, santiye):
    """Kayıtsız ay boş dosya DEĞİL, başlıklı ve sıfır toplamlı bir çalışma kitabıdır."""
    sheet = _sayfa(await _indir(client, admin_headers, santiye.id, year=2030, month=2))

    assert _satir(sheet, HEADER_ROW)[: len(COLUMN_HEADERS)] == list(COLUMN_HEADERS)
    toplam = _kisi_satiri(sheet, DAY_TOTAL_LABEL)
    assert toplam[-2:] == ["0.0", "0.0"]
    assert len([c for c in _satir(sheet, HEADER_ROW) if c is not None]) == len(COLUMN_HEADERS) + 30


# --- İzin ve IDOR ---


async def test_saha_muhendisi_indirebilir(client, saha_headers, santiye, ornek_ay):
    """`timesheet:view` YETER: indirme bir okumadır, `full` istenmez (spec §3)."""
    yanit = await _indir(client, saha_headers, santiye.id)
    assert yanit.content


async def test_proje_muduru_403(client, pm_headers, santiye, ornek_ay):
    yanit = await client.get(
        f"/sites/{santiye.id}/timesheet/export.xlsx",
        params={"year": YIL, "month": AY},
        headers=pm_headers,
    )
    assert yanit.status_code == 403, yanit.text


async def test_gorunmeyen_santiye_404(client, sef_headers, gorunmeyen_santiye):
    """Görünmeyen projenin GERÇEK şantiyesi ile var olmayan kimlik AYIRT EDİLEMEZ."""
    yanit = await client.get(
        f"/sites/{gorunmeyen_santiye.id}/timesheet/export.xlsx",
        params={"year": YIL, "month": AY},
        headers=sef_headers,
    )
    assert yanit.status_code == 404, yanit.text


async def test_baska_santiyenin_bolumu_404(client, admin_headers, santiye, yabanci_bolum):
    yanit = await client.get(
        f"/sites/{santiye.id}/timesheet/export.xlsx",
        params={"year": YIL, "month": AY, "section_id": str(yabanci_bolum.id)},
        headers=admin_headers,
    )
    assert yanit.status_code == 404, yanit.text


# --- İK-3 regresyonu: `worker_source` enum genişlemesi -----------------------


async def test_yeni_kaynak_tipleri_export_kirmaz(
    client, admin_headers, santiye, admin_kullanicisi, personel_fabrikasi, hucre_fabrikasi
):
    """İK-3 `worker_source`a `freelance` + `intern` EKLEDİ (BY 243/271).

    `SOURCE_LABELS` üç etiketliydi ve doğrudan indeksleniyordu: yeni kaynaklı bir
    personel matrise girdiği anda `KeyError` → **500**. Delik canlıda henüz
    yoktu (enum değerleri deploy edilmemişti) ama İK-3 merge'üyle açılırdı —
    yani bu dilimin ürettiği bir kırıktır, borç olarak devredilemez.

    Etiketler mockup'tan: BY 253 rozeti "Serbest" · BY 281 rozeti "Stajyer".
    """
    serbest = await personel_fabrikasi(
        "Kemal Tunç", trade="Mühendis", source=WorkerSource.freelance
    )
    stajyer = await personel_fabrikasi(
        "Burak Aydın", trade="İnşaat Müh. Staj", source=WorkerSource.intern
    )
    for kisi in (serbest, stajyer):
        await hucre_fabrikasi(santiye, kisi, gun(1), admin_kullanicisi, hours=9)

    sheet = _sayfa(await _indir(client, admin_headers, santiye.id))

    assert _kisi_satiri(sheet, "Kemal Tunç")[2] == "Serbest"
    assert _kisi_satiri(sheet, "Burak Aydın")[2] == "Stajyer"


async def test_bilinmeyen_kaynak_etiketi_sessiz_degil():
    """Enum bir daha genişlerse export 500 VERMEZ ama sessizce de yalan söylemez.

    Boş string ya da "Genel" gibi bir düşüş, tanınmayan kaynağı tanınmış gibi
    gösterirdi (WORKFLOW §3: sessiz düşüş yok). Ham değer görünür bir işaretle
    basılır — hücreyi okuyan eksik etiketi hemen fark eder.
    """

    class _SahteKaynak(str, enum.Enum):
        yeni_tip = "yeni_tip"

    etiket = source_label(_SahteKaynak.yeni_tip)

    assert "yeni_tip" in etiket
    assert etiket not in SOURCE_LABELS.values()
