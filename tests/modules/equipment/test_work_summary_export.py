"""EXPORT-XLSX — `GET /equipment/work-summary/export.xlsx` (mockup M3 "Excel İndir").

## Neden ÖZET yüzeyi (ölçülmüş karar)

`work-logs` satırları yalnız UUID taşır (ad YOK); o yüzeyden üretilen bir Excel
ya okunmaz UUID kolonları basardı ya da adları çözmek için İKİNCİ bir sorgu yolu
açardı. `work-summary` sayfalanmamıştır ve `equipment_name` taşır.

## Bu dosyanın ASIL iddiası: dosya = EKRAN

Satırlar sayılmaz, **KÜMELER karşılaştırılır**. Fixture'lar ürünün kendi
uçlarından kurulur (`POST /equipment`, `POST /equipment/work-logs`).

## 🔴 K16 — toplam ZARFTAN gelir, satırlardan YENİDEN TOPLANMAZ

Kapasitesi 0 olan makinenin `usage_pct`i `null`dur ve `usage_pct_avg`
paydasına GİRMEZ. Export kendi ortalamasını kursaydı (payda = TÜM satırlar)
başka bir sayı çıkardı — `test_toplam_satiri_zarftan_gelir` tam olarak bu farkı
ölçer.
"""

import json
from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest
from sqlalchemy import func, select

from app.modules.audit.models import AuditLog
from app.modules.equipment.models import EquipmentCategory, EquipmentRatePeriod, WorkLogType
from app.modules.equipment.work_summary_export import (
    COLUMN_HEADERS,
    SHEET_TITLE,
    TOTAL_LABEL,
    XLSX_MEDIA_TYPE,
    filename,
)

pytestmark = pytest.mark.asyncio

YIL = 2026
AY = 7
YOL = "/equipment/work-summary/export.xlsx"
EKRAN = "/equipment/work-summary"
DONEM = {"year": YIL, "month": AY}


def _sayfa(resp):
    return openpyxl.load_workbook(BytesIO(resp.content))[SHEET_TITLE]


def _dosya_satirlari(sheet) -> set[tuple]:
    """Başlık ve TOPLAM satırı DIŞINDAKİ satırların kümesi."""
    return {
        tuple(sheet.cell(row=row, column=col).value for col in range(1, len(COLUMN_HEADERS) + 1))
        for row in range(2, sheet.max_row)
    }


def _toplam_satiri(sheet) -> tuple:
    row = sheet.max_row
    satir = tuple(
        sheet.cell(row=row, column=col).value for col in range(1, len(COLUMN_HEADERS) + 1)
    )
    assert satir[0] == TOTAL_LABEL, satir
    return satir


def _ekran_satiri(row: dict) -> tuple:
    """M3 sütun sırasına göre BEKLENEN hücre demeti — export'un `_cells`i
    KULLANILMAZ (test kendi kendini doğrulamamalı)."""
    return (
        row["equipment_name"],
        row["hours"],
        row["usage_pct"],
        row["breakdown_hours"],
        row["cost"],
    )


def _ekran_zarfi(resp) -> dict:
    """`parse_float=str`: para/saat hücreleri API'nin YAZDIĞI METİNLE ölçülür."""
    return json.loads(resp.text, parse_float=str)


async def _ekipman(client, headers, name: str, **govde) -> dict:
    """`purchase_amount` K2 gereği ŞİRKETE AİT ekipmanda ZORUNLUDUR (422).

    M3 tablosunun kolonu değildir; yalnız kaydın açılabilmesi için verilir.
    """
    resp = await client.post(
        "/equipment",
        json={
            "name": name,
            "category": EquipmentCategory.machinery.value,
            "purchase_amount": "100000.00",
            **govde,
        },
        headers=headers,
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


async def _kayit(client, headers, ekipman: dict, *, gun: int, hours: str, tip: str) -> dict:
    resp = await client.post(
        "/equipment/work-logs",
        json={
            "equipment_id": ekipman["id"],
            "work_date": f"{YIL}-{AY:02d}-{gun:02d}",
            "site_id": ekipman["site_id"],
            "record_type": tip,
            "hours": hours,
        },
        headers=headers,
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


@pytest.fixture
async def park(client, admin_headers, gorunen_santiye):
    """Üç makine — biri bedelsiz + kapasitesiz (K16'nın taşıyıcısı).

    Hepsi ürünün kendi uçlarından açılır ve kayıtları `POST /work-logs`ten
    doğar; saat/maliyet sunucu hesabıdır.
    """
    vinc = await _ekipman(
        client,
        admin_headers,
        "Tower Crane TC-48",
        site_id=str(gorunen_santiye.id),
        rate_amount="3200.00",
        rate_period=EquipmentRatePeriod.daily.value,
        monthly_capacity_hours=200,
    )
    pompa = await _ekipman(
        client,
        admin_headers,
        "Beton Pompası BP-36",
        site_id=str(gorunen_santiye.id),
        rate_amount="2200.00",
        rate_period=EquipmentRatePeriod.daily.value,
        monthly_capacity_hours=200,
    )
    # 🔴 K16: bedeli ve kapasitesi TANIMSIZ — `cost` ve `usage_pct` `null` kalır.
    forklift = await _ekipman(
        client,
        admin_headers,
        "Forklift Linde H30",
        site_id=str(gorunen_santiye.id),
        monthly_capacity_hours=0,
    )

    await _kayit(client, admin_headers, vinc, gun=1, hours="9.00", tip=WorkLogType.worked.value)
    await _kayit(client, admin_headers, vinc, gun=2, hours="8.00", tip=WorkLogType.worked.value)
    await _kayit(client, admin_headers, pompa, gun=1, hours="6.00", tip=WorkLogType.worked.value)
    await _kayit(client, admin_headers, pompa, gun=3, hours="4.00", tip=WorkLogType.breakdown.value)
    await _kayit(client, admin_headers, forklift, gun=1, hours="5.00", tip=WorkLogType.worked.value)
    return {"vinc": vinc, "pompa": pompa, "forklift": forklift}


# --- Kapılar: POZİTİF KONTROLLÜ ---------------------------------------------


async def test_yetkisiz_rol_indiremez_403(client, yetkisiz_headers, park):
    """`hr_manager` (`equipment=_N`) — okumada bile 403."""
    resp = await client.get(YOL, params=DONEM, headers=yetkisiz_headers)
    assert resp.status_code == 403


async def test_view_seviyesi_dosyayi_GERCEKTEN_alir(client, muhendis_headers, park):
    """🔴 POZİTİF KONTROL — uç herkese 403 verecek şekilde bozulursa yakalanır.

    `field_engineer` (`equipment=_V`) ekranı görür; dosyayı da GERÇEKTEN alır.
    """
    resp = await client.get(YOL, params=DONEM, headers=muhendis_headers)
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == XLSX_MEDIA_TYPE
    assert _sayfa(resp).max_row == 2 + len(park)  # başlık + üç satır + toplam


async def test_kimliksiz_401(client):
    resp = await client.get(YOL, params=DONEM)
    assert resp.status_code == 401


# --- Küme eşitliği: dosya = EKRAN -------------------------------------------


async def test_kume_esitligi(client, admin_headers, park):
    ekran = await client.get(EKRAN, params=DONEM, headers=admin_headers)
    assert ekran.status_code == 200, ekran.text
    dosya = await client.get(YOL, params=DONEM, headers=admin_headers)
    assert dosya.status_code == 200, dosya.text

    beklenen = {_ekran_satiri(satir) for satir in _ekran_zarfi(ekran)["rows"]}
    assert len(beklenen) == 3
    assert _dosya_satirlari(_sayfa(dosya)) == beklenen


async def test_kume_esitligi_santiye_suzgecli(client, admin_headers, park, gorunmeyen_santiye):
    """Ekranın `site_id` süzgeci dosyada da AYNI kümeyi vermelidir."""
    params = {**DONEM, "site_id": str(gorunmeyen_santiye.id)}
    ekran = await client.get(EKRAN, params=params, headers=admin_headers)
    dosya = await client.get(YOL, params=params, headers=admin_headers)
    assert ekran.status_code == 200 and dosya.status_code == 200

    assert _ekran_zarfi(ekran)["rows"] == []
    assert _dosya_satirlari(_sayfa(dosya)) == set()


async def test_gorunmeyen_proje_dosyaya_SIZMAZ(
    client, sef_headers, admin_headers, park, gorunmeyen_santiye
):
    """Kapsam kapısı (`_visible_project_ids`) dosyada da geçerlidir.

    Şefin göremediği projedeki makine, admin'in dosyasında VARDIR ama şefin
    dosyasında YOKTUR — dosya ekranın kapsamını GENİŞLETMEZ.
    """
    gizli = await _ekipman(
        client,
        admin_headers,
        "Gizli Ekskavatör",
        site_id=str(gorunmeyen_santiye.id),
        rate_amount="2800.00",
        rate_period=EquipmentRatePeriod.daily.value,
    )
    await _kayit(client, admin_headers, gizli, gun=4, hours="7.00", tip=WorkLogType.worked.value)

    adminin = _dosya_satirlari(_sayfa(await client.get(YOL, params=DONEM, headers=admin_headers)))
    sefin = _dosya_satirlari(_sayfa(await client.get(YOL, params=DONEM, headers=sef_headers)))
    adlar_admin = {satir[0] for satir in adminin}
    adlar_sef = {satir[0] for satir in sefin}
    assert "Gizli Ekskavatör" in adlar_admin
    assert "Gizli Ekskavatör" not in adlar_sef


# --- 🔴 K16: toplam ZARFTAN gelir -------------------------------------------


async def test_toplam_satiri_zarftan_gelir(client, admin_headers, park):
    """Toplam satırı `totals`tan OKUNUR — satırlardan yeniden toplanmaz.

    Ayırt edici sayı `usage_pct_avg`dir: zarf yalnız BİLİNEN kullanımları
    ortalar; export kendi ortalamasını kursaydı payda TÜM satırlar olur ve
    başka bir sayı çıkardı. Aşağıda iki sayının FARKLI olduğu da ölçülür,
    yoksa test hiçbir şeyi ayırt etmezdi.
    """
    zarf = _ekran_zarfi(await client.get(EKRAN, params=DONEM, headers=admin_headers))
    toplamlar = zarf["totals"]
    bilinen = [Decimal(s["usage_pct"]) for s in zarf["rows"] if s["usage_pct"] is not None]
    assert len(bilinen) < len(zarf["rows"]), "hiç `null` kullanım yok — mutant ayırt edilemez"
    naif_ortalama = sum(bilinen) / len(zarf["rows"])
    assert Decimal(toplamlar["usage_pct_avg"]) != naif_ortalama

    satir = _toplam_satiri(_sayfa(await client.get(YOL, params=DONEM, headers=admin_headers)))
    assert satir[COLUMN_HEADERS.index("Çalışma (Saat)")] == toplamlar["hours"]
    assert satir[COLUMN_HEADERS.index("Kullanım %")] == toplamlar["usage_pct_avg"]
    assert satir[COLUMN_HEADERS.index("Arıza (Saat)")] == toplamlar["breakdown_hours"]
    assert satir[COLUMN_HEADERS.index("Maliyet")] == toplamlar["cost"]


async def test_bilinmeyen_maliyet_hucresine_dokunulmaz(client, admin_headers, park):
    """K16: `cost`/`usage_pct` `null` ise hücre `None` kalır — 0 YAZILMAZ."""
    sheet = _sayfa(await client.get(YOL, params=DONEM, headers=admin_headers))
    satir = next(s for s in _dosya_satirlari(sheet) if s[0] == "Forklift Linde H30")
    assert satir[COLUMN_HEADERS.index("Maliyet")] is None
    assert satir[COLUMN_HEADERS.index("Kullanım %")] is None


# --- Boş küme ---------------------------------------------------------------


async def test_bos_kume_gecerli_dosya(client, admin_headers):
    resp = await client.get(YOL, params={"year": YIL, "month": 1}, headers=admin_headers)
    assert resp.status_code == 200, resp.text
    sheet = _sayfa(resp)
    assert [c.value for c in sheet[1]] == list(COLUMN_HEADERS)
    assert sheet.max_row == 2
    assert _toplam_satiri(sheet)[COLUMN_HEADERS.index("Maliyet")] == "0"


# --- Para: API metniyle BAYT BİREBİR ----------------------------------------


async def test_para_hucresi_api_metniyle_birebir(client, admin_headers, park):
    zarf = _ekran_zarfi(await client.get(EKRAN, params=DONEM, headers=admin_headers))
    beklenen = next(s for s in zarf["rows"] if s["equipment_name"] == "Tower Crane TC-48")
    assert beklenen["cost"] is not None, "maliyet hesaplanmadı — test ölçmüyor"

    sheet = _sayfa(await client.get(YOL, params=DONEM, headers=admin_headers))
    satir = next(s for s in _dosya_satirlari(sheet) if s[0] == "Tower Crane TC-48")
    hucre = satir[COLUMN_HEADERS.index("Maliyet")]
    assert hucre == beklenen["cost"]
    assert isinstance(hucre, str), "para hücresi SAYI olarak yazılmış (FLOAT-YASAK)"


# --- Zarf -------------------------------------------------------------------


async def test_mime_ve_content_disposition(client, admin_headers, park):
    resp = await client.get(YOL, params=DONEM, headers=admin_headers)
    assert resp.headers["content-type"] == XLSX_MEDIA_TYPE
    disposition = resp.headers["content-disposition"]
    assert disposition.startswith("attachment; ")
    assert filename(YIL, AY) in disposition
    # İki ay aynı klasöre indiğinde birbirini EZMEZ.
    assert filename(YIL, AY) != filename(YIL, AY + 1)


async def test_zorunlu_donem_parametreleri(client, admin_headers):
    """Ekranın imzası neyse dosyanınki de odur: `year`/`month` ZORUNLUDUR."""
    assert (await client.get(YOL, headers=admin_headers)).status_code == 422
    assert (
        await client.get(YOL, params={"year": YIL, "month": 13}, headers=admin_headers)
    ).status_code == 422


async def test_okuma_denetim_kaydi_URETMEZ(client, admin_headers, park, seeded_db):
    once = (await seeded_db.execute(select(func.count()).select_from(AuditLog))).scalar_one()
    resp = await client.get(YOL, params=DONEM, headers=admin_headers)
    assert resp.status_code == 200
    sonra = (await seeded_db.execute(select(func.count()).select_from(AuditLog))).scalar_one()
    assert sonra == once
