"""Denetim gunlugu Excel disa aktarimi: GET /audit-log/export.xlsx (plan Task 5).

Sutun basliklari mockup (`Ayarlar - Denetim Gunlugu.dc.html`) tablosuyla BIREBIR.
FLOAT-YASAK: tum hucreler string yazilir (B4 dersi) — openpyxl'in sessiz tip
donusumune alan birakilmaz.
"""

import uuid
from datetime import UTC, datetime, timedelta
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.core.timezone import DISPLAY_TIMEZONE
from app.modules.audit.export import build_audit_workbook
from app.modules.audit.models import AuditAction, AuditLog
from app.modules.roles.models import Role
from app.modules.users.models import User
from tests.modules.test_audit_api import _add_row, _auth, _make_user

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_HEADER_ROW = ["Zaman", "Kullanıcı", "İşlem", "Detay", "IP Adresi"]


def _row(
    *,
    action: AuditAction = AuditAction.login,
    detail: str = "Sisteme giriş yapıldı",
    ip: str | None = "192.168.1.100",
    occurred_at: datetime | None = None,
    full_name: str | None = "Ahmet Yılmaz",
    role_name: str = "Patron",
) -> tuple[AuditLog, User | None, Role | None]:
    """Repository'nin dondurdugu (entry, actor, role) uclusunu bellekte kurar."""
    entry = AuditLog(
        id=uuid.uuid4(),
        occurred_at=occurred_at or datetime(2026, 7, 17, 9, 14, tzinfo=UTC),
        action=action,
        detail=detail,
        ip_address=ip,
    )
    if full_name is None:
        return (entry, None, None)
    actor = User(id=uuid.uuid4(), email="a@t.co", password_hash="x", full_name=full_name)
    return (entry, actor, Role(id=uuid.uuid4(), key="patron", name=role_name))


def _sheet(rows):
    buffer = build_audit_workbook(rows)
    assert isinstance(buffer, BytesIO)
    return load_workbook(buffer).active


def _values(sheet) -> list[list]:
    return [list(r) for r in sheet.iter_rows(values_only=True)]


# --- saf fonksiyon: build_audit_workbook -------------------------------------


def test_baslik_satiri_mockup_ile_birebir():
    assert _values(_sheet([_row()]))[0] == _HEADER_ROW


def test_bos_sonuc_yalnizca_baslik_satiri_olan_gecerli_dosya():
    values = _values(_sheet([]))
    assert values == [_HEADER_ROW]


def test_tum_hucreler_string_float_yasak():
    """Hicbir hucre sayi/tarih/None olarak yazilmaz — hepsi `str` (B4 dersi)."""
    rows = [
        _row(detail="Hakediş #47 onaylandı · 1240000", action=AuditAction.approve),
        _row(full_name=None, ip=None, action=AuditAction.backup, detail="2,3 GB"),
    ]
    for row in _sheet(rows).iter_rows():
        for cell in row:
            assert isinstance(cell.value, str), (cell.coordinate, type(cell.value))


@pytest.mark.parametrize(
    "action,label",
    [
        (AuditAction.login, "Giriş"),
        (AuditAction.create, "Oluşturma"),
        (AuditAction.update, "Güncelleme"),
        (AuditAction.delete, "Silme"),
        (AuditAction.approve, "Onay"),
        (AuditAction.backup, "Yedekleme"),
    ],
)
def test_islem_sutununda_turkce_rozet_etiketi(action, label):
    """Ham enum degeri (`login`, `create`...) DEGIL, mockup'in Turkce etiketi yazilir."""
    values = _values(_sheet([_row(action=action)]))
    assert values[1][2] == label
    assert values[1][2] != action.value


def test_zaman_sutunu_dd_mm_yyyy_hh_mm_stringi():
    """UTC saklanan deger TR'ye (UTC+3) cevrilerek yazilir."""
    values = _values(_sheet([_row(occurred_at=datetime(2026, 7, 17, 9, 14, tzinfo=UTC))]))
    assert values[1][0] == "17.07.2026 12:14"


def test_zaman_sutunu_utc_gece_yarisi_oncesini_ertesi_gune_tasir():
    """UTC 21:30 → Excel'de ERTESI gun 00:30 (TR=UTC+3); gun da kayar."""
    values = _values(_sheet([_row(occurred_at=datetime(2026, 7, 17, 21, 30, tzinfo=UTC))]))
    assert values[1][0] == "18.07.2026 00:30"


def test_zaman_sutunu_naive_degeri_utc_sayar():
    values = _values(_sheet([_row(occurred_at=datetime(2026, 7, 17, 21, 30))]))
    assert values[1][0] == "18.07.2026 00:30"


def test_zaman_sutunu_zoneinfo_kullanir_sabit_ofset_degil():
    """2016 kisinda TR ofseti +02 idi; sabit +3 gomulmus olsa 3 saat kayardi."""
    values = _values(_sheet([_row(occurred_at=datetime(2016, 1, 15, 12, 0, tzinfo=UTC))]))
    assert values[1][0] == "15.01.2016 14:00"


def test_kullanici_sutununda_ad_ve_rol():
    values = _values(_sheet([_row(full_name="Ayşe Demir", role_name="Muhasebe")]))
    assert "Ayşe Demir" in values[1][1]
    assert "Muhasebe" in values[1][1]


def test_aktorsuz_satirda_kullanici_sistem():
    values = _values(_sheet([_row(full_name=None)]))
    assert values[1][1] == "Sistem"


def test_ip_yoksa_tire_gosterilir():
    values = _values(_sheet([_row(ip=None)]))
    assert values[1][4] == "—"


def test_detay_ve_ip_oldugu_gibi_yazilir():
    values = _values(_sheet([_row(detail="Şirket bilgileri güncellendi", ip="10.0.0.45")]))
    assert values[1][3] == "Şirket bilgileri güncellendi"
    assert values[1][4] == "10.0.0.45"


def test_satir_sirasi_girdi_sirasini_korur():
    rows = [_row(detail="yeni"), _row(detail="orta"), _row(detail="eski")]
    assert [v[3] for v in _values(_sheet(rows))[1:]] == ["yeni", "orta", "eski"]


# --- uc: GET /audit-log/export.xlsx ------------------------------------------


async def _download(client, headers, query: str = ""):
    resp = await client.get(f"/audit-log/export.xlsx{query}", headers=headers)
    assert resp.status_code == 200, resp.text
    return resp


async def test_mime_ve_content_disposition(client, user_factory, seeded_db):
    headers = await _auth(client, user_factory, seeded_db, "system_admin")
    await _add_row(seeded_db)

    resp = await _download(client, headers)
    assert resp.headers["content-type"].startswith(_XLSX_MIME)
    assert resp.headers["content-disposition"] == 'attachment; filename="denetim-gunlugu.xlsx"'


async def test_indirilen_dosya_acilabilir_ve_baslik_satiri_dogru(client, user_factory, seeded_db):
    headers = await _auth(client, user_factory, seeded_db, "system_admin")
    actor = await _make_user(seeded_db, "Ahmet Yılmaz", "patron")
    await _add_row(seeded_db, actor=actor, detail="Sisteme giriş yapıldı")

    sheet = load_workbook(BytesIO((await _download(client, headers)).content)).active
    values = _values(sheet)
    assert values[0] == _HEADER_ROW
    assert values[1][3] == "Sisteme giriş yapıldı"
    assert "Ahmet Yılmaz" in values[1][1]


async def test_bos_sonuc_gecerli_dosya_doner(client, user_factory, seeded_db):
    headers = await _auth(client, user_factory, seeded_db, "system_admin")
    sheet = load_workbook(BytesIO((await _download(client, headers)).content)).active
    assert _values(sheet) == [_HEADER_ROW]


async def test_action_filtresi_export_a_uygulanir(client, user_factory, seeded_db):
    headers = await _auth(client, user_factory, seeded_db, "system_admin")
    await _add_row(seeded_db, action=AuditAction.create, detail="c1")
    await _add_row(seeded_db, action=AuditAction.delete, detail="d1")

    tumu = _values(load_workbook(BytesIO((await _download(client, headers)).content)).active)
    filtreli = _values(
        load_workbook(BytesIO((await _download(client, headers, "?action=create")).content)).active
    )
    assert len(tumu) == 3  # baslik + 2 satir
    assert len(filtreli) == 2  # baslik + 1 satir — filtreli export != tum kayitlar
    assert filtreli[1][3] == "c1"


async def test_q_filtresi_export_a_uygulanir(client, user_factory, seeded_db):
    headers = await _auth(client, user_factory, seeded_db, "system_admin")
    await _add_row(seeded_db, detail="Kullanıcı oluşturuldu: Ayşe Demir")
    await _add_row(seeded_db, detail="Şirket bilgileri güncellendi")

    values = _values(
        load_workbook(BytesIO((await _download(client, headers, "?q=oluşturuldu")).content)).active
    )
    assert len(values) == 2
    assert values[1][3].startswith("Kullanıcı oluşturuldu")


async def test_tarih_araligi_export_a_uygulanir(client, user_factory, seeded_db):
    headers = await _auth(client, user_factory, seeded_db, "system_admin")
    now = datetime.now(UTC)
    await _add_row(seeded_db, detail="bugun", occurred_at=now)
    await _add_row(seeded_db, detail="eski", occurred_at=now - timedelta(days=15))

    # Gun siniri TR takvimine gore; "bugun" TR gunudur.
    d_today = datetime.now(DISPLAY_TIMEZONE).date().isoformat()
    values = _values(
        load_workbook(
            BytesIO((await _download(client, headers, f"?date_from={d_today}")).content)
        ).active
    )
    assert [v[3] for v in values[1:]] == ["bugun"]


async def test_actor_filtresi_export_a_uygulanir(client, user_factory, seeded_db):
    headers = await _auth(client, user_factory, seeded_db, "system_admin")
    a = await _make_user(seeded_db, "Ahmet Yılmaz", "patron")
    b = await _make_user(seeded_db, "Sercan Öztürk", "site_chief")
    await _add_row(seeded_db, actor=a, detail="a1")
    await _add_row(seeded_db, actor=b, detail="b1")

    values = _values(
        load_workbook(
            BytesIO((await _download(client, headers, f"?actor_user_id={a.id}")).content)
        ).active
    )
    assert [v[3] for v in values[1:]] == ["a1"]


async def test_limit_yok_200_den_fazla_kayit_da_doner(client, user_factory, seeded_db):
    """Okuma ucunun `limit<=200` siniri export'a uygulanmaz — sessiz kirpma YOK."""
    headers = await _auth(client, user_factory, seeded_db, "system_admin")
    now = datetime.now(UTC)
    for i in range(205):
        await _add_row(seeded_db, detail=f"r{i}", occurred_at=now - timedelta(seconds=i))

    values = _values(load_workbook(BytesIO((await _download(client, headers)).content)).active)
    assert len(values) == 206  # baslik + 205 satir


async def test_aktorsuz_satir_sistem_olarak_disa_aktarilir(client, user_factory, seeded_db):
    headers = await _auth(client, user_factory, seeded_db, "system_admin")
    await _add_row(seeded_db, action=AuditAction.backup, detail="Yedekleme", actor=None, ip=None)

    values = _values(load_workbook(BytesIO((await _download(client, headers)).content)).active)
    assert values[1][1] == "Sistem"
    assert values[1][4] == "—"
    assert values[1][2] == "Yedekleme"


# --- yetki -------------------------------------------------------------------


async def test_yetkisiz_rol_403(client, user_factory, seeded_db):
    """`settings < view` olan HER rol reddedilir — patron dahil."""
    headers = await _auth(client, user_factory, seeded_db, "patron")
    assert (await client.get("/audit-log/export.xlsx", headers=headers)).status_code == 403


@pytest.mark.parametrize("role_key", ["site_chief", "accounting", "hr_manager", "procurement"])
async def test_diger_roller_de_403(client, user_factory, seeded_db, role_key):
    headers = await _auth(client, user_factory, seeded_db, role_key)
    assert (await client.get("/audit-log/export.xlsx", headers=headers)).status_code == 403


async def test_kimliksiz_istek_401(client, seeded_db):
    assert (await client.get("/audit-log/export.xlsx")).status_code == 401


async def test_gecersiz_action_422(client, user_factory, seeded_db):
    headers = await _auth(client, user_factory, seeded_db, "system_admin")
    resp = await client.get("/audit-log/export.xlsx?action=hacimsiz", headers=headers)
    assert resp.status_code == 422
