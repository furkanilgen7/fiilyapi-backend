"""EXPORT-XLSX — muhasebenin ÜÇ Excel ucu (Mizan · Hesap Planı · Yevmiye).

`GET /trial-balance/export.xlsx` · `GET /chart-of-accounts/export.xlsx` ·
`GET /journal/export.xlsx`

Dosyanın tamamına DÖRT kural hâkimdir:

1. **İçerik ÖLÇÜLÜR.** 200 + bayt uzunluğu HİÇBİR ŞEY kanıtlamaz: üretilen
   dosya `openpyxl` ile geri okunur ve hücreler tek tek doğrulanır.
2. **KÜME EŞİTLİĞİ, sayım değil.** Ekranın liste ucu ile dosya AYNI senaryoda
   çağrılır ve satır KÜMELERİ karşılaştırılır. Yalnız uzunluk karşılaştırılsaydı
   süzgeci farklı uygulayan bir dosya (veri kaçağı) testi GEÇERDİ — kusur tam
   olarak buydu ve bir kez üretimde yaşandı.
3. **Yetki kapısında POZİTİF KONTROL.** Yalnız 403 ölçülseydi, herkese 403
   veren bozuk bir uç da yeşil kalırdı. Her kapı testi yetkili aktörün dosyayı
   GERÇEKTEN aldığını da ölçer.
4. **Kırpma yok.** Sayfalama tavanını (200) AŞAN kümeler kurulur; ekran hâlâ
   kırpar, dosya kırpmaz. Bu, `limit=None`ın SQL'e gerçekten ulaştığının tek
   kanıtıdır.

`openpyxl`in kendi davranışı TEST EDİLMEZ.
"""

import uuid
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.modules.accounting.models import ChartAccountType, JournalEntryStatus

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_MIZAN = "/trial-balance/export.xlsx"
_MIZAN_LISTE = "/trial-balance"
_HP = "/chart-of-accounts/export.xlsx"
_HP_LISTE = "/chart-of-accounts"
_DEFTER = "/journal/export.xlsx"
_DEFTER_LISTE = "/journal"

#: Mockup `Muhasebe - Mizan.dc.html` satır 84-99.
_MIZAN_BASLIKLARI = (
    "Hesap Kodu",
    "Hesap Adı",
    "Açılış Borç",
    "Açılış Alacak",
    "Dönem Borç",
    "Dönem Alacak",
    "Kapanış Borç",
    "Kapanış Alacak",
)
#: Şerit (2) + boş ayraç (1) → başlık 4. satırdadır.
_MIZAN_BASLIK_SATIRI = 4

#: Mockup `Muhasebe - Hesap Planı.dc.html` satır 59-63.
_HP_BASLIKLARI = ("Kod", "Hesap Adı", "Tür", "Bakiye", "Durum")

#: Mockup `Ekran 8 - Muhasebe.dc.html` satır 101-106.
_DEFTER_BASLIKLARI = ("Tarih", "Hesap Kodu", "Açıklama", "Borç", "Alacak", "Bakiye")

_YIL = 2026
_AY = 7


def _sheet(content: bytes):
    return load_workbook(BytesIO(content)).active


def _rows(content: bytes) -> list[tuple]:
    return [tuple(row) for row in _sheet(content).iter_rows(values_only=True)]


def _dosya_mi(resp) -> None:
    """Yanıtın GERÇEKTEN açılabilir bir xlsx olduğunu ölçer (pozitif kontrol)."""
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(_XLSX_MIME)
    assert "attachment;" in resp.headers["content-disposition"]
    assert _sheet(resp.content) is not None


async def _hesaplar(hesap_fabrikasi):
    kasa = await hesap_fabrikasi("100", name="Kasa", account_type=ChartAccountType.asset)
    saticilar = await hesap_fabrikasi(
        "320", name="Satıcılar", account_type=ChartAccountType.liability
    )
    return kasa, saticilar


# =========================================================================== #
# 1 — MİZAN
# =========================================================================== #


def _mizan_json_kumesi(govde: dict) -> set[tuple]:
    return {
        (
            satir["account_code"],
            satir["account_name"],
            satir["opening_debit"],
            satir["opening_credit"],
            satir["period_debit"],
            satir["period_credit"],
            satir["closing_debit"],
            satir["closing_credit"],
        )
        for satir in govde["rows"]
    }


def _mizan_dosya_kumesi(content: bytes) -> set[tuple]:
    satirlar = _rows(content)
    assert satirlar[_MIZAN_BASLIK_SATIRI - 1] == _MIZAN_BASLIKLARI
    # Başlıktan sonrası; SON satır `GENEL TOPLAM`dır ve kümeye girmez.
    return {tuple(satir) for satir in satirlar[_MIZAN_BASLIK_SATIRI:-1]}


async def test_mizan_dosyasi_ekranin_kumesiyle_AYNI(
    client, muhasebe_headers, hesap_fabrikasi, fis_fabrikasi
):
    """🔴 KÜME EŞİTLİĞİ — dosyanın satır kümesi ekranın satır kümesiyle AYNI.

    MUTASYON: uç `include_empty=True` ile çağrılsa (ya da başka bir süzgeç
    uygulasa) kümeler AYRIŞIR ve bu test düşer.
    """
    kasa, saticilar = await _hesaplar(hesap_fabrikasi)
    await fis_fabrikasi([(kasa, "1234.56", "0"), (saticilar, "0", "1234.56")])

    sorgu = f"year={_YIL}&month={_AY}"
    ekran = await client.get(f"{_MIZAN_LISTE}?{sorgu}", headers=muhasebe_headers)
    dosya = await client.get(f"{_MIZAN}?{sorgu}", headers=muhasebe_headers)

    assert ekran.status_code == 200, ekran.text
    _dosya_mi(dosya)
    assert _mizan_json_kumesi(ekran.json()) == _mizan_dosya_kumesi(dosya.content)
    assert len(ekran.json()["rows"]) == 2


async def test_mizan_include_empty_suzgeci_DOSYAYA_da_gecer(
    client, muhasebe_headers, hesap_fabrikasi, fis_fabrikasi
):
    """Hareketsiz hesap varsayılanda dosyada YOKTUR, `include_empty=true` ile VARDIR.

    MUTASYON: uç `include_empty` parametresini yok sayıp sabit bir değer
    geçirseydi iki dosyadan biri yanlış kümeyi taşırdı.
    """
    kasa, saticilar = await _hesaplar(hesap_fabrikasi)
    await hesap_fabrikasi("102", name="Bankalar")
    await fis_fabrikasi([(kasa, "500.00", "0"), (saticilar, "0", "500.00")])

    varsayilan = await client.get(f"{_MIZAN}?year={_YIL}&month={_AY}", headers=muhasebe_headers)
    hepsi = await client.get(
        f"{_MIZAN}?year={_YIL}&month={_AY}&include_empty=true", headers=muhasebe_headers
    )

    kodlar = lambda icerik: {satir[0] for satir in _mizan_dosya_kumesi(icerik)}  # noqa: E731
    assert "102" not in kodlar(varsayilan.content)
    assert "102" in kodlar(hepsi.content)


async def test_mizan_genel_toplam_satiri_zarfin_totals_ALANINDAN_gelir(
    client, muhasebe_headers, hesap_fabrikasi, fis_fabrikasi
):
    """tfoot YENİDEN TOPLANMAZ: hücreler `totals`ın metniyle BİREBİR.

    MUTASYON: `GENEL TOPLAM` satırları `rows` üzerinden yeniden toplasaydı
    (ör. yuvarlayarak ya da açılışı atlayarak) bu eşitlik bozulurdu.
    """
    kasa, saticilar = await _hesaplar(hesap_fabrikasi)
    await fis_fabrikasi([(kasa, "1234.56", "0"), (saticilar, "0", "1234.56")])

    sorgu = f"year={_YIL}&month={_AY}"
    ekran = (await client.get(f"{_MIZAN_LISTE}?{sorgu}", headers=muhasebe_headers)).json()
    dosya = await client.get(f"{_MIZAN}?{sorgu}", headers=muhasebe_headers)

    toplam_satiri = _rows(dosya.content)[-1]
    totals = ekran["totals"]
    assert toplam_satiri == (
        "GENEL TOPLAM",
        None,
        totals["opening_debit"],
        totals["opening_credit"],
        totals["period_debit"],
        totals["period_credit"],
        totals["closing_debit"],
        totals["closing_credit"],
    )


async def test_mizan_para_hucresi_JSON_metniyle_BAYT_BAYT_ayni(
    client, muhasebe_headers, hesap_fabrikasi, fis_fabrikasi
):
    """`str(Decimal)` kuralı: kuruş ölçeği korunur, yeniden yuvarlama YOK.

    MUTASYON: hücreye `Decimal`/`float` yazılsa ya da biçimlendirilse
    (`f"{x:,.2f}"`) metin `1234.56`dan ayrılırdı.
    """
    kasa, saticilar = await _hesaplar(hesap_fabrikasi)
    await fis_fabrikasi([(kasa, "1234.56", "0"), (saticilar, "0", "1234.56")])

    dosya = await client.get(f"{_MIZAN}?year={_YIL}&month={_AY}", headers=muhasebe_headers)

    kasa_satiri = next(s for s in _mizan_dosya_kumesi(dosya.content) if s[0] == "100")
    assert kasa_satiri[4] == "1234.56"
    assert all(isinstance(hucre, str) for hucre in kasa_satiri)


async def test_mizan_bant_donem_ve_denge_durumunu_TASIR(
    client, muhasebe_headers, hesap_fabrikasi, fis_fabrikasi
):
    kasa, saticilar = await _hesaplar(hesap_fabrikasi)
    await fis_fabrikasi([(kasa, "500.00", "0"), (saticilar, "0", "500.00")])

    dosya = await client.get(f"{_MIZAN}?year={_YIL}&month={_AY}", headers=muhasebe_headers)

    satirlar = _rows(dosya.content)
    assert satirlar[0][:2] == ("Dönem", f"01.{_YIL}–{_AY:02d}.{_YIL}")
    assert satirlar[1][:2] == ("Denge", "Mizan Dengede")


async def test_mizan_DENGESIZ_defterde_bant_dengesizligi_soyler(
    client, muhasebe_headers, hesap_fabrikasi, fis_fabrikasi
):
    """`is_balanced` zarftan gelir; dosya onu SESSİZCE "dengede" yazmaz."""
    kasa, _ = await _hesaplar(hesap_fabrikasi)
    await fis_fabrikasi([(kasa, "500.00", "0")], header_totals=("500.00", "500.00"))

    dosya = await client.get(f"{_MIZAN}?year={_YIL}&month={_AY}", headers=muhasebe_headers)

    assert _rows(dosya.content)[1][:2] == ("Denge", "Mizan Dengede Değil")


async def test_mizan_BOS_kume_gecerli_dosya_uretir(client, muhasebe_headers):
    dosya = await client.get(f"{_MIZAN}?year={_YIL}&month={_AY}", headers=muhasebe_headers)

    _dosya_mi(dosya)
    satirlar = _rows(dosya.content)
    assert satirlar[_MIZAN_BASLIK_SATIRI - 1] == _MIZAN_BASLIKLARI
    assert satirlar[-1][0] == "GENEL TOPLAM"
    assert _mizan_dosya_kumesi(dosya.content) == set()


async def test_mizan_kapisi_yetkisiz_403_yetkili_DOSYAYI_ALIR(
    client, yetkisiz_headers, muhasebe_headers, hesap_fabrikasi, fis_fabrikasi
):
    """🔴 POZİTİF KONTROLLÜ kapı: `accounting=none` 403, `accounting=full` DOSYA.

    MUTASYON: `dependencies=[_VIEW]` silinseydi ilk assert düşer; kapı herkese
    403 verecek biçimde bozulsaydı ikinci assert düşer.
    """
    kasa, saticilar = await _hesaplar(hesap_fabrikasi)
    await fis_fabrikasi([(kasa, "500.00", "0"), (saticilar, "0", "500.00")])
    sorgu = f"year={_YIL}&month={_AY}"

    assert (await client.get(f"{_MIZAN}?{sorgu}", headers=yetkisiz_headers)).status_code == 403
    _dosya_mi(await client.get(f"{_MIZAN}?{sorgu}", headers=muhasebe_headers))


async def test_mizan_kimliksiz_401(client):
    assert (await client.get(f"{_MIZAN}?year={_YIL}&month={_AY}")).status_code == 401


async def test_mizan_donem_ZORUNLU_ve_bant_liste_ucuyle_AYNI(client, muhasebe_headers):
    """Liste ucuyla aynı zorunluluk/bantlar — sunucunun "bugün"ü okunmaz."""
    assert (await client.get(_MIZAN, headers=muhasebe_headers)).status_code == 422
    assert (
        await client.get(f"{_MIZAN}?year={_YIL}&month=13", headers=muhasebe_headers)
    ).status_code == 422


# =========================================================================== #
# 2 — HESAP PLANI
# =========================================================================== #


def _hp_json_kumesi(govde: dict) -> set[tuple]:
    turler = {
        "asset": "Aktif",
        "liability": "Pasif",
        "revenue": "Gelir",
        "expense": "Gider",
        "equity": "Özkaynak",
    }
    return {
        (
            satir["code"],
            satir["name"],
            turler[satir["account_type"]],
            satir["balance"],
            "Kullanımda" if satir["is_active"] else "Kullanım Dışı",
        )
        for satir in govde["items"]
    }


def _hp_dosya_kumesi(content: bytes) -> set[tuple]:
    satirlar = _rows(content)
    assert satirlar[0] == _HP_BASLIKLARI
    return {tuple(satir) for satir in satirlar[1:]}


async def test_hesap_plani_dosyasi_ekranin_kumesiyle_AYNI(
    client, muhasebe_headers, hesap_fabrikasi, fis_fabrikasi
):
    """🔴 KÜME EŞİTLİĞİ — beş kolonun HEPSİ ekranın zarfıyla aynı."""
    kasa, saticilar = await _hesaplar(hesap_fabrikasi)
    await hesap_fabrikasi("600", name="Yurt İçi Satışlar", account_type=ChartAccountType.revenue)
    await hesap_fabrikasi("103", name="Kapalı Hesap", is_active=False)
    await fis_fabrikasi([(kasa, "1234.56", "0"), (saticilar, "0", "1234.56")])

    ekran = await client.get(f"{_HP_LISTE}?limit=200", headers=muhasebe_headers)
    dosya = await client.get(_HP, headers=muhasebe_headers)

    assert ekran.status_code == 200, ekran.text
    _dosya_mi(dosya)
    assert _hp_json_kumesi(ekran.json()) == _hp_dosya_kumesi(dosya.content)
    assert len(ekran.json()["items"]) == 4


async def test_hesap_plani_suzgecleri_DOSYAYA_da_gecer(client, muhasebe_headers, hesap_fabrikasi):
    """🔴 VERİ KAÇAĞI BEKÇİSİ: `q`/`account_type`/`is_active` dosyada da uygulanır.

    MUTASYON: uç süzgeçleri servise geçirmeseydi (ya da farklı uygulasaydı)
    süzgeç dışı satırlar dosyaya sızardı ve sayım testi bunu GÖREMEZDİ.
    """
    await hesap_fabrikasi("100", name="Kasa")
    await hesap_fabrikasi("320", name="Satıcılar", account_type=ChartAccountType.liability)
    await hesap_fabrikasi("103", name="Kapalı Kasa", is_active=False)

    async def kodlar(sorgu: str) -> set[str]:
        resp = await client.get(f"{_HP}?{sorgu}", headers=muhasebe_headers)
        _dosya_mi(resp)
        return {satir[0] for satir in _hp_dosya_kumesi(resp.content)}

    assert await kodlar("q=Kasa") == {"100", "103"}
    assert await kodlar("account_type=liability") == {"320"}
    assert await kodlar("is_active=false") == {"103"}
    assert await kodlar("q=Kasa&is_active=true") == {"100"}


async def test_hesap_plani_dosyasi_SAYFALAMA_TAVANINI_asar(
    client, muhasebe_headers, hesap_fabrikasi
):
    """🔴 KIRPMA YOK — ekran 200'de kırpar, dosya 205'in HEPSİNİ yazar.

    MUTASYON: uç `limit=None` yerine `limit=200` geçirseydi (ya da repository
    `if limit is not None` koşulunu kaybetseydi) dosya 200 satırda kalır ve bu
    test düşerdi. Karşılaştırma SAYIM DEĞİL küme eşitliğidir: ekranın iki
    sayfasının birleşimi ile dosyanın kümesi aynı olmalıdır.
    """
    beklenen = set()
    for i in range(205):
        kod = f"{700 + i // 99}.{i % 99 + 1:02d}"
        await hesap_fabrikasi(kod, name=f"Hesap {kod}", account_type=ChartAccountType.expense)
        beklenen.add(kod)

    sayfa1 = await client.get(f"{_HP_LISTE}?limit=200&offset=0", headers=muhasebe_headers)
    sayfa2 = await client.get(f"{_HP_LISTE}?limit=200&offset=200", headers=muhasebe_headers)
    dosya = await client.get(_HP, headers=muhasebe_headers)

    assert sayfa1.json()["total"] == 205
    assert len(sayfa1.json()["items"]) == 200  # ekran KIRPAR
    assert len(sayfa2.json()["items"]) == 5

    dosya_kodlari = {satir[0] for satir in _hp_dosya_kumesi(dosya.content)}
    ekran_kodlari = {s["code"] for s in sayfa1.json()["items"] + sayfa2.json()["items"]}
    assert dosya_kodlari == beklenen == ekran_kodlari


async def test_hesap_plani_liste_ucunun_TAVANI_DEGISMEDI(client, muhasebe_headers):
    """Servis genişledi ama LİSTE ucunun sözleşmesi aynı: 201 → **422**."""
    resp = await client.get(f"{_HP_LISTE}?limit=201", headers=muhasebe_headers)
    assert resp.status_code == 422


async def test_hesap_plani_para_hucresi_JSON_metniyle_BAYT_BAYT_ayni(
    client, muhasebe_headers, hesap_fabrikasi, fis_fabrikasi
):
    kasa, saticilar = await _hesaplar(hesap_fabrikasi)
    await fis_fabrikasi([(kasa, "1234.56", "0"), (saticilar, "0", "1234.56")])

    ekran = await client.get(f"{_HP_LISTE}?limit=200", headers=muhasebe_headers)
    dosya = await client.get(_HP, headers=muhasebe_headers)

    json_bakiye = next(s["balance"] for s in ekran.json()["items"] if s["code"] == "100")
    dosya_bakiye = next(s[3] for s in _hp_dosya_kumesi(dosya.content) if s[0] == "100")
    assert dosya_bakiye == json_bakiye == "1234.56"


async def test_hesap_plani_BOS_kume_gecerli_dosya_uretir(client, muhasebe_headers):
    dosya = await client.get(_HP, headers=muhasebe_headers)

    _dosya_mi(dosya)
    assert _rows(dosya.content) == [_HP_BASLIKLARI]


async def test_hesap_plani_kapisi_yetkisiz_403_yetkili_DOSYAYI_ALIR(
    client, yetkisiz_headers, pm_headers, hesap_fabrikasi
):
    """`accounting=none` 403; `accounting=view` (PM) dosyayı ALIR — `view` YETER."""
    await hesap_fabrikasi("100", name="Kasa")

    assert (await client.get(_HP, headers=yetkisiz_headers)).status_code == 403
    _dosya_mi(await client.get(_HP, headers=pm_headers))


async def test_hesap_plani_kimliksiz_401(client):
    assert (await client.get(_HP)).status_code == 401


async def test_hesap_plani_export_yolu_UUID_SANILMAZ(client, muhasebe_headers):
    """🔴 ROTA SIRASI (MK-2): literal yol UUID rotasının ÜSTÜNDE kayıtlı.

    MUTASYON: uç `/chart-of-accounts/{account_id}` altına taşınsaydı `export.xlsx`
    bir UUID sanılır ve **422** dönerdi.
    """
    resp = await client.get(_HP, headers=muhasebe_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(_XLSX_MIME)


# =========================================================================== #
# 3 — YEVMİYE DEFTERİ
# =========================================================================== #


def _defter_json_kumesi(govde: dict) -> set[tuple]:
    return {
        (
            satir["entry_date"],
            satir["account_code"],
            satir["description"]
            if satir["detail_note"] is None
            else f"{satir['description']}\n{satir['detail_note']}",
            satir["debit"],
            satir["credit"],
            satir["running_balance"],
        )
        for satir in govde["items"]
    }


def _defter_dosya_kumesi(content: bytes) -> set[tuple]:
    satirlar = _rows(content)
    assert satirlar[0] == _DEFTER_BASLIKLARI
    return {tuple(satir) for satir in satirlar[1:]}


async def test_defter_dosyasi_ekranin_kumesiyle_AYNI(
    client, muhasebe_headers, hesap_fabrikasi, fis_fabrikasi
):
    """🔴 KÜME EŞİTLİĞİ — altı kolonun HEPSİ (koşan bakiye dahil) ekranla aynı."""
    kasa, saticilar = await _hesaplar(hesap_fabrikasi)
    await fis_fabrikasi([(kasa, "1234.56", "0"), (saticilar, "0", "1234.56")])
    await fis_fabrikasi([(kasa, "0", "34.56"), (saticilar, "34.56", "0")])
    # `draft` deftere GİRMEZ — dosya da onu göstermemeli (küme eşitliği ölçer).
    await fis_fabrikasi(
        [(kasa, "999.00", "0"), (saticilar, "0", "999.00")], status=JournalEntryStatus.draft
    )

    sorgu = f"year={_YIL}&month={_AY}"
    ekran = await client.get(f"{_DEFTER_LISTE}?{sorgu}&limit=200", headers=muhasebe_headers)
    dosya = await client.get(f"{_DEFTER}?{sorgu}", headers=muhasebe_headers)

    assert ekran.status_code == 200, ekran.text
    _dosya_mi(dosya)
    assert _defter_json_kumesi(ekran.json()) == _defter_dosya_kumesi(dosya.content)
    assert len(ekran.json()["items"]) == 4


async def test_defter_suzgecleri_DOSYAYA_da_gecer(
    client, muhasebe_headers, hesap_fabrikasi, fis_fabrikasi
):
    """🔴 VERİ KAÇAĞI BEKÇİSİ: `account_id` ve `status` dosyada da uygulanır."""
    kasa, saticilar = await _hesaplar(hesap_fabrikasi)
    fis = await fis_fabrikasi([(kasa, "500.00", "0"), (saticilar, "0", "500.00")])
    await fis_fabrikasi(
        [(kasa, "10.00", "0"), (saticilar, "0", "10.00")], status=JournalEntryStatus.reversed
    )
    assert fis is not None

    sorgu = f"year={_YIL}&month={_AY}"

    async def kodlar(ek: str) -> list[str]:
        resp = await client.get(f"{_DEFTER}?{sorgu}{ek}", headers=muhasebe_headers)
        _dosya_mi(resp)
        return [satir[1] for satir in _defter_dosya_kumesi(resp.content)]

    assert set(await kodlar(f"&account_id={kasa.id}")) == {"100"}
    assert len(await kodlar("&status=reversed")) == 2
    assert len(await kodlar("")) == 4


async def test_defter_dosyasi_SAYFALAMA_TAVANINI_asar_ve_bakiye_BOZULMAZ(
    client, muhasebe_headers, hesap_fabrikasi, fis_fabrikasi
):
    """🔴 KIRPMA YOK + koşan bakiye KORUNUR (206 satır, tavan 200).

    İki şey birden ölçülür:
    * dosya 206 satırın HEPSİNİ taşır (ekran 200'de kırpar);
    * ekranın İKİ sayfasının birleşimi ile dosyanın kümesi — `running_balance`
      DAHİL — AYNIDIR. Pencere fonksiyonu `LIMIT` ile birleştirilseydi
      (`ledger.py` 3. tuzak) sınırsız koşuda bakiyeler kayar ve bu test düşerdi.

    MUTASYON: uç `limit=None` yerine `limit=200` geçirse ya da `ledger.py`nin
    `if limit is not None` koşulu kaybolsa test düşer.
    """
    kasa, saticilar = await _hesaplar(hesap_fabrikasi)
    satirlar = []
    for i in range(103):
        satirlar.append((kasa, f"{i + 1}.00", "0"))
        satirlar.append((saticilar, "0", f"{i + 1}.00"))
    await fis_fabrikasi(satirlar)

    sorgu = f"year={_YIL}&month={_AY}"
    sayfa1 = await client.get(
        f"{_DEFTER_LISTE}?{sorgu}&limit=200&offset=0", headers=muhasebe_headers
    )
    sayfa2 = await client.get(
        f"{_DEFTER_LISTE}?{sorgu}&limit=200&offset=200", headers=muhasebe_headers
    )
    dosya = await client.get(f"{_DEFTER}?{sorgu}", headers=muhasebe_headers)

    assert sayfa1.json()["total"] == 206
    assert len(sayfa1.json()["items"]) == 200  # ekran KIRPAR
    assert len(sayfa2.json()["items"]) == 6

    dosya_kumesi = _defter_dosya_kumesi(dosya.content)
    ekran_kumesi = _defter_json_kumesi({"items": sayfa1.json()["items"] + sayfa2.json()["items"]})
    assert len(_rows(dosya.content)) == 207  # başlık + 206 satır
    assert dosya_kumesi == ekran_kumesi


async def test_defter_liste_ucunun_TAVANI_DEGISMEDI(client, muhasebe_headers):
    resp = await client.get(f"{_DEFTER_LISTE}?limit=201", headers=muhasebe_headers)
    assert resp.status_code == 422


async def test_defter_para_hucresi_JSON_metniyle_BAYT_BAYT_ayni(
    client, muhasebe_headers, hesap_fabrikasi, fis_fabrikasi
):
    kasa, saticilar = await _hesaplar(hesap_fabrikasi)
    await fis_fabrikasi([(kasa, "1234.56", "0"), (saticilar, "0", "1234.56")])

    sorgu = f"year={_YIL}&month={_AY}"
    ekran = await client.get(f"{_DEFTER_LISTE}?{sorgu}&limit=200", headers=muhasebe_headers)
    dosya = await client.get(f"{_DEFTER}?{sorgu}", headers=muhasebe_headers)

    json_satiri = next(s for s in ekran.json()["items"] if s["account_code"] == "100")
    dosya_satiri = next(s for s in _defter_dosya_kumesi(dosya.content) if s[1] == "100")
    assert dosya_satiri[3] == json_satiri["debit"] == "1234.56"
    assert dosya_satiri[5] == json_satiri["running_balance"]


async def test_defter_BOS_ay_gecerli_dosya_uretir(client, muhasebe_headers):
    dosya = await client.get(f"{_DEFTER}?year={_YIL}&month=1", headers=muhasebe_headers)

    _dosya_mi(dosya)
    assert _rows(dosya.content) == [_DEFTER_BASLIKLARI]


async def test_defter_kapisi_yetkisiz_403_yetkili_DOSYAYI_ALIR(
    client, yetkisiz_headers, muhasebe_headers, hesap_fabrikasi, fis_fabrikasi
):
    kasa, saticilar = await _hesaplar(hesap_fabrikasi)
    await fis_fabrikasi([(kasa, "500.00", "0"), (saticilar, "0", "500.00")])
    sorgu = f"year={_YIL}&month={_AY}"

    assert (await client.get(f"{_DEFTER}?{sorgu}", headers=yetkisiz_headers)).status_code == 403
    _dosya_mi(await client.get(f"{_DEFTER}?{sorgu}", headers=muhasebe_headers))


async def test_defter_kimliksiz_401(client):
    assert (await client.get(f"{_DEFTER}?year={_YIL}&month={_AY}")).status_code == 401


async def test_defter_donemi_verilmezse_liste_ucuyle_AYNI_VARSAYILANI_kullanir(
    client, muhasebe_headers
):
    """Dönem verilmediğinde `ledger.default_period()` — dosya adı da onu taşır."""
    from app.modules.accounting import ledger

    yil, ay = ledger.default_period()
    resp = await client.get(_DEFTER, headers=muhasebe_headers)

    _dosya_mi(resp)
    assert f"yevmiye-{yil}-{ay:02d}.xlsx" in resp.headers["content-disposition"]


async def test_defter_aciklama_ve_detay_notu_TEK_hucrede_IKI_SATIR(
    client, muhasebe_headers, hesap_fabrikasi, fis_fabrikasi, seeded_db
):
    """E8:113 iki satırlıdır; not `null` iken ikinci satır AÇILMAZ."""
    from sqlalchemy import select

    from app.modules.accounting.models import JournalLine

    kasa, saticilar = await _hesaplar(hesap_fabrikasi)
    fis = await fis_fabrikasi(
        [(kasa, "500.00", "0"), (saticilar, "0", "500.00")], description="Hakediş Tahsilatı"
    )
    satir = (
        await seeded_db.execute(select(JournalLine).where(JournalLine.entry_id == fis.id).limit(1))
    ).scalar_one()
    assert satir is not None
    fis.detail_note = "Ziraat Bank TRF-0717"
    await seeded_db.flush()

    dosya = await client.get(f"{_DEFTER}?year={_YIL}&month={_AY}", headers=muhasebe_headers)

    aciklamalar = {satir[2] for satir in _defter_dosya_kumesi(dosya.content)}
    assert aciklamalar == {"Hakediş Tahsilatı\nZiraat Bank TRF-0717"}


# =========================================================================== #
# ORTAK — dosya adları çakışmaz
# =========================================================================== #


async def test_dosya_adlari_KAPSAMI_tasir(client, muhasebe_headers):
    """İki dönem aynı klasöre indiğinde birbirini EZMEMELİ."""
    mizan = await client.get(f"{_MIZAN}?year={_YIL}&month={_AY}", headers=muhasebe_headers)
    defter = await client.get(f"{_DEFTER}?year={_YIL}&month={_AY}", headers=muhasebe_headers)
    hp = await client.get(_HP, headers=muhasebe_headers)

    assert "mizan-2026-07.xlsx" in mizan.headers["content-disposition"]
    assert "yevmiye-2026-07.xlsx" in defter.headers["content-disposition"]
    assert "hesap-plani.xlsx" in hp.headers["content-disposition"]


async def test_uuid_tipli_account_id_suzgeci_gecersizse_422(client, muhasebe_headers):
    resp = await client.get(
        f"{_DEFTER}?year={_YIL}&month={_AY}&account_id={uuid.uuid4()}x", headers=muhasebe_headers
    )
    assert resp.status_code == 422


def test_decimal_metni_kuruş_olcegini_KORUR():
    """Kural sabiti: `str(Decimal)` yeniden yuvarlamaz (hücre yazımının tabanı)."""
    assert str(Decimal("1234.50")) == "1234.50"
