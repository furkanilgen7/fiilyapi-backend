"""P9 T4 — paylasim tablosu Excel disa aktarimi (spec §5, S3; KKP 24 "Excel").

`GET /projects/{project_id}/units/export.xlsx` — KKP 86-92'nin YEDI sutunu.

Uc kural bu dosyanin tamamina hakimdir:

1. **Icerik OLCULUR.** 200 + bayt uzunlugu bir sey KANITLAMAZ: uretilen dosya
   `openpyxl` ile GERI OKUNUR ve basliklar + hucre degerleri tek tek dogrulanir
   (`test_boq_export.py` / `test_export.py` (puantaj) emsali).
2. **Kaynak LISTE UCUYLA AYNI.** Sutun degerleri `service.list_units` zarfindan
   gelir; Excel ikinci bir hesap yolu ACMAZ. Bu yuzden ayni senaryoda liste
   ucunun JSON'u ile dosyanin hucreleri KARSILASTIRILIR — ikisi ayrisirsa
   hangisinin dogru oldugu tartisma konusu olurdu.
3. **Gorunmeyen proje 404.** Dosya ucu bir PROJE VARLIK ORAKULU olamaz
   (`units_import_template_endpoint` ile ayni gerekce, IDOR-8).
"""

import uuid
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import update

from app.modules.customers.models import Customer, CustomerType
from app.modules.projects.models import LandShareShareholder
from app.modules.sales.models import SaleType, UnitSale, UnitSaleStatus
from app.modules.units.models import Unit, UnitOwnerSide, UnitSalesStatus
from tests.modules.units._units_api import _auth, _block, _login, _site, _unit

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

#: KKP 86-92 — sira ve metin mockup'tan; degistirilmesi ekranla ayrisma demektir.
_HEADERS = ("Ünite", "Tip", "m²", "Rayiç Değer", "Sahip", "Hissedar / Alıcı", "Satış Durumu")

_PROJECT_MISSING = "Proje bulunamadı"


def _url(project_id: uuid.UUID) -> str:
    return f"/projects/{project_id}/units/export.xlsx"


def _sheet(content: bytes):
    return load_workbook(BytesIO(content)).active


def _rows(sheet) -> list[tuple]:
    return [tuple(row) for row in sheet.iter_rows(values_only=True)]


async def _shareholder(session, project, name: str = "Ahmet Yılmaz") -> LandShareShareholder:
    row = LandShareShareholder(project_id=project.id, name=name, share_pct=Decimal("50.00"))
    session.add(row)
    await session.flush()
    return row


async def _sale(session, project, unit, actor_id: uuid.UUID, name: str = "Serkan Öz") -> UnitSale:
    """Alici adi ACIK satis kaydindan gelir (P8 T5) — Excel de oradan okur."""
    customer = Customer(name=name, customer_type=CustomerType.person)
    session.add(customer)
    await session.flush()
    sale = UnitSale(
        unit_id=unit.id,
        project_id=project.id,
        customer_id=customer.id,
        sale_type=SaleType.sale,
        status=UnitSaleStatus.active,
        sale_price=Decimal("1380000.00"),
        created_by=actor_id,
    )
    session.add(sale)
    await session.flush()
    return sale


# --- Yetki + gorunurluk ---


async def test_export_requires_token(client, project_factory):
    project = await project_factory("P9-EX-1")

    resp = await client.get(_url(project.id))

    assert resp.status_code == 401


async def test_export_requires_projects_permission(client, user_factory, project_factory):
    """`procurement` rolunun `projects` izni `none` — dosya ucu de 403 verir."""
    project = await project_factory("P9-EX-2")
    token = await _login(client, user_factory, "procurement")

    resp = await client.get(_url(project.id), headers=_auth(token))

    assert resp.status_code == 403


async def test_export_invisible_project_returns_404(client, user_factory, project_factory):
    """Kapsam satiri OLMAYAN `patron`: proje vardir ama gorunmez → 404 (403 DEGIL)."""
    project = await project_factory("P9-EX-3")
    token = await _login(client, user_factory, "patron")

    resp = await client.get(_url(project.id), headers=_auth(token))

    assert resp.status_code == 404
    assert resp.json()["detail"] == _PROJECT_MISSING


async def test_export_unknown_project_same_answer_as_invisible(
    client, user_factory, project_factory
):
    """Var olmayan UUID ile gorunmez proje AYNI cevabi verir — govde de sizdirmaz."""
    project = await project_factory("P9-EX-4")
    token = await _login(client, user_factory, "patron")

    invisible = await client.get(_url(project.id), headers=_auth(token))
    unknown = await client.get(_url(uuid.uuid4()), headers=_auth(token))

    assert unknown.status_code == invisible.status_code == 404
    assert unknown.json() == invisible.json()


# --- Dosya zarfi ---


async def test_export_returns_xlsx_with_content_disposition(
    client, db_session, user_factory, project_factory
):
    project = await project_factory("P9-EX-5")
    token = await _login(client, user_factory, "system_admin")

    resp = await client.get(_url(project.id), headers=_auth(token))

    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == _XLSX_MIME
    disposition = resp.headers["content-disposition"]
    assert disposition.startswith("attachment; ")
    # Proje kodu dosya adinda GORUNUR: iki projenin dosyasi ayni klasorde birbirini ezmemeli.
    assert "P9-EX-5" in disposition


async def test_export_without_units_still_returns_valid_workbook(
    client, db_session, user_factory, project_factory
):
    """Unitesi olmayan projede de GECERLI dosya doner: yalniz baslik satiri."""
    project = await project_factory("P9-EX-6")
    token = await _login(client, user_factory, "system_admin")

    resp = await client.get(_url(project.id), headers=_auth(token))

    assert _rows(_sheet(resp.content)) == [_HEADERS]


# --- Icerik (KKP 86-92) ---


async def test_export_headers_match_mockup(client, db_session, user_factory, project_factory):
    project = await project_factory("P9-EX-7")
    site = await _site(db_session, project)
    block = await _block(db_session, project, site)
    await _unit(db_session, project, block, "1")
    token = await _login(client, user_factory, "system_admin")

    resp = await client.get(_url(project.id), headers=_auth(token))

    assert _rows(_sheet(resp.content))[0] == _HEADERS


async def test_export_landowner_row_carries_shareholder(
    client, db_session, user_factory, project_factory
):
    """KKP 91'in ARSA yarisi: hissedar adi (`shareholder_name`).

    KKP 92: ARSA unitesinin satis durumu "Arsa Sahibinde"dir — `sales_status`
    sutunu ne olursa olsun (o kume `owner_side` turevini KAPSAMAZ, models §110).
    """
    project = await project_factory("P9-EX-8", project_type="kat_karsiligi")
    site = await _site(db_session, project)
    block = await _block(db_session, project, site, name="A Blok")
    shareholder = await _shareholder(db_session, project)
    await _unit(
        db_session,
        project,
        block,
        "2",
        layout="3+1",
        gross_area_m2=Decimal("148.00"),
        appraisal_value=Decimal("1380000.00"),
        owner_side=UnitOwnerSide.landowner,
        shareholder_id=shareholder.id,
        sales_status=UnitSalesStatus.listed,
    )
    token = await _login(client, user_factory, "system_admin")

    resp = await client.get(_url(project.id), headers=_auth(token))

    assert _rows(_sheet(resp.content))[1] == (
        "A Blok · 2",
        "3+1",
        "148.00",
        "1380000.00",
        "ARSA",
        "Ahmet Yılmaz",
        "Arsa Sahibinde",
    )


async def test_export_contractor_row_carries_buyer(
    client, db_session, user_factory, project_factory
):
    """KKP 91'in BIZ yarisi: alici adi (`buyer_name`), hissedar DEGIL."""
    project = await project_factory("P9-EX-9", project_type="kat_karsiligi")
    site = await _site(db_session, project)
    block = await _block(db_session, project, site, name="A Blok")
    unit = await _unit(
        db_session,
        project,
        block,
        "1",
        layout="3+1",
        gross_area_m2=Decimal("148.00"),
        appraisal_value=Decimal("1380000.00"),
        owner_side=UnitOwnerSide.contractor,
        sales_status=UnitSalesStatus.sold,
    )
    user = await user_factory(email="p9export@t.co", password="parola1234", role_key="system_admin")
    await _sale(db_session, project, unit, user.id)
    resp = await client.post(
        "/auth/login", json={"email": "p9export@t.co", "password": "parola1234"}
    )
    token = resp.json()["access_token"]

    resp = await client.get(_url(project.id), headers=_auth(token))

    assert _rows(_sheet(resp.content))[1] == (
        "A Blok · 1",
        "3+1",
        "148.00",
        "1380000.00",
        "BİZ",
        "Serkan Öz",
        "Satıldı",
    )


async def test_export_empty_cells_use_dash(client, db_session, user_factory, project_factory):
    """Atanmamis unite: taraf da, taraf adi da, durum da UYDURULMAZ (KKP 119 "—").

    `sales_status` sutunu NULLABLE'dir ama `server_default`i `listed`tir: ORM'e
    `None` VERMEK yetmez (SQLAlchemy sutunu INSERT'ten duserir ve sunucu
    varsayilani devreye girer), bu yuzden bos durum ACIK bir UPDATE ile kurulur.
    """
    project = await project_factory("P9-EX-10", project_type="kat_karsiligi")
    site = await _site(db_session, project)
    block = await _block(db_session, project, site, name="B Blok")
    unit = await _unit(db_session, project, block, "1")
    await db_session.execute(update(Unit).where(Unit.id == unit.id).values(sales_status=None))
    token = await _login(client, user_factory, "system_admin")

    resp = await client.get(_url(project.id), headers=_auth(token))

    assert _rows(_sheet(resp.content))[1] == ("B Blok · 1", "—", "—", "—", "—", "—", "—")


async def test_export_landowner_without_shareholder_is_dash(
    client, db_session, user_factory, project_factory
):
    """ARSA + hissedar atanmamis GECERLIDIR (spec §4.2) — ad uydurulmaz."""
    project = await project_factory("P9-EX-11", project_type="kat_karsiligi")
    site = await _site(db_session, project)
    block = await _block(db_session, project, site, name="A Blok")
    await _unit(db_session, project, block, "3", owner_side=UnitOwnerSide.landowner)
    token = await _login(client, user_factory, "system_admin")

    resp = await client.get(_url(project.id), headers=_auth(token))

    row = _rows(_sheet(resp.content))[1]
    assert row[4:] == ("ARSA", "—", "Arsa Sahibinde")


async def test_export_rows_match_list_endpoint(client, db_session, user_factory, project_factory):
    """Kural 2: dosya LISTE UCUYLA AYNI kaynaktan besleniyor mu?

    Satir sayisi ve `Ünite` sutunu liste ucunun dondurdugu etiketlerle BIREBIR
    ayni sirada olmalidir — Excel kendi sorgusunu acsaydi bu iddia kirilirdi.
    """
    project = await project_factory("P9-EX-12", project_type="kat_karsiligi")
    site = await _site(db_session, project)
    first = await _block(db_session, project, site, name="A Blok")
    second = await _block(db_session, project, site, name="B Blok", code="B")
    await _unit(db_session, project, first, "1")
    await _unit(db_session, project, first, "2", owner_side=UnitOwnerSide.landowner)
    await _unit(db_session, project, second, "1")
    token = await _login(client, user_factory, "system_admin")

    listed = await client.get(f"/projects/{project.id}/units", headers=_auth(token))
    exported = await client.get(_url(project.id), headers=_auth(token))

    labels = [unit["label"] for group in listed.json()["blocks"] for unit in group["units"]]
    assert [row[0] for row in _rows(_sheet(exported.content))[1:]] == labels


async def test_export_values_match_list_endpoint_json(
    client, db_session, user_factory, project_factory
):
    """Sayisal sutunlar API JSON'u ile AYNI metin temsilini tasir (boq emsali):
    ekranda "148.00" gorunup dosyada "148" yazmasi ayrisma demektir."""
    project = await project_factory("P9-EX-13", project_type="kat_karsiligi")
    site = await _site(db_session, project)
    block = await _block(db_session, project, site, name="A Blok")
    await _unit(
        db_session,
        project,
        block,
        "9",
        gross_area_m2=Decimal("112.50"),
        appraisal_value=Decimal("1040000.00"),
    )
    token = await _login(client, user_factory, "system_admin")

    listed = await client.get(f"/projects/{project.id}/units", headers=_auth(token))
    exported = await client.get(_url(project.id), headers=_auth(token))

    satir = listed.json()["blocks"][0]["units"][0]
    row = _rows(_sheet(exported.content))[1]
    assert row[2] == satir["gross_area_m2"]
    assert row[3] == satir["appraisal_value"]
