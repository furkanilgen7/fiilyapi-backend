"""P3.1 T14 — `GET .../units/import/template` (spec §6.7, §12.5/53-54).

Sablon ucu akisin ILK adimidir, sus degildir: EI 54 "sablonu indirip doldurun,
sonra yukleyin" diye tarif eder ve dugme mockup'ta IKI KEZ gecer (EI 37, 87).

Bu dosyanin kilitledigi uc kural:

1. **Basliklar `importer.COLUMNS`'tan TURETILIR.** Sablonda elle yazilmis
   ikinci bir baslik listesi olsaydi, `COLUMNS` degistiginde sablon sessizce
   eskir ve kullanicinin indirdigi dosya "baslik eksik" 422'si alirdi. Son
   test (`..._import_ucunda_kabul_edilir`) bu ayrismayi UCTAN UCA yakalar.
2. **Veri satiri YOKTUR.** Ornek satir koymak, kullanicinin onu silmeyi unutup
   hatali satir olarak yuklemesine yol acar (spec §6.7).
3. **Izin `view`'dir** (spec §6.2 karari, §12.6/I6): bos bir baslik satiri
   hicbir proje verisi icermez; `full`'a kapatmak veri GIRECEK kullaniciyi
   sablona ulasamaz hâle getirirdi.
"""

import io

from openpyxl import load_workbook
from sqlalchemy import delete, func, select

from app.modules.audit.models import AuditLog
from app.modules.units.importer import COLUMNS
from tests.modules.units.test_units_api import _auth, _login, _login_with_access, _site

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Spec §6.4 sutun tablosu (EI 85). Bu liste `importer.COLUMNS`'tan BAGIMSIZ
# olarak BIREBIR yazilir: turetilmis bir beklenti, turetilmis bir uretimi
# dogrulayamaz (ikisi birlikte kayarsa test yesil kalirdi).
_EXPECTED_HEADERS = [
    "Blok",
    "Kat",
    "Ünite No",
    "Tür",
    "Oda Tipi",
    "Brüt m²",
    "Net m²",
    "Cephe",
    "Liste Fiyatı",
    "Rayiç Değer",
    "Maliyet",
    "Sahiplik",
]


def _sheet(content: bytes):
    return load_workbook(io.BytesIO(content)).active


async def _audit_count(session) -> int:
    return int((await session.execute(select(func.count()).select_from(AuditLog))).scalar_one())


async def test_template_returns_200_with_xlsx_content_type(
    client, db_session, user_factory, project_factory
):
    """Spec §12.5/53. `Content-Type` frontend'in ikili indirme kuralinin
    dayanagidir (`GOREV-SIRASI.md` §3): JSON donen bir uc sessizce bozuk bir
    `.xlsx` indirtirdi."""
    project = await project_factory("T14-1")
    token = await _login(client, user_factory, "system_admin")

    resp = await client.get(f"/projects/{project.id}/units/import/template", headers=_auth(token))

    assert resp.status_code == 200
    assert resp.headers["content-type"] == _XLSX_MIME
    assert "attachment" in resp.headers["content-disposition"]


async def test_template_first_row_has_twelve_headers_in_canonical_order(
    client, db_session, user_factory, project_factory
):
    """Spec §12.5/53 + §6.4. SIRA da iddianin parcasidir: kullanici sutunlari
    kaydirirsa `Kat` sutununa `Ünite No` yazar ve veri sessizce yanlis kolona
    duser."""
    project = await project_factory("T14-2")
    token = await _login(client, user_factory, "system_admin")

    resp = await client.get(f"/projects/{project.id}/units/import/template", headers=_auth(token))

    sheet = _sheet(resp.content)
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert headers == _EXPECTED_HEADERS
    assert len(headers) == 12
    # Ikinci otorite yasagi: uretim listesi `importer.COLUMNS` ile AYNI olmali.
    assert headers == [column.label for column in COLUMNS]


async def test_template_has_no_data_rows(client, db_session, user_factory, project_factory):
    """Spec §12.5/53. Ornek satir koymak, kullanicinin onu silmeyi unutup
    hatali satir olarak yuklemesine yol acar (spec §6.7)."""
    project = await project_factory("T14-3")
    token = await _login(client, user_factory, "system_admin")

    resp = await client.get(f"/projects/{project.id}/units/import/template", headers=_auth(token))

    sheet = _sheet(resp.content)
    assert sheet.max_row == 1


async def test_template_writes_no_audit(client, db_session, user_factory, project_factory):
    """Spec §12.5/54 + §9. Okuma ucudur: "Sablon Indir"e her basista gunluge
    satir dusen bir ekran ortaya cikmamalidir (P4 T7 kurali)."""
    project = await project_factory("T14-4")
    token = await _login(client, user_factory, "system_admin")
    await db_session.execute(delete(AuditLog))

    resp = await client.get(f"/projects/{project.id}/units/import/template", headers=_auth(token))

    assert resp.status_code == 200
    assert await _audit_count(db_session) == 0


async def test_template_allowed_with_view_permission_200(
    client, db_session, user_factory, project_factory
):
    """Spec §12.6/I6 + §6.2 karari. `site_chief` (`projects`=view) sablonu
    INDIREBILIR — diger tum yeni uclar (preview/validate) ayni rolde 403 doner.
    Fark bilinclidir: bos bir baslik satiri proje verisi tasimaz."""
    project = await project_factory("T14-5")
    token = await _login_with_access(client, db_session, user_factory, "site_chief")

    resp = await client.get(f"/projects/{project.id}/units/import/template", headers=_auth(token))

    assert resp.status_code == 200


async def test_downloaded_template_is_accepted_by_import_endpoint(
    client, db_session, user_factory, project_factory
):
    """DONGU TESTI. Sablon indirilir, TEK satir doldurulur, `import`'a verilir.

    Baslik listesi sablon ile cozumleyici arasinda ayrisirsa (iki otorite
    tuzagi) bu test kirmizi doner — diger testler ayrismayi GOREMEZ, cunku
    ikisi de ayni yanlis listeyi kullanirdi.
    """
    # `Sahiplik` sutunu YALNIZ kat karsiligi projede dolabilir (spec §3.3) —
    # dongu testi 12 sutunun HEPSINI doldurdugu icin proje tipi de ona gore secilir.
    project = await project_factory("T14-6", project_type="kat_karsiligi")
    await _site(db_session, project)
    token = await _login(client, user_factory, "system_admin")

    template = await client.get(
        f"/projects/{project.id}/units/import/template", headers=_auth(token)
    )
    workbook = load_workbook(io.BytesIO(template.content))
    workbook.active.append(
        ["A Blok", 3, "3", "Daire", "3+1", 120, 95, "Güney", 4500000, None, None, "BİZ"]
    )
    buffer = io.BytesIO()
    workbook.save(buffer)

    resp = await client.post(
        f"/projects/{project.id}/units/import",
        files={"file": ("uniteler.xlsx", buffer.getvalue(), _XLSX_MIME)},
        headers=_auth(token),
    )

    assert resp.status_code == 200, resp.text
    assert resp.json()["created"] == 1
