"""T8 — GET /sites/{site_id}/boq/export (spec §5.3)."""

import uuid
from decimal import Decimal
from io import BytesIO

import openpyxl

from app.modules.boq.models import BoqGroup, BoqItem
from app.modules.sites.models import Site
from app.modules.users.models import UserProjectAccess

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


async def _login(client, user_factory, role_key: str, email: str | None = None) -> str:
    address = email or f"{role_key}@boq-export.co"
    await user_factory(email=address, password="parola1234", role_key=role_key)
    resp = await client.post("/auth/login", json={"email": address, "password": "parola1234"})
    return resp.json()["access_token"]


async def _login_with_access(client, session, user_factory, role_key: str, email: str) -> str:
    user = await user_factory(email=email, password="parola1234", role_key=role_key)
    session.add(UserProjectAccess(user_id=user.id, project_id=None, all_projects=True))
    await session.flush()
    resp = await client.post("/auth/login", json={"email": email, "password": "parola1234"})
    return resp.json()["access_token"]


def _auth(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


async def _site(session, project, code: str = "A-BLOK", **kwargs) -> Site:
    site = Site(project_id=project.id, code=code, name=kwargs.pop("name", "A-Blok Şantiyesi"))
    for field, value in kwargs.items():
        setattr(site, field, value)
    session.add(site)
    await session.flush()
    return site


async def _group(session, site, name: str = "TOPRAK VE TEMEL İŞLERİ", **kwargs) -> BoqGroup:
    group = BoqGroup(site_id=site.id, name=name, **kwargs)
    session.add(group)
    await session.flush()
    return group


async def _item(session, site, group, code: str = "01.001", **kwargs) -> BoqItem:
    defaults = {
        "description": "Kazı (Makine ile)",
        "unit": "m³",
        "quantity": Decimal("1240.000"),
        "unit_price": Decimal("280.00"),
    }
    defaults.update(kwargs)
    item = BoqItem(site_id=site.id, group_id=group.id, code=code, **defaults)
    session.add(item)
    await session.flush()
    return item


def _load_workbook(content: bytes):
    return openpyxl.load_workbook(BytesIO(content))


async def test_export_boq_unauthenticated(client, db_session, project_factory):
    project = await project_factory("BOQ-EXPORT-1")
    site = await _site(db_session, project)

    resp = await client.get(f"/sites/{site.id}/boq/export")

    assert resp.status_code == 401


async def test_export_boq_role_without_boq_permission_forbidden(
    client, db_session, user_factory, project_factory
):
    project = await project_factory("BOQ-EXPORT-2")
    site = await _site(db_session, project)
    token = await _login_with_access(
        client, db_session, user_factory, "hr_manager", "hr@boq-export-2.co"
    )

    resp = await client.get(f"/sites/{site.id}/boq/export", headers=_auth(token))

    assert resp.status_code == 403


async def test_export_boq_invisible_site_returns_404(
    client, db_session, user_factory, project_factory
):
    project = await project_factory("BOQ-EXPORT-3")
    site = await _site(db_session, project)
    token = await _login(client, user_factory, "site_chief", "sc@boq-export-3.co")

    resp = await client.get(f"/sites/{site.id}/boq/export", headers=_auth(token))

    assert resp.status_code == 404
    assert resp.json()["detail"] == "Şantiye bulunamadı"


async def test_export_boq_missing_site_returns_404(client, user_factory):
    token = await _login(client, user_factory, "system_admin")

    resp = await client.get(f"/sites/{uuid.uuid4()}/boq/export", headers=_auth(token))

    assert resp.status_code == 404


async def test_export_boq_view_only_role_allowed(client, db_session, user_factory, project_factory):
    """boq:view yeterlidir (yazma degil) — site_chief indirebilir."""
    project = await project_factory("BOQ-EXPORT-4")
    site = await _site(db_session, project)
    token = await _login_with_access(
        client, db_session, user_factory, "site_chief", "sc@boq-export-4.co"
    )

    resp = await client.get(f"/sites/{site.id}/boq/export", headers=_auth(token))

    assert resp.status_code == 200
    assert resp.headers["content-type"] == XLSX_MEDIA_TYPE


async def test_export_boq_happy_path_matches_mockup_headers(
    client, db_session, user_factory, project_factory
):
    project = await project_factory("BOQ-EXPORT-5")
    site = await _site(db_session, project, code="A-BLOK-5")
    group = await _group(db_session, site, name="TOPRAK VE TEMEL İŞLERİ", sort_order=1)
    await _item(
        db_session,
        site,
        group,
        code="01.001",
        description="Kazı (Makine ile)",
        unit="m³",
        quantity=Decimal("1240.000"),
        unit_price=Decimal("280.00"),
        sort_order=1,
    )
    token = await _login(client, user_factory, "system_admin")

    resp = await client.get(f"/sites/{site.id}/boq/export", headers=_auth(token))

    assert resp.status_code == 200
    assert resp.headers["content-type"] == XLSX_MEDIA_TYPE
    disposition = resp.headers["content-disposition"]
    assert "is-kalemleri-A-BLOK-5.xlsx" in disposition

    workbook = _load_workbook(resp.content)
    sheet = workbook.active

    header_row = [cell.value for cell in sheet[1]]
    assert header_row == [
        "Poz No",
        "İş Kalemi Tarifi",
        "Birim",
        "Miktar",
        "Birim Fiyat",
        "Tutar",
        "Gerç. %",
    ]

    group_row = [cell.value for cell in sheet[2]]
    assert group_row[0] == "1. TOPRAK VE TEMEL İŞLERİ"

    item_row = [cell.value for cell in sheet[3]]
    assert item_row == [
        "01.001",
        "Kazı (Makine ile)",
        "m³",
        "1240.000",
        "280.00",
        "347200.00",
        None,
    ]

    total_row = [cell.value for cell in sheet[4]]
    assert total_row[0] == "GENEL TOPLAM"
    assert total_row[5] == "347200.00"
    assert total_row[6] is None


async def test_export_boq_empty_site_still_returns_valid_workbook(
    client, db_session, user_factory, project_factory
):
    project = await project_factory("BOQ-EXPORT-6")
    site = await _site(db_session, project)
    token = await _login(client, user_factory, "system_admin")

    resp = await client.get(f"/sites/{site.id}/boq/export", headers=_auth(token))

    assert resp.status_code == 200
    workbook = _load_workbook(resp.content)
    sheet = workbook.active
    total_row = [cell.value for cell in sheet[2]]
    assert total_row[0] == "GENEL TOPLAM"
    assert total_row[5] == "0.00"


async def test_export_boq_does_not_record_audit(client, db_session, user_factory, project_factory):
    """Okuma ucu (export dahil) denetim kaydi YAZMAZ (T7 kurali)."""
    from sqlalchemy import select

    from app.modules.audit.models import AuditAction, AuditLog

    project = await project_factory("BOQ-EXPORT-7")
    site = await _site(db_session, project)
    token = await _login(client, user_factory, "system_admin")

    resp = await client.get(f"/sites/{site.id}/boq/export", headers=_auth(token))

    assert resp.status_code == 200
    rows = (await db_session.execute(select(AuditLog))).scalars().all()
    assert all(row.action != AuditAction.create for row in rows)
