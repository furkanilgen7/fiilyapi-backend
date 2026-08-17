"""BOQ-SEC — iş kalemlerinin bölümlere MİKTAR TAHSİSİ (K1-K6).

Emirdeki dokuz AYRIŞMA NOKTASININ sekizi buradadır; dokuzuncusu (eşzamanlı iki
replace) `test_boq_allocation_concurrency.py`dedir — o senaryo tek bağlantılı
`db_session` üzerinde YAPILAMAZ (SAVEPOINT sarmalayıcısı gerçek satır kilidini
gizler, FAT-1 dersi).

🔴 Bu dosyanın en kritik kuralı: bir DEĞER testi tek başına yetmez. Sınırda
GEÇEN ve sınırın bir birim üstünde REDDEDİLEN durumlar AYRI AYRI iddia edilir —
"reddediyor mu" sorusu, eşiğin doğru YERDE olduğunu kanıtlamaz.
"""

import uuid
from decimal import Decimal

from sqlalchemy import select

from app.core.access import AccessLevel
from app.modules.audit.models import AuditAction, AuditLog
from app.modules.boq.models import BoqGroup, BoqItem, BoqItemSectionAllocation
from app.modules.roles.models import Module, Role, RolePermission
from app.modules.sites.models import Section, Site
from app.modules.users.models import UserProjectAccess

# --- Kurulum yardımcıları (test_boq_api.py deseniyle birebir) ---------------


async def _set_permission(session, role_key: str, module_key: str, level: AccessLevel) -> None:
    role_id = (await session.execute(select(Role.id).where(Role.key == role_key))).scalar_one()
    module_id = (
        await session.execute(select(Module.id).where(Module.key == module_key))
    ).scalar_one()
    permission = (
        await session.execute(
            select(RolePermission).where(
                RolePermission.role_id == role_id, RolePermission.module_id == module_id
            )
        )
    ).scalar_one()
    permission.access_level = level
    await session.flush()


async def _login(client, session, user_factory, role_key: str, email: str) -> str:
    user = await user_factory(email=email, password="parola1234", role_key=role_key)
    session.add(UserProjectAccess(user_id=user.id, project_id=None, all_projects=True))
    await session.flush()
    resp = await client.post("/auth/login", json={"email": email, "password": "parola1234"})
    return resp.json()["access_token"]


def _auth(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


async def _site(session, project, code: str = "A-BLOK") -> Site:
    site = Site(project_id=project.id, code=code, name=f"{code} Şantiyesi")
    session.add(site)
    await session.flush()
    return site


async def _section(session, site, name: str, code: str | None = None) -> Section:
    section = Section(site_id=site.id, name=name, code=code)
    session.add(section)
    await session.flush()
    return section


async def _group(session, site, name: str = "TOPRAK VE TEMEL İŞLERİ", sort_order: int = 0):
    group = BoqGroup(site_id=site.id, name=name, sort_order=sort_order)
    session.add(group)
    await session.flush()
    return group


async def _item(session, site, group, code: str = "01.001", **kwargs) -> BoqItem:
    defaults = {
        "description": "Beton Dökümü C30",
        "unit": "m³",
        "quantity": Decimal("1200.000"),
        "unit_price": Decimal("100.00"),
    }
    defaults.update(kwargs)
    item = BoqItem(site_id=site.id, group_id=group.id, code=code, **defaults)
    session.add(item)
    await session.flush()
    return item


async def _allocation(session, item, section, quantity: str) -> BoqItemSectionAllocation:
    row = BoqItemSectionAllocation(
        boq_item_id=item.id, section_id=section.id, quantity=Decimal(quantity)
    )
    session.add(row)
    await session.flush()
    return row


def _put_url(item_id) -> str:
    return f"/boq/items/{item_id}/allocations"


# --- Ayrışma 1-4: SINIR ARİTMETİĞİ -----------------------------------------


async def test_toplam_quantitye_TAM_ESITSE_gecer(client, db_session, user_factory, project_factory):
    """🔴 Ayrışma 1 — sınırda GEÇMELİ. `<` yazılsaydı burası kırmızı olurdu."""
    project = await project_factory("BOQSEC-1")
    site = await _site(db_session, project)
    group = await _group(db_session, site)
    item = await _item(db_session, site, group)
    kat_a = await _section(db_session, site, "Kat 6-10")
    kat_b = await _section(db_session, site, "Kat 11-15")
    token = await _login(client, db_session, user_factory, "system_admin", "s1@boqsec.co")

    resp = await client.put(
        _put_url(item.id),
        json={
            "allocations": [
                {"section_id": str(kat_a.id), "quantity": "700.000"},
                {"section_id": str(kat_b.id), "quantity": "500.000"},
            ]
        },
        headers=_auth(token),
    )

    assert resp.status_code == 200, resp.text
    govde = resp.json()
    assert Decimal(govde["item"]["allocated_quantity"]) == Decimal("1200.000")
    assert Decimal(govde["item"]["unallocated_quantity"]) == Decimal("0.000")


async def test_toplam_quantityden_BIR_BIRIM_fazlaysa_409(
    client, db_session, user_factory, project_factory
):
    """🔴 Ayrışma 2 — sınırın 0,001 üstü REDDEDİLMELİ.

    Ayrışma 1 ile birlikte eşiğin TAM YERİNİ çiviler: tek başına ikisinden biri
    `<` ile `<=` arasındaki farkı göremez.
    """
    project = await project_factory("BOQSEC-2")
    site = await _site(db_session, project)
    group = await _group(db_session, site)
    item = await _item(db_session, site, group)
    kat = await _section(db_session, site, "Kat 6-10")
    token = await _login(client, db_session, user_factory, "system_admin", "s2@boqsec.co")

    resp = await client.put(
        _put_url(item.id),
        json={"allocations": [{"section_id": str(kat.id), "quantity": "1200.001"}]},
        headers=_auth(token),
    )

    assert resp.status_code == 409, resp.text
    assert "aşamaz" in resp.json()["detail"]
    assert (await _tahsis_sayisi(db_session, item)) == 0, "reddedilen istek satır YAZDI"


async def test_tek_tahsis_quantitynin_TAMAMI(client, db_session, user_factory, project_factory):
    """Ayrışma 3 — tek bölüme tam kota."""
    project = await project_factory("BOQSEC-3")
    site = await _site(db_session, project)
    group = await _group(db_session, site)
    item = await _item(db_session, site, group)
    kat = await _section(db_session, site, "Kat 6-10")
    token = await _login(client, db_session, user_factory, "system_admin", "s3@boqsec.co")

    resp = await client.put(
        _put_url(item.id),
        json={"allocations": [{"section_id": str(kat.id), "quantity": "1200.000"}]},
        headers=_auth(token),
    )

    assert resp.status_code == 200, resp.text
    assert Decimal(resp.json()["item"]["unallocated_quantity"]) == Decimal("0.000")


async def test_kismi_dagitim_atanmamis_miktari_birakir(
    client, db_session, user_factory, project_factory
):
    """Ayrışma 4 — KK-B1'in birebir örneği: 1.200'ün 400+300'ü dağıtılır, 500 kalır."""
    project = await project_factory("BOQSEC-4")
    site = await _site(db_session, project)
    group = await _group(db_session, site)
    item = await _item(db_session, site, group)
    kat_a = await _section(db_session, site, "Kat 6-10")
    kat_b = await _section(db_session, site, "Kat 11-15")
    token = await _login(client, db_session, user_factory, "system_admin", "s4@boqsec.co")

    resp = await client.put(
        _put_url(item.id),
        json={
            "allocations": [
                {"section_id": str(kat_a.id), "quantity": "400.000"},
                {"section_id": str(kat_b.id), "quantity": "300.000"},
            ]
        },
        headers=_auth(token),
    )

    assert resp.status_code == 200, resp.text
    govde = resp.json()
    assert Decimal(govde["item"]["allocated_quantity"]) == Decimal("700.000")
    assert Decimal(govde["item"]["unallocated_quantity"]) == Decimal("500.000")
    assert {a["section_name"] for a in govde["allocations"]} == {"Kat 6-10", "Kat 11-15"}


# --- Ayrışma 5: REPLACE ile toplamı AZALTMA --------------------------------


async def test_replace_toplami_azaltirken_ARA_DURUM_ihlali_uydurmaz(
    client, db_session, user_factory, project_factory
):
    """🔴 Ayrışma 5 — eski satırlar silinip yenileri yazılırken invariant ARA
    DURUMDA ihlal edilmiş GİBİ görünmemeli.

    1.200'ün tamamı iki bölümdeyken gövde "hepsi tek bölümde 1.200" der. Naif bir
    uygulama önce 1.200'lük yeni satırı yazıp sonra eskileri silseydi ara toplam
    2.400 olur ve kendi kontrolüne takılırdı. Kimlik koruyan replace bunu
    yapısal olarak yaşamaz.
    """
    project = await project_factory("BOQSEC-5")
    site = await _site(db_session, project)
    group = await _group(db_session, site)
    item = await _item(db_session, site, group)
    kat_a = await _section(db_session, site, "Kat 6-10")
    kat_b = await _section(db_session, site, "Kat 11-15")
    await _allocation(db_session, item, kat_a, "700.000")
    await _allocation(db_session, item, kat_b, "500.000")
    token = await _login(client, db_session, user_factory, "system_admin", "s5@boqsec.co")

    resp = await client.put(
        _put_url(item.id),
        json={"allocations": [{"section_id": str(kat_a.id), "quantity": "1200.000"}]},
        headers=_auth(token),
    )

    assert resp.status_code == 200, resp.text
    assert len(resp.json()["allocations"]) == 1
    assert (await _tahsis_sayisi(db_session, item)) == 1


async def test_bos_dizi_TUM_tahsisleri_kaldirir(client, db_session, user_factory, project_factory):
    """K4: `[]` = hepsini kaldır. Miktar tamamen "atanmamış"a döner."""
    project = await project_factory("BOQSEC-6")
    site = await _site(db_session, project)
    group = await _group(db_session, site)
    item = await _item(db_session, site, group)
    kat = await _section(db_session, site, "Kat 6-10")
    await _allocation(db_session, item, kat, "400.000")
    token = await _login(client, db_session, user_factory, "system_admin", "s6@boqsec.co")

    resp = await client.put(_put_url(item.id), json={"allocations": []}, headers=_auth(token))

    assert resp.status_code == 200, resp.text
    assert resp.json()["allocations"] == []
    assert Decimal(resp.json()["item"]["unallocated_quantity"]) == Decimal("1200.000")
    assert (await _tahsis_sayisi(db_session, item)) == 0


# --- K4: SESSİZ YORUM REDDİ ------------------------------------------------


async def test_alan_HIC_gonderilmezse_422_sessizce_yorumlanmaz(
    client, db_session, user_factory, project_factory
):
    """🔴 K4 — bu uçta "dokunma" anlamı YOKTUR.

    Eksik alanı "değiştirme" ya da "hepsini sil" diye yorumlamak, kullanıcının
    niyetini sunucunun UYDURMASI olurdu; ikisi de veri kaybı üretir.
    """
    project = await project_factory("BOQSEC-7")
    site = await _site(db_session, project)
    group = await _group(db_session, site)
    item = await _item(db_session, site, group)
    kat = await _section(db_session, site, "Kat 6-10")
    await _allocation(db_session, item, kat, "400.000")
    token = await _login(client, db_session, user_factory, "system_admin", "s7@boqsec.co")

    eksik = await client.put(_put_url(item.id), json={}, headers=_auth(token))
    boslukta = await client.put(_put_url(item.id), json={"allocations": None}, headers=_auth(token))

    assert eksik.status_code == 422, eksik.text
    assert boslukta.status_code == 422, boslukta.text
    # 🔴 "Reddedildi" yetmez: kayıt DOKUNULMAMIŞ olmalı.
    assert (await _tahsis_sayisi(db_session, item)) == 1


async def test_ayni_bolum_iki_kez_422_ve_SESSIZCE_TOPLANMAZ(
    client, db_session, user_factory, project_factory
):
    """🔴 Sunucu iki satırı toplamaz: 400+300 = 700 yazsaydı kullanıcı ekranda
    hiç girmediği bir sayı görürdü."""
    project = await project_factory("BOQSEC-8")
    site = await _site(db_session, project)
    group = await _group(db_session, site)
    item = await _item(db_session, site, group)
    kat = await _section(db_session, site, "Kat 6-10")
    token = await _login(client, db_session, user_factory, "system_admin", "s8@boqsec.co")

    resp = await client.put(
        _put_url(item.id),
        json={
            "allocations": [
                {"section_id": str(kat.id), "quantity": "400.000"},
                {"section_id": str(kat.id), "quantity": "300.000"},
            ]
        },
        headers=_auth(token),
    )

    assert resp.status_code == 422, resp.text
    assert (await _tahsis_sayisi(db_session, item)) == 0


async def test_sifir_ve_negatif_miktar_422(client, db_session, user_factory, project_factory):
    """Sıfır tahsis bir SATIR olarak tutulmaz (K1 CHECK'iyle aynı kural)."""
    project = await project_factory("BOQSEC-9")
    site = await _site(db_session, project)
    group = await _group(db_session, site)
    item = await _item(db_session, site, group)
    kat = await _section(db_session, site, "Kat 6-10")
    token = await _login(client, db_session, user_factory, "system_admin", "s9@boqsec.co")

    for miktar in ("0", "0.000", "-5.000"):
        resp = await client.put(
            _put_url(item.id),
            json={"allocations": [{"section_id": str(kat.id), "quantity": miktar}]},
            headers=_auth(token),
        )
        assert resp.status_code == 422, f"{miktar}: {resp.text}"


# --- Kapsam (IDOR) ---------------------------------------------------------


async def test_baska_santiyenin_bolumu_404(client, db_session, user_factory, project_factory):
    """K4: başka şantiyenin bölüm UUID'si, var olmayan UUID ile AYNI 404'ü alır."""
    project = await project_factory("BOQSEC-10")
    site = await _site(db_session, project, "A-BLOK")
    komsu = await _site(db_session, project, "B-BLOK")
    group = await _group(db_session, site)
    item = await _item(db_session, site, group)
    yabanci = await _section(db_session, komsu, "Komşu Kat")
    token = await _login(client, db_session, user_factory, "system_admin", "s10@boqsec.co")

    yabanci_resp = await client.put(
        _put_url(item.id),
        json={"allocations": [{"section_id": str(yabanci.id), "quantity": "10.000"}]},
        headers=_auth(token),
    )
    hayali_resp = await client.put(
        _put_url(item.id),
        json={"allocations": [{"section_id": str(uuid.uuid4()), "quantity": "10.000"}]},
        headers=_auth(token),
    )

    assert yabanci_resp.status_code == 404
    assert hayali_resp.status_code == 404
    # 🔴 İki gövde AYIRT EDİLEMEZ olmalı — aksi hâlde uç bir bölüm tarayıcısıdır.
    assert yabanci_resp.json() == hayali_resp.json()


async def test_tahsis_yazma_full_izin_ister(client, db_session, user_factory, project_factory):
    """K8: yazma `boq:full`. `view` yeterli DEĞİL, yeni izin modülü de yok."""
    project = await project_factory("BOQSEC-11")
    site = await _site(db_session, project)
    group = await _group(db_session, site)
    item = await _item(db_session, site, group)
    kat = await _section(db_session, site, "Kat 6-10")
    await _set_permission(db_session, "site_chief", "boq", AccessLevel.view)
    token = await _login(client, db_session, user_factory, "site_chief", "s11@boqsec.co")

    resp = await client.put(
        _put_url(item.id),
        json={"allocations": [{"section_id": str(kat.id), "quantity": "10.000"}]},
        headers=_auth(token),
    )

    assert resp.status_code == 403


async def test_tahsis_yazma_denetim_gunlugune_yazar(
    client, db_session, user_factory, project_factory
):
    project = await project_factory("BOQSEC-12")
    site = await _site(db_session, project)
    group = await _group(db_session, site)
    item = await _item(db_session, site, group)
    kat = await _section(db_session, site, "Kat 6-10")
    token = await _login(client, db_session, user_factory, "system_admin", "s12@boqsec.co")

    await client.put(
        _put_url(item.id),
        json={"allocations": [{"section_id": str(kat.id), "quantity": "400.000"}]},
        headers=_auth(token),
    )

    kayitlar = (
        (await db_session.execute(select(AuditLog).where(AuditLog.action == AuditAction.update)))
        .scalars()
        .all()
    )
    assert any("bölüm tahsisleri" in k.detail and "01.001" in k.detail for k in kayitlar)


# --- Ayrışma 8-9: OKUMA SÜZGECİ (K5) ---------------------------------------


async def test_section_id_suzgeci_TAHSIS_MIKTARINI_dondurur_ve_bos_gruplari_duser(
    client, db_session, user_factory, project_factory
):
    """🔴 Ayrışma 8 — süzgeçte tahsissiz poz listede YOK, boş grup da yok.

    Üç iddia birbirinden ayrıdır: (a) dönen `quantity` bölüm payıdır, poz kotası
    DEĞİL; (b) `amount` o paydan türer; (c) tahsissiz kalemin TEK BAŞINA olduğu
    grup listeden tamamen düşer.
    """
    project = await project_factory("BOQSEC-13")
    site = await _site(db_session, project)
    grup_a = await _group(db_session, site, "BETON İŞLERİ", sort_order=0)
    grup_b = await _group(db_session, site, "BOYA İŞLERİ", sort_order=1)
    beton = await _item(db_session, site, grup_a, code="01.001")
    boya = await _item(db_session, site, grup_b, code="02.001", quantity=Decimal("50.000"))
    kat = await _section(db_session, site, "Kat 6-10")
    await _allocation(db_session, beton, kat, "400.000")
    token = await _login(client, db_session, user_factory, "system_admin", "s13@boqsec.co")

    resp = await client.get(
        f"/sites/{site.id}/boq", params={"section_id": str(kat.id)}, headers=_auth(token)
    )

    assert resp.status_code == 200, resp.text
    govde = resp.json()
    assert [g["name"] for g in govde["groups"]] == ["BETON İŞLERİ"], "boş grup düşmedi"
    (kalem,) = govde["groups"][0]["items"]
    assert Decimal(kalem["quantity"]) == Decimal("400.000"), "poz kotası dönmüş"
    assert Decimal(kalem["amount"]) == Decimal("40000.00"), "amount bölüm payından türemiyor"
    # 🔴 İKİ ANLAM: `unallocated` süzgeçte de POZUN GERÇEK kotasından türer.
    assert Decimal(kalem["allocated_quantity"]) == Decimal("400.000")
    assert Decimal(kalem["unallocated_quantity"]) == Decimal("800.000")
    assert Decimal(kalem["allocated_quantity"]) + Decimal(kalem["unallocated_quantity"]) == Decimal(
        "1200.000"
    ), "şantiye kotası yanıttan türetilemiyor"
    assert boya.id is not None  # kurulumun ikinci kalemi gerçekten yazıldı


async def test_section_id_VERILMEZSE_bugunku_davranis(
    client, db_session, user_factory, project_factory
):
    """🔴 Ayrışma 9 — VARSAYILAN YOL. Süzgeç yokken poz KOTASI döner ve tahsissiz
    kalemler listede KALIR (mevcut ekran kırılmaz)."""
    project = await project_factory("BOQSEC-14")
    site = await _site(db_session, project)
    grup_a = await _group(db_session, site, "BETON İŞLERİ", sort_order=0)
    grup_b = await _group(db_session, site, "BOYA İŞLERİ", sort_order=1)
    beton = await _item(db_session, site, grup_a, code="01.001")
    await _item(db_session, site, grup_b, code="02.001", quantity=Decimal("50.000"))
    kat = await _section(db_session, site, "Kat 6-10")
    await _allocation(db_session, beton, kat, "400.000")
    token = await _login(client, db_session, user_factory, "system_admin", "s14@boqsec.co")

    resp = await client.get(f"/sites/{site.id}/boq", headers=_auth(token))

    assert resp.status_code == 200
    govde = resp.json()
    assert [g["name"] for g in govde["groups"]] == ["BETON İŞLERİ", "BOYA İŞLERİ"]
    kalem = govde["groups"][0]["items"][0]
    assert Decimal(kalem["quantity"]) == Decimal("1200.000")
    assert Decimal(kalem["allocated_quantity"]) == Decimal("400.000")
    assert Decimal(kalem["unallocated_quantity"]) == Decimal("800.000")
    # Tahsissiz kalem: sıfır tahsis, tam atanmamış.
    boya_kalem = govde["groups"][1]["items"][0]
    assert Decimal(boya_kalem["allocated_quantity"]) == Decimal("0.000")
    assert Decimal(boya_kalem["unallocated_quantity"]) == Decimal("50.000")


async def test_okumada_baska_santiyenin_bolumu_BOS_LISTE_degil_404(
    client, db_session, user_factory, project_factory
):
    """K5 — boş liste dönmek "bu bölüme kalem atanmamış" YALANINI söylerdi."""
    project = await project_factory("BOQSEC-15")
    site = await _site(db_session, project, "A-BLOK")
    komsu = await _site(db_session, project, "B-BLOK")
    group = await _group(db_session, site)
    await _item(db_session, site, group)
    yabanci = await _section(db_session, komsu, "Komşu Kat")
    token = await _login(client, db_session, user_factory, "system_admin", "s15@boqsec.co")

    yabanci_resp = await client.get(
        f"/sites/{site.id}/boq", params={"section_id": str(yabanci.id)}, headers=_auth(token)
    )
    hayali_resp = await client.get(
        f"/sites/{site.id}/boq", params={"section_id": str(uuid.uuid4())}, headers=_auth(token)
    )

    assert yabanci_resp.status_code == 404
    assert hayali_resp.status_code == 404
    assert yabanci_resp.json() == hayali_resp.json()


async def test_export_AYNI_cagridan_beslenir(client, db_session, user_factory, project_factory):
    """🔴 K5 — Excel ile ekran AYRIŞAMAZ: iki uç aynı servis çağrısını kullanır.

    İddia sayı düzeyindedir: süzülmüş xlsx'te YALNIZ tahsisi olan kalem ve onun
    BÖLÜM PAYI bulunur; ikinci bir süzme kodu yazılsaydı burada poz kotası
    (1200.000) görünürdü.
    """
    from io import BytesIO

    from openpyxl import load_workbook

    project = await project_factory("BOQSEC-16")
    site = await _site(db_session, project)
    grup_a = await _group(db_session, site, "BETON İŞLERİ", sort_order=0)
    grup_b = await _group(db_session, site, "BOYA İŞLERİ", sort_order=1)
    beton = await _item(db_session, site, grup_a, code="01.001")
    await _item(db_session, site, grup_b, code="02.001", quantity=Decimal("50.000"))
    kat = await _section(db_session, site, "Kat 6-10")
    await _allocation(db_session, beton, kat, "400.000")
    token = await _login(client, db_session, user_factory, "system_admin", "s16@boqsec.co")

    resp = await client.get(
        f"/sites/{site.id}/boq/export", params={"section_id": str(kat.id)}, headers=_auth(token)
    )

    assert resp.status_code == 200
    sheet = load_workbook(BytesIO(resp.content)).active
    hucreler = [[c.value for c in row] for row in sheet.iter_rows()]
    duz = [str(v) for row in hucreler for v in row if v is not None]
    assert "01.001" in duz
    assert "02.001" not in duz, "süzgeç xlsx'e uygulanmadı"
    assert "400.000" in duz
    assert "1200.000" not in duz, "xlsx poz kotasını bastı — ikinci süzme kodu var"
    assert "40000.00" in duz


# --- Ayrışma 7 + K2: BÖLÜM SİLİNMESİ (ÜÇ AYRI İDDİA) -----------------------


async def test_bolum_silinince_UC_AYRI_iddia(client, db_session, user_factory, project_factory):
    """🔴 K2 CASCADE — üç iddia AYRI AYRI kurulur (biri diğerini kanıtlamaz):

    (a) tahsis satırı GİTTİ,
    (b) poz satırı ve `quantity`si DEĞİŞMEDİ (veri kaybı DEĞİL),
    (c) o pozun ATANMAMIŞ miktarı tahsis kadar ARTTI (miktar havuza döndü).

    `SET NULL` seçilseydi (a) tutar ama (c) TUTMAZDI: sahipsiz satır toplamda
    kalmaya devam eder ve kotayı görünmez biçimde bloke ederdi.
    """
    project = await project_factory("BOQSEC-17")
    site = await _site(db_session, project)
    group = await _group(db_session, site)
    item = await _item(db_session, site, group)
    kat_a = await _section(db_session, site, "Kat 6-10")
    kat_b = await _section(db_session, site, "Kat 11-15")
    await _allocation(db_session, item, kat_a, "400.000")
    await _allocation(db_session, item, kat_b, "300.000")
    token = await _login(client, db_session, user_factory, "system_admin", "s17@boqsec.co")

    once = await client.get(f"/sites/{site.id}/boq", headers=_auth(token))
    once_kalem = once.json()["groups"][0]["items"][0]
    assert Decimal(once_kalem["unallocated_quantity"]) == Decimal("500.000")

    # Kimlikler silmeden ÖNCE okunur: `expire_all` sonrası bir kimliğe dokunmak
    # senkron bağlamda tembel yükleme tetikler (MissingGreenlet).
    site_id, item_id, kat_a_id, kat_b_id = site.id, item.id, kat_a.id, kat_b.id

    silme = await client.delete(f"/sections/{kat_a_id}", headers=_auth(token))
    assert silme.status_code in (204, 200), silme.text
    db_session.expire_all()

    # (a) tahsis satırı gitti
    kalanlar = (
        (
            await db_session.execute(
                select(BoqItemSectionAllocation).where(
                    BoqItemSectionAllocation.boq_item_id == item_id
                )
            )
        )
        .scalars()
        .all()
    )
    assert [r.section_id for r in kalanlar] == [kat_b_id]

    # (b) poz satırı ve kotası dokunulmadı
    poz = (
        await db_session.execute(
            select(BoqItem).where(BoqItem.id == item_id).execution_options(populate_existing=True)
        )
    ).scalar_one_or_none()
    assert poz is not None, "bölüm silme POZU götürdü — veri kaybı"
    assert poz.quantity == Decimal("1200.000")

    # (c) atanmamış miktar tahsis kadar arttı
    sonra = await client.get(f"/sites/{site_id}/boq", headers=_auth(token))
    sonra_kalem = sonra.json()["groups"][0]["items"][0]
    assert Decimal(sonra_kalem["allocated_quantity"]) == Decimal("300.000")
    assert Decimal(sonra_kalem["unallocated_quantity"]) == Decimal("900.000")


async def test_poz_silinince_tahsisleri_de_gider(client, db_session, user_factory, project_factory):
    """`boq_item_id` CASCADE — tahsis satırı pozun ALT PARÇASIDIR."""
    project = await project_factory("BOQSEC-18")
    site = await _site(db_session, project)
    group = await _group(db_session, site)
    item = await _item(db_session, site, group)
    kat = await _section(db_session, site, "Kat 6-10")
    await _allocation(db_session, item, kat, "400.000")
    token = await _login(client, db_session, user_factory, "system_admin", "s18@boqsec.co")

    resp = await client.delete(f"/boq/items/{item.id}", headers=_auth(token))

    assert resp.status_code == 204, resp.text
    db_session.expire_all()
    assert (await _tahsis_sayisi(db_session, item)) == 0


# --- İNVARIANTIN İKİNCİ KAPISI: PATCH ile kotayı düşürmek ------------------


async def test_PATCH_kotayi_tahsis_toplaminin_ALTINA_indiremez(
    client, db_session, user_factory, project_factory
):
    """🔴 Bir kümeyle çalışan bekçi KÜMENİN KENDİSİNİ de sınamalıdır (MT-2 dersi).

    Tahsis ucu invariantı YUKARIDAN sınırlar. AYNI invariant `PATCH /boq/items`
    ile AŞAĞIDAN kırılabilir: 700'ü dağıtılmış bir pozun kotasını 500'e çekmek
    `SUM > quantity` bırakır ve hiçbir uç bunu bir daha fark etmez.
    """
    project = await project_factory("BOQSEC-19")
    site = await _site(db_session, project)
    group = await _group(db_session, site)
    item = await _item(db_session, site, group)
    kat = await _section(db_session, site, "Kat 6-10")
    await _allocation(db_session, item, kat, "700.000")
    token = await _login(client, db_session, user_factory, "system_admin", "s19@boqsec.co")

    dusur = await client.patch(
        f"/boq/items/{item.id}", json={"quantity": "500.000"}, headers=_auth(token)
    )
    tam_sinir = await client.patch(
        f"/boq/items/{item.id}", json={"quantity": "700.000"}, headers=_auth(token)
    )

    assert dusur.status_code == 409, dusur.text
    # 🔴 Sınırda GEÇMELİ: yalnız reddi iddia etmek eşiğin yerini kanıtlamaz.
    assert tam_sinir.status_code == 200, tam_sinir.text
    await db_session.refresh(item)
    assert item.quantity == Decimal("700.000")


async def test_PATCH_kotayi_YUKARI_cekmek_serbest(
    client, db_session, user_factory, project_factory
):
    """Kapı yalnız invariantı koruduğu yönde kapalıdır; kotayı büyütmek serbesttir."""
    project = await project_factory("BOQSEC-20")
    site = await _site(db_session, project)
    group = await _group(db_session, site)
    item = await _item(db_session, site, group)
    kat = await _section(db_session, site, "Kat 6-10")
    await _allocation(db_session, item, kat, "700.000")
    token = await _login(client, db_session, user_factory, "system_admin", "s20@boqsec.co")

    resp = await client.patch(
        f"/boq/items/{item.id}", json={"quantity": "2000.000"}, headers=_auth(token)
    )

    assert resp.status_code == 200, resp.text
    assert Decimal(resp.json()["unallocated_quantity"]) == Decimal("1300.000")


# --- Yardımcı --------------------------------------------------------------


async def _tahsis_sayisi(session, item) -> int:
    rows = (
        (
            await session.execute(
                select(BoqItemSectionAllocation).where(
                    BoqItemSectionAllocation.boq_item_id == item.id
                )
            )
        )
        .scalars()
        .all()
    )
    return len(rows)
