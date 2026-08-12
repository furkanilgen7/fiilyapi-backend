"""SA T4 — teklif karşılaştırma Excel'i (TEK 38 "Excel" düğmesi, §7 S5).

Sütunlar TEK kartlarıyla BİREBİR: Tedarikçi (57) · Birim Fiyat (64) · Toplam
(65) · Teslimat (68) · Garanti (69) · Ödeme (70) · Nakliye (71).

FLOAT-YASAK (B4 dersi, `test_audit_export` emsali): her hücre açıkça `str`
yazılır — openpyxl'in sessiz tip tahminine alan bırakılmaz. GERİ-OKUMA testi
zorunludur: dosya openpyxl ile açılıp hücre hücre doğrulanır.
"""

from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.modules.procurement.export import (
    COLUMN_HEADERS,
    SHEET_TITLE,
    build_quote_comparison_workbook,
)
from app.modules.procurement.models import PaymentTerms
from app.modules.procurement.schemas import PurchaseQuoteCard

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _kart(
    *,
    supplier_name: str = "Demirsan A.Ş.",
    unit_price: str = "21500.00",
    total_cost: str = "322500.00",
    delivery_time: str = "3 iş günü",
    warranty_note: str | None = "TS 708 standart",
    payment_terms: PaymentTerms = PaymentTerms.days_30,
    shipping_included: bool = True,
    shipping_cost: str | None = None,
) -> PurchaseQuoteCard:
    import uuid
    from datetime import UTC, datetime

    return PurchaseQuoteCard(
        id=uuid.uuid4(),
        request_id=uuid.uuid4(),
        supplier_id=uuid.uuid4(),
        unit_price=Decimal(unit_price),
        delivery_time=delivery_time,
        warranty_note=warranty_note,
        payment_terms=payment_terms,
        shipping_included=shipping_included,
        shipping_cost=None if shipping_cost is None else Decimal(shipping_cost),
        is_selected=False,
        created_at=datetime(2026, 7, 14, tzinfo=UTC),
        supplier_name=supplier_name,
        total_cost=Decimal(total_cost),
        is_best_price=False,
    )


def _sheet(kartlar):
    buffer = build_quote_comparison_workbook(kartlar)
    assert isinstance(buffer, BytesIO)
    return load_workbook(buffer).active


def _values(sheet) -> list[list]:
    return [list(satir) for satir in sheet.iter_rows(values_only=True)]


# --- Saf fonksiyon: geri-okuma ---


def test_baslik_satiri_TEK_kartiyla_birebir():
    assert _values(_sheet([_kart()]))[0] == [
        "Tedarikçi",
        "Birim Fiyat",
        "Toplam",
        "Teslimat",
        "Garanti",
        "Ödeme",
        "Nakliye",
    ]
    assert COLUMN_HEADERS[0] == "Tedarikçi"
    assert _sheet([]).title == SHEET_TITLE


def test_bos_karsilastirma_yalnizca_baslik_satiri():
    assert len(_values(_sheet([]))) == 1


def test_satir_degerleri_geri_okundu():
    """TEK 1. kartının BİREBİR karşılığı."""
    assert _values(_sheet([_kart()]))[1] == [
        "Demirsan A.Ş.",
        "₺ 21.500,00",
        "₺ 322.500,00",
        "3 iş günü",
        "TS 708 standart",
        "30 gün vadeli",
        "Dahil",
    ]


def test_nakliye_haric_tutari_yazilir():
    """TEK 90 "Hariç (+₺8.000)" — iki hâl de ayırt edilir."""
    satir = _values(_sheet([_kart(shipping_included=False, shipping_cost="8000.00")]))[1]
    assert satir[6] == "Hariç (+₺ 8.000,00)"


def test_nakliye_harici_tutarsiz_hali():
    assert _values(_sheet([_kart(shipping_included=False)]))[1][6] == "Hariç"


def test_bos_garanti_uzun_tire():
    assert _values(_sheet([_kart(warranty_note=None)]))[1][4] == "—"


@pytest.mark.parametrize(
    ("terms", "etiket"),
    [
        (PaymentTerms.cash, "Peşin"),
        (PaymentTerms.days_15, "15 gün vadeli"),
        (PaymentTerms.days_30, "30 gün vadeli"),
        (PaymentTerms.days_60, "60 gün vadeli"),
    ],
)
def test_odeme_kosulu_TR_etiketi(terms, etiket):
    """Ham enum değeri (`days_30`) dışa aktarılmaz — TEK 70'in dili yazılır."""
    assert _values(_sheet([_kart(payment_terms=terms)]))[1][5] == etiket


def test_tum_hucreler_string_float_yasak():
    kartlar = [_kart(), _kart(shipping_included=False, shipping_cost="8000.00", warranty_note=None)]
    for satir in _sheet(kartlar).iter_rows():
        for hucre in satir:
            assert isinstance(hucre.value, str), (hucre.coordinate, type(hucre.value))


def test_kart_sirasi_korunur():
    kartlar = [_kart(supplier_name="A Ltd."), _kart(supplier_name="B Ltd.")]
    assert [satir[0] for satir in _values(_sheet(kartlar))[1:]] == ["A Ltd.", "B Ltd."]


# --- Uç ---


async def test_uc_xlsx_dondurur(
    client, satinalma_headers, gorunen_proje, talep_fabrikasi, tedarikci_fabrikasi, teklif_fabrikasi
):
    from app.modules.procurement.models import PurchaseRequestStatus

    talep = await talep_fabrikasi(
        gorunen_proje, status=PurchaseRequestStatus.quote_wait, lines=[("15.000", "21500.00")]
    )
    await teklif_fabrikasi(talep, await tedarikci_fabrikasi("Demirsan A.Ş."), unit_price="21500.00")

    yanit = await client.get(
        f"/purchase-requests/{talep.id}/quotes/export.xlsx", headers=satinalma_headers
    )

    assert yanit.status_code == 200, yanit.text
    assert yanit.headers["content-type"] == _XLSX_MIME
    assert talep.request_no in yanit.headers["content-disposition"]
    assert "attachment" in yanit.headers["content-disposition"]

    sheet = load_workbook(BytesIO(yanit.content)).active
    satirlar = _values(sheet)
    assert satirlar[0][0] == "Tedarikçi"
    # Toplam = teklif × talebin toplam miktarı (15) — `order_total_from_quote`.
    assert satirlar[1] == [
        "Demirsan A.Ş.",
        "₺ 21.500,00",
        "₺ 322.500,00",
        "3 iş günü",
        "—",
        "30 gün vadeli",
        "Dahil",
    ]


async def test_gorunmeyen_talebin_excel_i_404(
    client, satinalma_headers, gorunmeyen_proje, talep_fabrikasi
):
    talep = await talep_fabrikasi(gorunmeyen_proje)

    yanit = await client.get(
        f"/purchase-requests/{talep.id}/quotes/export.xlsx", headers=satinalma_headers
    )

    assert yanit.status_code == 404, yanit.text
    assert yanit.json()["detail"] == "Satın alma talebi bulunamadı"


async def test_excel_okuma_izni_ister(client, yetkisiz_headers, gorunen_proje, talep_fabrikasi):
    talep = await talep_fabrikasi(gorunen_proje)

    yanit = await client.get(
        f"/purchase-requests/{talep.id}/quotes/export.xlsx", headers=yetkisiz_headers
    )

    assert yanit.status_code == 403, yanit.text
