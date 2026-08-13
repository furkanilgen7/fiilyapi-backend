"""İK-3 T5 — `GET /payroll/periods/{id}/export` (spec §5, BY 55 "Excel").

Emsal `timesheet/export.py` + `tests/timesheet/test_export.py`: aynı kütüphane
(openpyxl), aynı ikili gövde (`Response` + `Content-Disposition`), aynı
FLOAT-YASAK (her hücre AÇIKÇA `str`) ve aynı okuma kuralı — **indirme bir
OKUMADIR, `record_audit` ÇAĞIRMAZ.**

İçerik BY tablosunun BİREBİR karşılığıdır: dokuz sütun (BY 110-118) + tip bölüm
başlıkları (BY 127/175/243/271) + toplam satırı (BY 298-301).

## 🔴 `null` alan BOŞ basılır — 0 BASILMAZ (S4)

Hesaplanamamış satırın brütü/neti `null`dur ve hücreye HİÇ DOKUNULMAZ.
0 yazmak "ödenecek bir şey yok" yalanı olurdu ve Excel'i açan kişi eksik veriyi
göremezdi. `days` de `null` olabilir: serbest meslekte (S7) ve puantaj kaydı
olmayan personelde (S4.1, T4b'den sonra ÜÇ ücret tipinde de).
"""

from decimal import Decimal
from io import BytesIO
from urllib.parse import quote

import openpyxl
import pytest
from sqlalchemy import func, select

from app.modules.audit.models import AuditLog
from app.modules.payroll.export import (
    COLUMN_HEADERS,
    HEADER_ROW,
    INFO_LABELS,
    SECTION_LABELS,
    SHEET_TITLE,
    STATUS_LABELS,
    TOTAL_LABEL_PREFIX,
    XLSX_MEDIA_TYPE,
    filename,
)
from app.modules.payroll.models import PayrollLineStatus
from app.modules.site_diary.models import WorkerSource

from .conftest import AY, YIL

pytestmark = pytest.mark.asyncio


async def _indir(client, headers, donem):
    resp = await client.get(f"/payroll/periods/{donem.id}/export", headers=headers)
    assert resp.status_code == 200, resp.text
    return resp


def _sayfa(resp):
    return openpyxl.load_workbook(BytesIO(resp.content))[SHEET_TITLE]


def _satir(sheet, row: int) -> list:
    return [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]


def _kisi_satiri(sheet, ad: str) -> list:
    for row in range(HEADER_ROW + 1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == ad:
            return _satir(sheet, row)
    raise AssertionError(f"bordro satırı yok: {ad}")


@pytest.fixture
async def hesaplanmis(client, ik_headers, donem, dort_tip):
    await client.post(f"/payroll/periods/{donem.id}/compute", headers=ik_headers)
    return donem


# --- Kapılar + zarf --------------------------------------------------------


async def test_yetkisiz_rol_indiremez_403(client, yetkisiz_headers, donem):
    resp = await client.get(f"/payroll/periods/{donem.id}/export", headers=yetkisiz_headers)
    assert resp.status_code == 403


async def test_olmayan_donem_404(client, ik_headers, seeded_db):
    import uuid

    resp = await client.get(f"/payroll/periods/{uuid.uuid4()}/export", headers=ik_headers)
    assert resp.status_code == 404


async def test_IKILI_govde_ve_basliklar(client, ik_headers, hesaplanmis):
    """xlsx ikili gövdedir: JSON'a çevrilmez, doğru tip ve ad ile iner."""
    resp = await _indir(client, ik_headers, hesaplanmis)

    assert resp.headers["content-type"].startswith(XLSX_MEDIA_TYPE)
    assert resp.content[:2] == b"PK"  # zip damgası — gerçek bir xlsx
    ad = filename(YIL, AY)
    assert ad == f"bordro-{YIL}-{AY:02d}.xlsx"
    disposition = resp.headers["content-disposition"]
    assert disposition.startswith("attachment;")
    assert quote(ad) in disposition


async def test_indirme_denetim_YAZMAZ(client, ik_headers, hesaplanmis, db_session):
    once = (await db_session.execute(select(func.count()).select_from(AuditLog))).scalar_one()
    await _indir(client, ik_headers, hesaplanmis)
    sonra = (await db_session.execute(select(func.count()).select_from(AuditLog))).scalar_one()
    assert sonra == once


# --- BY tablosu birebir ----------------------------------------------------


async def test_baslik_seridi_ve_sutun_basliklari(client, ik_headers, hesaplanmis):
    """Şerit (BY 52/61/63 + dört kart) + dokuz sütun başlığı (BY 110-118)."""
    sheet = _sayfa(await _indir(client, ik_headers, hesaplanmis))

    for index, etiket in enumerate(INFO_LABELS, start=1):
        assert sheet.cell(row=index, column=1).value == etiket
    assert sheet.cell(row=1, column=2).value == f"{AY:02d}.{YIL}"
    assert _satir(sheet, HEADER_ROW)[: len(COLUMN_HEADERS)] == list(COLUMN_HEADERS)


async def test_tip_bolum_basliklari_BY_127_175_243_271(client, ik_headers, hesaplanmis):
    sheet = _sayfa(await _indir(client, ik_headers, hesaplanmis))
    basliklar = [
        sheet.cell(row=row, column=1).value for row in range(HEADER_ROW + 1, sheet.max_row + 1)
    ]
    for kaynak in (
        WorkerSource.company,
        WorkerSource.subcontractor,
        WorkerSource.freelance,
        WorkerSource.intern,
    ):
        assert SECTION_LABELS[kaynak] in basliklar


async def test_sirket_satiri_dokuz_sutun(client, ik_headers, hesaplanmis):
    """Ayşe Demir: 5 gün × 1.800 → brüt 9.000,00 · kesinti %25,759 · net 6.681,69.

    Sayılar ORANLARDAN türer (S1), BY 133-148'in tutarlarından DEĞİL.
    """
    sheet = _sayfa(await _indir(client, ik_headers, hesaplanmis))
    assert _kisi_satiri(sheet, "Ayşe Demir")[: len(COLUMN_HEADERS)] == [
        "Ayşe Demir",
        "Şirket",
        "5",
        "9000.00",
        "2318.31",
        "6681.69",
        "6681.69",
        "0.00",
        STATUS_LABELS[PayrollLineStatus.pending],
    ]


async def test_serbest_meslekte_GUN_bos_basilir(client, ik_headers, hesaplanmis):
    """S7 — serbest meslekte gün YOKTUR (BY 254 "—"): hücreye DOKUNULMAZ."""
    satir = _kisi_satiri(_sayfa(await _indir(client, ik_headers, hesaplanmis)), "Kemal Tunç")
    assert satir[2] is None
    assert satir[3] == "12500.00"


async def test_S4_hesaplanamamis_satirda_PARA_HUCRELERI_BOS_sifir_DEGIL(
    client, ik_headers, hesaplanmis
):
    """🔴 S4 — `null` para alanı BOŞ iner, **0 BASILMAZ**.

    0 basılsaydı Excel'i açan kişi "bu kişiye ödenecek bir şey yok" diye okurdu;
    oysa ücret tanımı EKSİKTİR. **Gün yine basılır** ve bu bilinçlidir: gün
    puantajdan OKUNAN bir olgudur (`compute._uncomputed`), ücret tanımsız diye
    kaybolması eksik verinin NEREDE olduğunu gizlerdi.
    """
    satir = _kisi_satiri(_sayfa(await _indir(client, ik_headers, hesaplanmis)), "Zeynep Ak")

    assert satir[2] == "5"
    assert satir[3:8] == [None, None, None, None, None]  # brüt/kesinti/net/banka/elden
    assert satir[8] == STATUS_LABELS[PayrollLineStatus.uncomputed]


async def test_toplam_satiri_BY_298_301(client, ik_headers, hesaplanmis):
    """Toplam satırı ÖDEME tabanındandır (`summary.py`): taşeron ve
    hesaplanamamış satır NET toplamına GİRMEZ (K2/S4).

    Etiketteki sayı ise dönemin TÜM satırlarıdır (BY 298 "TOPLAM (48 çalışan)"
    = tfoot 12+29+5+2) — iki sayı AYNI DEĞİLDİR ve tek alana indirgenmez.
    """
    sheet = _sayfa(await _indir(client, ik_headers, hesaplanmis))
    toplam = _satir(sheet, sheet.max_row)

    assert toplam[0] == f"{TOTAL_LABEL_PREFIX} (5 çalışan)"
    odenebilir = Decimal("6681.69") + Decimal("10000.00") + Decimal("7500.00")
    assert toplam[5] == str(odenebilir)
    assert toplam[6] == str(odenebilir)  # üçü de banka (ŞEF KARARI 3)
    assert toplam[7] == "0.00"


async def test_bos_donem_GECERLI_dosya_dondurur(client, ik_headers, donem):
    """Satırsız dönemde de geçerli bir dosya iner — uydurma satır üretilmez."""
    sheet = _sayfa(await _indir(client, ik_headers, donem))
    assert _satir(sheet, HEADER_ROW)[: len(COLUMN_HEADERS)] == list(COLUMN_HEADERS)
    assert sheet.cell(row=HEADER_ROW + 1, column=1).value == f"{TOTAL_LABEL_PREFIX} (0 çalışan)"
