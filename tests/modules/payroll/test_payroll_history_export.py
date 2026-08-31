"""EXPORT-XLSX — `GET /payroll/periods/export.xlsx` (mockup BG "Excel İndir").

🔴 `test_payroll_export.py` ile KARIŞTIRILMAZ: o, TEK dönemin DETAYINI ölçer;
bu dosya BORDRO GEÇMİŞİNİ (dönem listesi) ölçer. İki ayrı ekran, iki ayrı zarf.

## Bu dosyanın ASIL iddiası: dosya = EKRAN

Satırlar sayılmaz, **KÜMELER karşılaştırılır**: liste ucunun JSON'ından
türetilen satır kümesi ile xlsx'ten okunan satır kümesi BİREBİR aynı olmalıdır.

Dönemler ÜRÜNÜN KENDİ UCUNDAN açılır (`POST /payroll/periods`) ve tutarlar
ÜRÜNÜN KENDİ hesabından (`POST /payroll/periods/{id}/compute`) doğar — ORM
satırı elle yazılsaydı "Excel ekranla aynı" iddiası ölçülmemiş olurdu.
"""

import json
from io import BytesIO

import openpyxl
import pytest
from sqlalchemy import func, select

from app.modules.audit.models import AuditLog
from app.modules.payroll.history_export import (
    COLUMN_HEADERS,
    PERIOD_STATUS_LABELS,
    SHEET_TITLE,
    XLSX_MEDIA_TYPE,
    filename,
)
from app.modules.payroll.models import PayrollPeriodStatus

from .conftest import AY, YIL

pytestmark = pytest.mark.asyncio

YOL = "/payroll/periods/export.xlsx"
LISTE = "/payroll/periods"


def _sayfa(resp):
    return openpyxl.load_workbook(BytesIO(resp.content))[SHEET_TITLE]


def _dosya_satirlari(sheet) -> set[tuple]:
    return {
        tuple(sheet.cell(row=row, column=col).value for col in range(1, len(COLUMN_HEADERS) + 1))
        for row in range(2, sheet.max_row + 1)
    }


def _ekran_satiri(item: dict) -> tuple:
    """BG sütun sırasına göre BEKLENEN hücre demeti — export'un `_cells`i
    KULLANILMAZ (test kendi kendini doğrulamamalı)."""
    return (
        f"{item['month']:02d}.{item['year']}",
        str(item["personnel_count"]),
        item["gross_total"],
        item["sgk_employer_total"],
        item["net_total"],
        item["total_cost"],
        item["payment_due_date"],
        PERIOD_STATUS_LABELS[PayrollPeriodStatus(item["status"])],
    )


def _ekran_kumesi(resp) -> set[tuple]:
    """`parse_float=str`: para hücreleri API'nin YAZDIĞI METİNLE karşılaştırılır."""
    return {_ekran_satiri(item) for item in json.loads(resp.text, parse_float=str)["items"]}


async def _donem_ac(client, headers, *, year: int, month: int, due: str | None = None) -> dict:
    govde: dict = {"year": year, "month": month}
    if due is not None:
        govde["payment_due_date"] = due
    resp = await client.post(LISTE, json=govde, headers=headers)
    assert resp.status_code == 201, resp.text
    return resp.json()


@pytest.fixture
async def gecmis(client, ik_headers, oranlar, personel_fabrikasi, puantaj_fabrikasi):
    """Üç dönem — ikisi boş, biri GERÇEKTEN hesaplanmış (para dolu satır).

    Hepsi ürünün kendi uçlarından açılır; tutarlar `compute` ucundan doğar.
    """
    kisi = await personel_fabrikasi("Ayşe Demir")
    await puantaj_fabrikasi(kisi, [1, 2, 3, 4, 5])

    dolu = await _donem_ac(client, ik_headers, year=YIL, month=AY, due=f"{YIL}-07-20")
    hesap = await client.post(f"{LISTE}/{dolu['id']}/compute", headers=ik_headers)
    assert hesap.status_code == 200, hesap.text

    await _donem_ac(client, ik_headers, year=YIL, month=AY - 1, due=f"{YIL}-06-20")
    # Ödeme tarihi GİRİLMEMİŞ dönem: hücreye DOKUNULMAZ (sunucu tarih uydurmaz).
    await _donem_ac(client, ik_headers, year=YIL, month=AY - 2)
    return dolu


# --- Kapılar: POZİTİF KONTROLLÜ ---------------------------------------------


async def test_yetkisiz_rol_indiremez_403(client, yetkisiz_headers, gecmis):
    """`site_chief` (`payroll=_N`) — okumada bile 403."""
    resp = await client.get(YOL, headers=yetkisiz_headers)
    assert resp.status_code == 403


async def test_yetkili_rol_dosyayi_GERCEKTEN_alir(client, ik_headers, gecmis):
    """🔴 POZİTİF KONTROL — uç herkese 403 verecek şekilde bozulursa yakalanır."""
    resp = await client.get(YOL, headers=ik_headers)
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == XLSX_MEDIA_TYPE
    assert _sayfa(resp).max_row == 4  # başlık + üç dönem


async def test_kimliksiz_401(client):
    resp = await client.get(YOL)
    assert resp.status_code == 401


# --- Küme eşitliği: dosya = EKRAN -------------------------------------------


async def test_kume_esitligi(client, ik_headers, gecmis):
    ekran = await client.get(LISTE, params={"limit": 200}, headers=ik_headers)
    assert ekran.status_code == 200, ekran.text
    dosya = await client.get(YOL, headers=ik_headers)
    assert dosya.status_code == 200, dosya.text

    beklenen = _ekran_kumesi(ekran)
    assert len(beklenen) == 3
    assert _dosya_satirlari(_sayfa(dosya)) == beklenen


# --- Sessiz kırpma YOK ------------------------------------------------------


async def test_tavanin_UZERINDE_donem_tamami_iner(client, ik_headers):
    """🔴 204 dönem: liste 200'de KIRPAR, dosya HEPSİNİ yazar.

    `limit=None`ın SQL'e gerçekten ulaştığını ölçer.
    """
    adet = 0
    for yil in range(2000, 2017):
        for ay in range(1, 13):
            await _donem_ac(client, ik_headers, year=yil, month=ay)
            adet += 1
    assert adet == 204

    ekran = await client.get(LISTE, params={"limit": 200}, headers=ik_headers)
    assert ekran.status_code == 200, ekran.text
    assert len(ekran.json()["items"]) == 200
    assert ekran.json()["total"] == adet

    sheet = _sayfa(await client.get(YOL, headers=ik_headers))
    assert sheet.max_row == 1 + adet
    assert len(_dosya_satirlari(sheet)) == adet


async def test_liste_ucunun_tavani_DEGISMEDI(client, ik_headers):
    resp = await client.get(LISTE, params={"limit": 201}, headers=ik_headers)
    assert resp.status_code == 422


# --- Boş küme ---------------------------------------------------------------


async def test_bos_kume_gecerli_dosya(client, ik_headers):
    resp = await client.get(YOL, headers=ik_headers)
    assert resp.status_code == 200, resp.text
    sheet = _sayfa(resp)
    assert [c.value for c in sheet[1]] == list(COLUMN_HEADERS)
    assert sheet.max_row == 1


# --- Para: API metniyle BAYT BİREBİR ----------------------------------------


async def test_para_hucreleri_api_metniyle_birebir(client, ik_headers, gecmis):
    ekran = await client.get(LISTE, params={"limit": 200}, headers=ik_headers)
    item = next(
        satir
        for satir in json.loads(ekran.text, parse_float=str)["items"]
        if satir["id"] == gecmis["id"]
    )
    assert item["gross_total"] != "0.00", "hesaplanmamış dönem — test hiçbir şey ölçmüyor"

    sheet = _sayfa(await client.get(YOL, headers=ik_headers))
    satir = next(s for s in _dosya_satirlari(sheet) if s[0] == f"{AY:02d}.{YIL}")
    for sutun, alan in (
        ("Brüt Maaş", "gross_total"),
        ("SGK İşveren", "sgk_employer_total"),
        ("Net Ödenen", "net_total"),
        ("Toplam Maliyet", "total_cost"),
    ):
        hucre = satir[COLUMN_HEADERS.index(sutun)]
        assert hucre == item[alan], sutun
        assert isinstance(hucre, str), f"{sutun} SAYI olarak yazılmış (FLOAT-YASAK)"


async def test_odeme_tarihi_yoksa_hucreye_dokunulmaz(client, ik_headers, gecmis):
    sheet = _sayfa(await client.get(YOL, headers=ik_headers))
    satir = next(s for s in _dosya_satirlari(sheet) if s[0] == f"{AY - 2:02d}.{YIL}")
    assert satir[COLUMN_HEADERS.index("Ödeme Tarihi")] is None


# --- Zarf -------------------------------------------------------------------


async def test_mime_ve_content_disposition(client, ik_headers, gecmis):
    resp = await client.get(YOL, headers=ik_headers)
    assert resp.headers["content-type"] == XLSX_MEDIA_TYPE
    disposition = resp.headers["content-disposition"]
    assert disposition.startswith("attachment; ")
    assert filename() in disposition
    assert filename() != "bordro-2026-07.xlsx"  # dönem detayıyla ÇAKIŞMAZ


async def test_okuma_denetim_kaydi_URETMEZ(client, ik_headers, gecmis, db_session):
    once = (await db_session.execute(select(func.count()).select_from(AuditLog))).scalar_one()
    resp = await client.get(YOL, headers=ik_headers)
    assert resp.status_code == 200
    sonra = (await db_session.execute(select(func.count()).select_from(AuditLog))).scalar_one()
    assert sonra == once
