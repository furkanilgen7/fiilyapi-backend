"""EXPORT-XLSX — `GET /personnel/export.xlsx` (mockup PE "Dışa Aktar").

Emsal `tests/modules/test_audit_export.py` + `tests/modules/units/test_units_export.py`:
aynı kütüphane (openpyxl), aynı ikili gövde (`Response` + `Content-Disposition`),
aynı okuma kuralı — **indirme bir OKUMADIR, `record_audit` ÇAĞIRMAZ.**

## Bu dosyanın ASIL iddiası: dosya = EKRAN

Satırlar sayılmaz, **KÜMELER karşılaştırılır** (`test_kume_esitligi_*`): liste
ucunun JSON'ından türetilen satır kümesi ile xlsx'ten okunan satır kümesi
BİREBİR aynı olmalıdır. Sayı karşılaştırması, süzgeci farklı uygulayan bir
export'u YAKALAYAMAZDI (aynı sayıda ama başka kayıtlar).

Fixture'lar ÜRÜNÜN KENDİ YOLUNDAN kurulur (`POST /personnel`): ORM satırı elle
yazılsaydı "Excel ekranla aynı" iddiası ölçülmemiş olurdu — ekranın gördüğü
kayıt ile testin yazdığı kayıt aynı doğrulama/normalizasyondan geçmezdi.

## 🔴 PII negatif bekçisi

`PersonnelResponse` TCKN/IBAN/telefon/adres/e-posta taşır; PE tablosu bunların
HİÇBİRİNİ göstermez. `test_pii_alanlari_dosyaya_girmez` bu alanların
DEĞERLERİNİ dosyanın TÜM hücrelerinde arar. Sonraki bir dilim "bir alan daha
ekleyeyim" derse test kırmızıya döner — sütun listesine bakan bir test bunu
göremezdi.
"""

import json
from io import BytesIO

import openpyxl
import pytest
from sqlalchemy import func, select

from app.modules.audit.models import AuditLog
from app.modules.personnel.export import (
    ACTIVE_LABEL,
    COLUMN_HEADERS,
    INACTIVE_LABEL,
    SHEET_TITLE,
    SOURCE_LABELS,
    XLSX_MEDIA_TYPE,
    filename,
)
from app.modules.site_diary.models import WorkerSource

pytestmark = pytest.mark.asyncio

YOL = "/personnel/export.xlsx"
LISTE = "/personnel"

#: Checksum'ı GEÇERLİ bir TCKN (`guards.validate_tckn`): 12345678950.
GECERLI_TCKN = "12345678950"
#: mod-97'si GEÇERLİ bir TR IBAN (`app/core/iban.py`).
GECERLI_IBAN = "TR330006100519786457841326"

#: Dosyada ASLA görünmemesi gereken DEĞERLER (sütun adları değil — değerler).
PII_DEGERLERI = (
    GECERLI_TCKN,
    GECERLI_IBAN,
    "05551112233",
    "Atatürk Cad. No:5 Kadıköy",
    "gizli@personel.co",
    "Acil Kişi Zeynep",
    "05559998877",
)


def _sayfa(resp):
    return openpyxl.load_workbook(BytesIO(resp.content))[SHEET_TITLE]


def _dosya_satirlari(sheet) -> set[tuple]:
    """Başlık DIŞINDAKİ satırların kümesi (sıra değil, KÜME karşılaştırılır)."""
    return {
        tuple(sheet.cell(row=row, column=col).value for col in range(1, len(COLUMN_HEADERS) + 1))
        for row in range(2, sheet.max_row + 1)
    }


def _ekran_satiri(item: dict) -> tuple:
    """Liste ucunun JSON'ından PE sütun sırasına göre BEKLENEN hücre demeti.

    Etiket sözlüğü ürün modülünden gelir (elle yazılsaydı iki metin ayrışırdı)
    ama SÜTUN SEÇİMİ burada AÇIKÇA yazılır: export'un kendi `_cells`i
    kullanılsaydı test kendi kendini doğrular, hiçbir şey bekçilemezdi.
    """
    return (
        item["full_name"],
        item["hire_date"],
        SOURCE_LABELS[WorkerSource(item["source"])],
        item["trade"],
        item["sgk_no"],
        item["wage_amount"],
        ACTIVE_LABEL if item["is_active"] else INACTIVE_LABEL,
    )


def _ekran_kumesi(resp) -> set[tuple]:
    """`parse_float=str`: para hücresi API'nin YAZDIĞI METİNLE karşılaştırılır.

    `json.loads` varsayılanı `Decimal`i `float`a çevirip "1200.00"ı 1200.0
    yapardı ve bayt-birebirlik iddiası ÖLÇÜLMEMİŞ olurdu.
    """
    return {_ekran_satiri(item) for item in json.loads(resp.text, parse_float=str)["items"]}


async def _ekle(client, headers, **govde) -> dict:
    resp = await client.post(LISTE, json=govde, headers=headers)
    assert resp.status_code == 201, resp.text
    return resp.json()


@pytest.fixture
async def kadro(client, ik_headers, taseron):
    """Üç personel — ÜRÜNÜN KENDİ UCUNDAN (`POST /personnel`) yazılır."""
    return [
        await _ekle(
            client,
            ik_headers,
            full_name="Mehmet Yılmaz",
            source=WorkerSource.company.value,
            trade="Kalıpçı Usta",
            sgk_no="12345678900",
            hire_date="2025-03-01",
            wage_amount="1200.00",
            is_active=True,
        ),
        await _ekle(
            client,
            ik_headers,
            full_name="Ali Kaya",
            source=WorkerSource.subcontractor.value,
            subcontractor_id=str(taseron.id),
            trade="Demir Ustası",
            sgk_no="23456789011",
            hire_date="2025-04-15",
            wage_amount="1100.50",
            is_active=False,
        ),
        # Boş alanlı satır: hücreye DOKUNULMAZ (None kalır), "" ya da 0 YAZILMAZ.
        await _ekle(
            client,
            ik_headers,
            full_name="Hasan Çelik",
            source=WorkerSource.freelance.value,
        ),
    ]


# --- Kapılar: POZİTİF KONTROLLÜ ---------------------------------------------


async def test_yetkisiz_rol_indiremez_403(client, yetkisiz_headers, kadro):
    """`procurement` (`personnel=_N`) — okumada bile 403."""
    resp = await client.get(YOL, headers=yetkisiz_headers)
    assert resp.status_code == 403


async def test_view_seviyesi_dosyayi_GERCEKTEN_alir(client, sef_headers, kadro):
    """🔴 POZİTİF KONTROL — 403 testinin tek başına anlamı YOKTUR.

    Uç kazara herkese 403 verecek şekilde bozulsaydı yukarıdaki test YEŞİL
    kalırdı. Ekranı görebilen rol (`site_chief`, `personnel=_V`) dosyayı da
    GERÇEKTEN almalı ve dosya AÇILABİLİR olmalıdır.
    """
    resp = await client.get(YOL, headers=sef_headers)
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == XLSX_MEDIA_TYPE
    sheet = _sayfa(resp)
    assert sheet.max_row == 1 + len(kadro)


async def test_kimliksiz_401(client):
    resp = await client.get(YOL)
    assert resp.status_code == 401


# --- Küme eşitliği: dosya = EKRAN -------------------------------------------


async def test_kume_esitligi_suzgecsiz(client, ik_headers, kadro):
    ekran = await client.get(LISTE, params={"limit": 200}, headers=ik_headers)
    assert ekran.status_code == 200, ekran.text
    dosya = await client.get(YOL, headers=ik_headers)
    assert dosya.status_code == 200, dosya.text

    assert _dosya_satirlari(_sayfa(dosya)) == _ekran_kumesi(ekran)


@pytest.mark.parametrize(
    "params",
    [
        {"source": WorkerSource.company.value},
        {"is_active": False},
        {"q": "Çelik"},
        {"is_draft": True},
    ],
    ids=["source", "is_active", "q", "is_draft"],
)
async def test_kume_esitligi_suzgecli(client, ik_headers, kadro, params):
    """Aynı süzgeç ekranda ve dosyada AYNI kümeyi vermelidir.

    Süzgeç kümesi tek `PersonnelFilters` dataclass'ındadır; bu test onun
    ekrana ve dosyaya AYNI şekilde bağlandığını ölçer.
    """
    ekran = await client.get(LISTE, params={**params, "limit": 200}, headers=ik_headers)
    assert ekran.status_code == 200, ekran.text
    dosya = await client.get(YOL, params=params, headers=ik_headers)
    assert dosya.status_code == 200, dosya.text

    beklenen = _ekran_kumesi(ekran)
    assert beklenen, f"süzgeç {params} boş küme verdi — test hiçbir şey ölçmüyor"
    assert _dosya_satirlari(_sayfa(dosya)) == beklenen


# --- Sessiz kırpma YOK ------------------------------------------------------


async def test_tavanin_UZERINDE_kayit_tamami_iner(client, ik_headers):
    """🔴 205 kayıt: liste 200'de KIRPAR, dosya HEPSİNİ yazar.

    `limit=None`ın SQL'e gerçekten ulaştığını ölçer. Repository'de
    `if limit is not None` dalı 200'e sabitlenseydi bu test kırmızı olurdu.
    """
    adet = 205
    for sira in range(adet):
        await _ekle(
            client,
            ik_headers,
            full_name=f"Tavan Personel {sira:03d}",
            source=WorkerSource.company.value,
        )

    ekran = await client.get(LISTE, params={"limit": 200}, headers=ik_headers)
    assert ekran.status_code == 200, ekran.text
    assert len(ekran.json()["items"]) == 200
    assert ekran.json()["total"] == adet

    sheet = _sayfa(await client.get(YOL, headers=ik_headers))
    assert sheet.max_row == 1 + adet
    assert len(_dosya_satirlari(sheet)) == adet


async def test_liste_ucunun_tavani_DEGISMEDI(client, ik_headers):
    """Liste ucu hâlâ 200'de kapalıdır — `limit=None` yalnız export'un hakkı."""
    resp = await client.get(LISTE, params={"limit": 201}, headers=ik_headers)
    assert resp.status_code == 422


# --- Boş küme ---------------------------------------------------------------


async def test_bos_kume_gecerli_dosya(client, ik_headers):
    resp = await client.get(YOL, params={"q": "hicbir-eslesme-yok"}, headers=ik_headers)
    assert resp.status_code == 200, resp.text
    sheet = _sayfa(resp)
    assert [c.value for c in sheet[1]] == list(COLUMN_HEADERS)
    assert sheet.max_row == 1


# --- Para: API metniyle BAYT BİREBİR ----------------------------------------


async def test_para_hucresi_api_metniyle_birebir(client, ik_headers, kadro):
    """`1100.50` ekranda neyse dosyada da ODUR — yeniden yuvarlama YOK."""
    ekran = await client.get(LISTE, params={"q": "Ali Kaya"}, headers=ik_headers)
    metin = json.loads(ekran.text, parse_float=str)["items"][0]["wage_amount"]

    sheet = _sayfa(await client.get(YOL, params={"q": "Ali Kaya"}, headers=ik_headers))
    hucre = sheet.cell(row=2, column=COLUMN_HEADERS.index("Ücret/Gün") + 1).value
    assert hucre == metin
    assert isinstance(hucre, str), "para hücresi SAYI olarak yazılmış (FLOAT-YASAK)"


async def test_bos_alan_hucresine_dokunulmaz(client, ik_headers, kadro):
    """Meslek/SGK/ücret boşsa hücre `None` kalır — `""` de `0` da YAZILMAZ."""
    sheet = _sayfa(await client.get(YOL, params={"q": "Hasan"}, headers=ik_headers))
    satir = next(iter(_dosya_satirlari(sheet)))
    assert satir[0] == "Hasan Çelik"
    for sutun in ("İşe giriş", "Meslek", "SGK", "Ücret/Gün"):
        assert satir[COLUMN_HEADERS.index(sutun)] is None


# --- 🔴 PII negatif bekçisi --------------------------------------------------


async def test_pii_alanlari_dosyaya_girmez(client, ik_headers):
    """TCKN · IBAN · telefon · adres · e-posta · acil kişi HİÇBİR hücrede yok.

    Değerler API zarfında VARDIR (aşağıda ölçülür) — yani bekçi "alan boştu"
    diye değil, export onları BASMADIĞI için yeşildir.
    """
    olusan = await _ekle(
        client,
        ik_headers,
        full_name="Gizli Kişi",
        source=WorkerSource.company.value,
        tc_no=GECERLI_TCKN,
        iban=GECERLI_IBAN,
        phone="05551112233",
        address="Atatürk Cad. No:5 Kadıköy",
        email="gizli@personel.co",
        emergency_contact_name="Acil Kişi Zeynep",
        emergency_contact_phone="05559998877",
    )
    # Pozitif yarı: veriler GERÇEKTEN kayıtlı ve zarfta görünüyor.
    for alan, deger in (
        ("tc_no", GECERLI_TCKN),
        ("iban", GECERLI_IBAN),
        ("phone", "05551112233"),
        ("address", "Atatürk Cad. No:5 Kadıköy"),
        ("email", "gizli@personel.co"),
    ):
        assert olusan[alan] == deger

    sheet = _sayfa(await client.get(YOL, params={"q": "Gizli"}, headers=ik_headers))
    hucreler = {
        str(sheet.cell(row=row, column=col).value)
        for row in range(1, sheet.max_row + 1)
        for col in range(1, sheet.max_column + 1)
    }
    for deger in PII_DEGERLERI:
        assert deger not in hucreler, f"PII sızıntısı: {deger} dosyada"


# --- Zarf -------------------------------------------------------------------


async def test_mime_ve_content_disposition(client, ik_headers, kadro):
    resp = await client.get(YOL, headers=ik_headers)
    assert resp.headers["content-type"] == XLSX_MEDIA_TYPE
    disposition = resp.headers["content-disposition"]
    assert disposition.startswith("attachment; ")
    assert filename() in disposition


async def test_okuma_denetim_kaydi_URETMEZ(client, ik_headers, kadro, seeded_db):
    """İndirme bir OKUMADIR (`units/router.py` P4 T7)."""
    once = (await seeded_db.execute(select(func.count()).select_from(AuditLog))).scalar_one()
    resp = await client.get(YOL, headers=ik_headers)
    assert resp.status_code == 200
    sonra = (await seeded_db.execute(select(func.count()).select_from(AuditLog))).scalar_one()
    assert sonra == once
