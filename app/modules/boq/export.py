"""BOQ (İş Kalemleri) Excel çıktısı (plan T8, spec §5.3).

Saf sunum katmani: `Request`/`Response` bilmez, DB'ye dokunmaz — yalnizca
`BoqListResponse` (servis katmaninin urettigi zarf) calisma kitabina cevirir.

Sutun basliklari mockup (`Ekran 13 - İş Kalemleri.dc.html` satir 96-102) ile
BIREBIR. Grup basliklari ("1. TOPRAK VE TEMEL İŞLERİ") ve GENEL TOPLAM satiri
basilir; "Gerç. %" sutunu basliktaki yerinde durur ama hucreleri BOS kalir
(veri P7'de yazilir — spec §3.2/§5.3, zarif dusus, sessiz atlama degil).

FLOAT-YASAK (B4 dersi, audit/export.py emsali): her hucre ACIKCA `str` yazilir.
Servis katmani zaten Decimal degerleri quantize edip API yanitiyla AYNI metin
temsiline (ornegin "1240.000", "347200.00") cevirdigi icin burada TEKRAR
yuvarlama/bicimlendirme YAPILMAZ — API ile Excel her zaman ayni degeri gosterir.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.modules.boq.schemas import BoqGroupResponse, BoqItemResponse, BoqListResponse

SHEET_TITLE = "İş Kalemleri"

#: Mockup tablosunun basliklari — sira ve metin degistirilmez (spec §5.3).
COLUMN_HEADERS: tuple[str, ...] = (
    "Poz No",
    "İş Kalemi Tarifi",
    "Birim",
    "Miktar",
    "Birim Fiyat",
    "Tutar",
    "Gerç. %",
)

GRAND_TOTAL_LABEL = "GENEL TOPLAM"

_COLUMN_COUNT = len(COLUMN_HEADERS)
_TUTAR_COLUMN = 6
_COLUMN_WIDTHS = (12, 42, 10, 14, 14, 16, 12)


def _group_title(index: int, group: BoqGroupResponse) -> str:
    """Grup adinin bastaki sira numarasi SAKLANMAZ (T1 modeli) — burada
    listedeki sirasindan (1'den baslayan sayac) turetilir, mockup'taki
    "1. TOPRAK VE TEMEL İŞLERİ" bicimiyle birebir."""
    return f"{index}. {group.name}"


def _item_row(item: BoqItemResponse) -> tuple[str, ...]:
    """ "Gerç. %" (7. sutun) icin deger DONDURULMEZ — hucre yazilmadan bos

    birakilir (spec §5.3: openpyxl'de acikca `""` yazmak geri okumada `None`'a
    donusuyor, bu yuzden hucreye hic dokunulmamak "bos" niyetini net anlatir).
    """
    return (
        str(item.code),
        str(item.description),
        str(item.unit),
        str(item.quantity),
        str(item.unit_price),
        str(item.amount),
    )


def _write_row(sheet: Worksheet, row: int, values: tuple[str, ...]) -> None:
    for column, value in enumerate(values, start=1):
        # Deger her zaman `str`; openpyxl'in tip tahminine alan birakilmaz.
        sheet.cell(row=row, column=column).value = value


def _write_group_header(sheet: Worksheet, row: int, index: int, group: BoqGroupResponse) -> None:
    sheet.cell(row=row, column=1).value = _group_title(index, group)
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_COLUMN_COUNT)


def _write_grand_total(sheet: Worksheet, row: int, boq: BoqListResponse) -> None:
    sheet.cell(row=row, column=1).value = GRAND_TOTAL_LABEL
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_TUTAR_COLUMN - 1)
    sheet.cell(row=row, column=_TUTAR_COLUMN).value = str(boq.totals.grand_total)
    # Gerç. % (sutun 7) icin deger DONDURULMEZ — hucre bos birakilir (spec §5.3).


def _apply_layout(sheet: Worksheet) -> None:
    for index, width in enumerate(_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width
    sheet.freeze_panes = "A2"


def build_boq_workbook(boq: BoqListResponse) -> BytesIO:
    """BOQ zarfindan xlsx calisma kitabi uretir ve bellekteki tamponu doner.

    Bos BOQ'da bile gecerli bir dosya doner: baslik satiri + "0.00" GENEL TOPLAM.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_TITLE

    _write_row(sheet, 1, COLUMN_HEADERS)

    row = 2
    for group_index, group in enumerate(boq.groups, start=1):
        _write_group_header(sheet, row, group_index, group)
        row += 1
        for item in group.items:
            _write_row(sheet, row, _item_row(item))
            row += 1

    _write_grand_total(sheet, row, boq)
    _apply_layout(sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
