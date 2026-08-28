"""Puantaj Excel çıktısı (T4, spec §3) — mockup ŞP `Şantiye - Puantaj.dc.html`.

`boq/export.py` ve `audit/export.py` deseninin kardeşi: saf SUNUM katmanı.
`Request`/`Response` bilmez, DB'ye dokunmaz, kapsam/yetki kararı VERMEZ — yalnız
`TimesheetMatrix` zarfını (T3'ün ürettiği matris) çalışma kitabına çevirir.

## Hiçbir toplam BURADA hesaplanmaz

Kişi saat toplamı, adam-günü ve günlük toplamlar `matrix.py`
tarafından hesaplanmış olarak GELİR. Excel kendi toplamını kurarsa ekran ile
dosya zamanla ayrışır ve hangisinin doğru olduğu tartışılır hâle gelir; bu yüzden
burada yalnız BİÇİMLENDİRME vardır.

## FLOAT-YASAK (B4 dersi, `audit/export.py` emsali)

Her hücre AÇIKÇA `str` yazılır: openpyxl bir hücreye `int`/`Decimal`/`date`
verildiğinde onu sessizce sayı+biçim olarak saklar ve geri okumada tip değişir.

## Boş gün hücresine DOKUNULMAZ

Kaydı olmayan gün için hücreye `""` yazmak openpyxl'de geri okumada `None`'a
dönüşür (boq §5.3 dersi); niyeti net anlatan davranış hücreye HİÇ dokunmamaktır.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.modules.site_diary.models import WorkerSource
from app.modules.timesheet.models import TimesheetCode
from app.modules.timesheet.schemas import (
    TimesheetCell,
    TimesheetDayTotal,
    TimesheetMatrix,
    TimesheetMatrixRow,
)

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

SHEET_TITLE = "Puantaj"

#: Başlık şeridi (ŞP 116-119) — etiket/değer çiftleri, tablodan ÖNCE gelir.
INFO_LABELS: tuple[str, ...] = (
    "Proje",
    "Şantiye",
    "Dönem",
    "Bölüm",
    "İşçi Sayısı",
    "Toplam Saat",
    "Toplam Adam-Gün",
)

#: Kişi sütunları (ŞP 148-150 / 168-170). Meslek ile taşeron firma AYRI
#: sütunlardır: mockup'taki "Demir Ustası — Akın İnşaat" bir SUNUM kararıdır.
COLUMN_HEADERS: tuple[str, ...] = ("Ad Soyad", "Meslek", "Tür", "Taşeron Firma")

HOURS_TOTAL_HEADER = "Toplam Saat"

TOTAL_HEADER = "Adam-Gün"

DAY_TOTAL_LABEL = "Günlük Toplam"

#: Kodlu (saatsiz) hücrenin harfleri; ham enum değeri yazılmaz.
#:
#: 🔴 PUAN-SAAT: `Ç` ve `FM` KALKTI — çalışılan gün artık hücreye SAAT olarak
#: yazılır (E5 236), harf değil. Geriye yalnız "çalışılmadı ama sebebi var"
#: kodları kaldı.
CODE_LABELS: dict[TimesheetCode, str] = {
    TimesheetCode.leave: "İ",
    TimesheetCode.holiday: "T",
    TimesheetCode.temporary_duty: "G",
}

#: ŞP 150/170 rozetleri + BY 253 "Serbest" / BY 281 "Stajyer".
#:
#: Son iki etiket İK-3'te EKLENDİ: bordro dilimi paylaşılan `worker_source` DB
#: tipine `freelance` + `intern` değerlerini takas etti (İK-3 spec S10) ve bu
#: sözlük üç etiketli kaldığı sürece o kaynaklı bir personel puantaj export'una
#: girdiği anda `KeyError` → **500** olurdu.
SOURCE_LABELS: dict[WorkerSource, str] = {
    WorkerSource.company: "Şirket",
    WorkerSource.subcontractor: "Taşeron",
    WorkerSource.general: "Genel",
    WorkerSource.freelance: "Serbest",
    WorkerSource.intern: "Stajyer",
}

#: Tanınmayan enum değerinin GÖRÜNÜR düşüşü. Boş string ya da "Genel" gibi bir
#: düşüş, tanınmayan kaynağı tanınmış gibi gösterirdi (WORKFLOW §3: sessiz
#: düşüş yok) — ham değer soru işaretiyle basılır, okuyan eksik etiketi fark eder.
UNKNOWN_LABEL_PREFIX = "?"


def source_label(source: WorkerSource) -> str:
    """Kaynak rozetinin metni — SÖZLÜK DOĞRUDAN İNDEKSLENMEZ.

    `SOURCE_LABELS[row.source]` yazımı, paylaşılan enum her genişlediğinde
    (İK-3 iki değer ekledi) export'u sessizce 500'e düşürüyordu. Erişim
    dayanıklıdır ama düşüş SESSİZ DEĞİLDİR: etiket eklemeyi unutan sonraki
    dilim hücrede `?yeni_deger` görür.
    """
    return SOURCE_LABELS.get(source, f"{UNKNOWN_LABEL_PREFIX}{source.value}")


#: Boş metin hücresi (audit/export.py emsali).
EMPTY_VALUE = "—"

#: Tablo başlık satırının numarası: şerit + bir boş ayraç satırı.
HEADER_ROW = len(INFO_LABELS) + 2

_PERSON_COLUMN_COUNT = len(COLUMN_HEADERS)
_PERSON_COLUMN_WIDTHS = (26, 20, 12, 22)
_DAY_COLUMN_WIDTH = 5
_TOTAL_COLUMN_WIDTH = 10


def _period_label(matrix: TimesheetMatrix) -> str:
    return f"{matrix.month:02d}.{matrix.year}"


def _info_values(matrix: TimesheetMatrix) -> tuple[str, ...]:
    return (
        matrix.project_name,
        matrix.site_name,
        _period_label(matrix),
        # Bölüm seçilmemişse şerit bir bölüm adı İDDİA ETMEZ (matris ile aynı kural).
        matrix.section_name or EMPTY_VALUE,
        str(matrix.worker_count),
        str(matrix.total_hours),
        str(matrix.total_man_days),
    )


def _day_total_label(total: TimesheetDayTotal) -> str:
    """Gün sütununun ayak hücresi — o günün SAAT toplamı + `G` işareti.

    Sayı `matrix.py`den GELİR (girilmiş saatlerin toplamı); `G` yalnızca bir
    İŞARETTİR, sayıyı değiştirmez — geçici görev saate 0 katar ama sütunda
    kimsenin bulunmadığı izlenimini vermemelidir.
    """
    return f"{total.total_hours}" + ("G" if total.temporary_duty_count else "")


def _cell_label(cell: TimesheetCell) -> str:
    """Hücrenin metni: SAAT ya da kod harfi — hücrenin sözleşmesi gereği tam biri.

    `CODE_LABELS` doğrudan indekslenmez: paylaşılan enum genişlerse export
    sessizce 500 olurdu (`source_label` ile aynı ders).
    """
    if cell.hours is not None:
        return str(cell.hours)
    if cell.code is None:  # pragma: no cover - DB CHECK'i bunu imkansiz kilar
        return EMPTY_VALUE
    return CODE_LABELS.get(cell.code, f"{UNKNOWN_LABEL_PREFIX}{cell.code.value}")


def _person_cells(row: TimesheetMatrixRow) -> tuple[str, ...]:
    return (
        row.full_name,
        row.trade or EMPTY_VALUE,
        source_label(row.source),
        row.subcontractor_name or EMPTY_VALUE,
    )


def _write_row(sheet: Worksheet, row: int, values: tuple[str | None, ...]) -> None:
    for column, value in enumerate(values, start=1):
        if value is None:
            # Kaydı olmayan güne DOKUNULMAZ (docstring: boş niyeti böyle anlatılır).
            continue
        sheet.cell(row=row, column=column).value = value


def _write_info(sheet: Worksheet, matrix: TimesheetMatrix) -> None:
    for index, (label, value) in enumerate(
        zip(INFO_LABELS, _info_values(matrix), strict=True), start=1
    ):
        sheet.cell(row=index, column=1).value = label
        sheet.cell(row=index, column=2).value = value


def _write_header(sheet: Worksheet, days: list[TimesheetDayTotal]) -> None:
    _write_row(
        sheet,
        HEADER_ROW,
        COLUMN_HEADERS
        + tuple(str(total.work_date.day) for total in days)
        + (HOURS_TOTAL_HEADER, TOTAL_HEADER),
    )


def _write_person(
    sheet: Worksheet, row_number: int, row: TimesheetMatrixRow, days: list[TimesheetDayTotal]
) -> None:
    values = {cell.work_date: _cell_label(cell) for cell in row.cells}
    _write_row(
        sheet,
        row_number,
        _person_cells(row)
        + tuple(values.get(total.work_date) for total in days)
        + (str(row.total_hours), str(row.man_days)),
    )


def _write_day_totals(
    sheet: Worksheet, row_number: int, matrix: TimesheetMatrix, days: list[TimesheetDayTotal]
) -> None:
    _write_row(
        sheet,
        row_number,
        (DAY_TOTAL_LABEL,)
        + (None,) * (_PERSON_COLUMN_COUNT - 1)
        + tuple(_day_total_label(total) for total in days)
        + (str(matrix.total_hours), str(matrix.total_man_days)),
    )


def _apply_layout(sheet: Worksheet, day_count: int) -> None:
    widths = (
        _PERSON_COLUMN_WIDTHS
        + (_DAY_COLUMN_WIDTH,) * day_count
        + (_TOTAL_COLUMN_WIDTH, _TOTAL_COLUMN_WIDTH)
    )
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[
            sheet.cell(row=HEADER_ROW, column=index).column_letter
        ].width = width
    sheet.freeze_panes = sheet.cell(row=HEADER_ROW + 1, column=_PERSON_COLUMN_COUNT + 1).coordinate


def build_timesheet_workbook(matrix: TimesheetMatrix) -> BytesIO:
    """Matris zarfından xlsx çalışma kitabı üretir ve bellekteki tamponu döner.

    Kayıtsız dönemde de GEÇERLİ bir dosya döner: gün iskeleti takvimden gelir
    (`day_totals`), kişi satırı yoktur ve alt toplam satırı sıfırlarla basılır.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_TITLE

    days = list(matrix.day_totals)
    _write_info(sheet, matrix)
    _write_header(sheet, days)

    row_number = HEADER_ROW
    for row in matrix.rows:
        row_number += 1
        _write_person(sheet, row_number, row, days)

    _write_day_totals(sheet, row_number + 1, matrix, days)
    _apply_layout(sheet, len(days))

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def filename(site_code: str, year: int, month: int) -> str:
    """`puantaj-{şantiye kodu}-{yıl}-{ay}.xlsx` — dönem dosya adında GÖRÜNÜR.

    Aynı şantiyenin iki ayı aynı klasöre indiğinde dosyalar birbirini ezmemeli.
    """
    return f"puantaj-{site_code}-{year}-{month:02d}.xlsx"
