"""Personel kartoteksi Excel çıktısı (EXPORT-XLSX) — mockup PE "Dışa Aktar".

`units/export.py` · `payroll/export.py` · `audit/export.py` deseninin kardeşi:
saf SUNUM katmanı. `Request`/`Response` bilmez, DB'ye DOKUNMAZ, kapsam/yetki
kararı VERMEZ — yalnız liste ucunun ürettiği `PersonnelResponse` satırlarını
çalışma kitabına çevirir.

## 🔴 SÜTUN KÜMESİ EKRANI AŞMAZ — bu bir VERİ SIZINTISI kapısıdır

`PersonnelResponse` **TCKN** (`tc_no`), **IBAN**, **telefon**, **adres**,
acil durum kişisi ve e-posta taşır. `Personel.dc.html` tablosunun HİÇBİRİ
bu alanları göstermez (thead: Ad Soyad · Tür · Meslek · Proje · SGK ·
Ücret/Gün · Durum). Excel API zarfının TAMAMINI bassaydı, ekranda TCKN'yi
göremeyen bir `personnel=view` kullanıcısı (şantiye şefi, saha mühendisi) tek
tıkla bütün kadronun kimlik ve banka bilgisini indirebilirdi — yani dosya,
ekranın kendi sınırını atlayan bir yan kapı olurdu.

Bu yüzden sütun kümesi mockup'ın thead'inden TÜRETİLİR, şemadan değil; ve
`tests/personnel/test_personnel_export.py` bunu NEGATİF bir bekçiyle sabitler
(dışlanan alanların DEĞERLERİ dosyada hiçbir hücrede geçmez). Sonraki bir
dilim "bir alan daha ekleyeyim" derse test kırmızıya döner.

## "Proje" sütunu YOKTUR — ölçülmüş sınır

Mockup'ın 4. sütunu proje ADInı gösterir; `PersonnelResponse` yalnız
`assigned_project_id` (UUID) taşır. Ham UUID basmak kullanıcıya OKUNMAZ bir
kolon verirdi; adı çözmek ise ekranın kullanmadığı İKİNCİ bir sorgu yolu açardı
(`equipment` work-logs ölçümüyle aynı gerekçe: ikinci yol, dosya ile ekranın
zamanla ayrışmasının kaynağıdır). Sütun bilinçli olarak DIŞARIDA bırakıldı —
"ekrandan AZ göstermek" meşrudur, "FAZLA göstermek" değildir.

## "İşe giriş" ayrı sütundur

Mockup'ta bu değer "Ad Soyad" hücresinin ALT SATIRIDIR (`İşe giriş:
01.03.2025`). Excel hücresinde iki satırlık kompozit metin süzülemez/sıralanamaz
olurdu; aynı OLGU kendi sütununa alındı. Ekranda GÖRÜNEN bir alandır, yani
kapsam genişlemesi değildir.

## FLOAT-YASAK (B4 dersi, `audit/export.py` emsali)

Her hücre AÇIKÇA `str` yazılır: openpyxl bir hücreye `Decimal`/`date` verilirse
onu sessizce sayı+biçim olarak saklar. `str(Decimal)` API yanıtının metniyle
BİREBİR aynıdır — yeniden yuvarlama YOK.

## `null` alana DOKUNULMAZ (`payroll/export.py` S4 kuralı)

Meslek/SGK/ücret/işe giriş boşsa hücreye HİÇ dokunulmaz. `""` niyeti
anlatmaz, `0` ise "ücreti sıfır" yalanı olurdu.
"""

from collections.abc import Sequence
from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.modules.personnel.schemas import PersonnelResponse
from app.modules.site_diary.models import WorkerSource

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

SHEET_TITLE = "Personel"

#: PE thead'i BİREBİR — "Proje" hariç (docstring), "İşe giriş" mockup'ın alt
#: satırından kendi sütununa alınmıştır.
COLUMN_HEADERS: tuple[str, ...] = (
    "Ad Soyad",
    "İşe giriş",
    "Tür",
    "Meslek",
    "SGK",
    "Ücret/Gün",
    "Durum",
)

#: `payroll/export.py` ve `timesheet/export.py` ile AYNI metinler — iki dosyada
#: iki farklı rozet metni kullanıcıya iki farklı kavram öğretirdi.
SOURCE_LABELS: dict[WorkerSource, str] = {
    WorkerSource.company: "Şirket",
    WorkerSource.subcontractor: "Taşeron",
    WorkerSource.general: "Genel",
    WorkerSource.freelance: "Serbest",
    WorkerSource.intern: "Stajyer",
}

#: PE "Durum" rozeti. Mockup yalnız "Aktif"i gösterir (tüm satırları aktiftir);
#: karşıtı ekranın `is_active` süzgecinin öteki ucudur. `is_draft` bu sütuna
#: GİRMEZ: ekranda da ayrı bir rozeti yoktur, kendi süzgecidir.
ACTIVE_LABEL = "Aktif"
INACTIVE_LABEL = "Pasif"

#: Tanınmayan enum değerinin GÖRÜNÜR düşüşü (`payroll/export.py` emsali).
UNKNOWN_LABEL_PREFIX = "?"

EMPTY_VALUE = "—"

XLSX_FILENAME = "personel.xlsx"

_COLUMN_WIDTHS = (28, 14, 12, 22, 18, 14, 10)


def _label(sozluk: dict, anahtar: object) -> str:
    """Etiket erişiminin TEK yolu — **sözlük DOĞRUDAN İNDEKSLENMEZ.**

    Paylaşılan `WorkerSource` genişlediğinde `SOZLUK[enum]` yazımı export'u
    sessizce `KeyError` → 500 yapardı. Düşüş dayanıklıdır ama SESSİZ DEĞİLDİR:
    etiket eklemeyi unutan dilim hücrede `?yeni_deger` görür.
    """
    return sozluk.get(anahtar, f"{UNKNOWN_LABEL_PREFIX}{getattr(anahtar, 'value', anahtar)}")


def source_label(source: object) -> str:
    return _label(SOURCE_LABELS, source)


def _text(value: object | None) -> str | None:
    """`null` ise **`None`** döner ve hücreye DOKUNULMAZ."""
    return None if value is None else str(value)


def _cells(person: PersonnelResponse) -> tuple[str | None, ...]:
    return (
        person.full_name,
        # `date` → ISO metin (`str(x)`); satır içi strftime YASAK.
        _text(person.hire_date),
        source_label(person.source),
        _text(person.trade),
        _text(person.sgk_no),
        # `str(Decimal)` — API metniyle birebir; ölçek KORUNUR.
        _text(person.wage_amount),
        ACTIVE_LABEL if person.is_active else INACTIVE_LABEL,
    )


def _write_row(sheet: Worksheet, row: int, values: tuple[str | None, ...]) -> None:
    for column, value in enumerate(values, start=1):
        if value is None:
            # `null` alana DOKUNULMAZ (docstring: boş niyeti böyle anlatılır).
            continue
        sheet.cell(row=row, column=column).value = value


def _apply_layout(sheet: Worksheet) -> None:
    for index, width in enumerate(_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width
    sheet.freeze_panes = "A2"


def build_personnel_workbook(items: Sequence[PersonnelResponse]) -> BytesIO:
    """Personel satırlarından xlsx üretir ve bellekteki tamponu döner.

    Kayıtsız süzgeçte de GEÇERLİ bir dosya döner — yalnız başlık satırı.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_TITLE

    _write_row(sheet, 1, COLUMN_HEADERS)
    for index, person in enumerate(items, start=2):
        _write_row(sheet, index, _cells(person))
    _apply_layout(sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def filename() -> str:
    """`personel.xlsx` — `audit/export.py` emsali sabit ad.

    Kartoteks TEK bir kümedir (proje/dönem gibi bir kapsam anahtarı yoktur);
    süzgeçler dosya adına KODLANMAZ çünkü aynı süzgecin iki farklı çalıştırması
    zaten aynı belgedir.
    """
    return XLSX_FILENAME
