"""Teklif karsilastirma Excel'i (TEK 38 "Excel" dugmesi, spec §7 S5) — T4.

`audit/export.py` deseninin kardesi: saf sunum katmani — `Request`/`Response`
bilmez, DB'ye dokunmaz, yalnizca HAZIR kartlari calisma kitabina cevirir.

Girdi bilincli olarak `PurchaseQuoteCard`tir (ekranin gordugu nesne): "Toplam"
sutunu boylece `service.list_quotes`in hesapladigi degerdir ve o da
`transitions.order_total_from_quote`tan gelir. Ham teklif satirlari alinsaydi
formul BURADA ikinci kez yazilir, Excel ile ekran ayni teklif icin farkli toplam
gosterebilirdi.

FLOAT-YASAK (B4 dersi): her hucre ACIKCA `str` yazilir. openpyxl bir hucreye
`Decimal` verildiginde onu sessizce sayi+bicim olarak saklar; para birimi ve
binlik ayraci da kaybolurdu.
"""

from collections.abc import Iterable, Sequence
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.modules.procurement.models import PaymentTerms
from app.modules.procurement.schemas import PurchaseQuoteCard

__all__ = [
    "COLUMN_HEADERS",
    "SHEET_TITLE",
    "XLSX_FILENAME_SUFFIX",
    "build_quote_comparison_workbook",
]

SHEET_TITLE = "Teklif Karşılaştırması"

#: TEK kartinin satirlari — sira ve metin degistirilmez (mockup sadakati):
#: tedarikci adi (57) · Birim Fiyat (64) · Toplam (65) · Teslimat (68) ·
#: Garanti (69) · Odeme (70) · Nakliye (71).
COLUMN_HEADERS: tuple[str, ...] = (
    "Tedarikçi",
    "Birim Fiyat",
    "Toplam",
    "Teslimat",
    "Garanti",
    "Ödeme",
    "Nakliye",
)

#: Dosya adinin sabit kuyrugu; onune talep NUMARASI gelir (`SAT-2026-0042-…`).
XLSX_FILENAME_SUFFIX = "teklif-karsilastirma.xlsx"

#: "EN IYI FIYAT"/"EN HIZLI" rozetleri sutun OLARAK yazilmaz: ekranda gorsel
#: bir vurgudur ve tabloda kendi basina siralanabilir bir veri degildir.
#: Karsilastirmanin kendisi zaten "Toplam" sutunundan okunur.

#: Mockup'in odeme dili (TEK 70) — ham enum degeri disa aktarilmaz.
PAYMENT_LABELS: dict[PaymentTerms, str] = {
    PaymentTerms.cash: "Peşin",
    PaymentTerms.days_15: "15 gün vadeli",
    PaymentTerms.days_30: "30 gün vadeli",
    PaymentTerms.days_60: "60 gün vadeli",
}

#: Bos hucrenin gosterimi (`audit/export.py` ile ayni).
EMPTY_VALUE = "—"

_COLUMN_WIDTHS = (28, 16, 18, 18, 24, 18, 22)


def _money(value: Decimal) -> str:
    """TR para bicimi: `₺ 322.500,00`.

    Ayraclar mockup'la birebirdir (binlik nokta, ondalik virgul). Bicim
    burada TEK yerde durur; iki sutun da ondan gecer.
    """
    ingiliz = f"{value:,.2f}"
    return "₺ " + ingiliz.replace(",", "\x00").replace(".", ",").replace("\x00", ".")


def _shipping(card: PurchaseQuoteCard) -> str:
    """TEK 71/90'in iki hâli: "Dahil" ya da "Hariç (+₺8.000)".

    Tutari olmayan haric nakliye yalniz "Hariç" yazar — uydurma bir 0 tutar
    kullanici icin "bedava" anlamina gelirdi.
    """
    if card.shipping_included:
        return "Dahil"
    if card.shipping_cost is None:
        return "Hariç"
    return f"Hariç (+{_money(card.shipping_cost)})"


def _cells(card: PurchaseQuoteCard) -> tuple[str, ...]:
    return (
        card.supplier_name,
        _money(card.unit_price),
        _money(card.total_cost),
        card.delivery_time,
        card.warranty_note or EMPTY_VALUE,
        PAYMENT_LABELS.get(card.payment_terms, str(card.payment_terms.value)),
        _shipping(card),
    )


def _apply_layout(sheet: Worksheet) -> None:
    for index, width in enumerate(_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width
    sheet.freeze_panes = "A2"


def _write_row(sheet: Worksheet, index: int, values: Sequence[str]) -> None:
    for column, value in enumerate(values, start=1):
        # Deger her zaman `str`; openpyxl'in tip tahminine alan birakilmaz.
        sheet.cell(row=index, column=column).value = value


def build_quote_comparison_workbook(cards: Iterable[PurchaseQuoteCard]) -> BytesIO:
    """Kartlardan xlsx uretir ve bellekteki tamponu doner.

    Teklifsiz talepte bile GECERLI bir dosya doner (yalniz baslik satiri):
    kullanici "Excel" dugmesine bastiginda hata degil bos bir tablo gormeli.

    Kart SIRASI korunur — `list_quotes` onu sunucuda sabitler (`created_at`,
    `id`), yani Excel ekranla ayni duzendedir.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_TITLE

    _write_row(sheet, 1, COLUMN_HEADERS)
    for index, card in enumerate(cards, start=2):
        _write_row(sheet, index, _cells(card))
    _apply_layout(sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
