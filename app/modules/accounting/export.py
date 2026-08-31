"""Muhasebe Excel çıktıları (EXPORT-XLSX) — Mizan · Hesap Planı · Yevmiye Defteri.

`payroll/export.py` ve `units/export.py` deseninin kardeşi: **saf SUNUM
katmanı**. `Request`/`Response` bilmez, DB'ye DOKUNMAZ, kapsam/yetki kararı
VERMEZ; yalnız ekranın kendi servisinin ürettiği zarfı çalışma kitabına çevirir.

## Üç yüzey TEK dosyada

Üçü de `accounting` modülünün ekranlarıdır ve aynı üç kuralı paylaşır (para
yazımı, etiket sözlüğü, düzen). Ayrı üç dosyaya bölünselerdi `_money`/`_label`
üç kez yazılır ve zamanla ayrışırlardı.

## Hiçbir toplam BURADA hesaplanmaz

Mizanın `GENEL TOPLAM` satırı `TrialBalanceTotals`tan OLDUĞU GİBİ basılır —
`rows` üzerinden yeniden toplanmaz. Yeniden toplansaydı dosyanın tfoot'u
ekranın tfoot'undan ayrışabilir ve hangisinin doğru olduğu tartışılırdı
(`trial_balance.py` zaten `totals`ı satırlardan türetir; ikinci bir toplayıcı
o tek kaynağı ikiye bölerdi).

## FLOAT-YASAK

Her para hücresi AÇIKÇA `str(Decimal)` yazılır. openpyxl'e `Decimal` verilseydi
hücre sayı+biçim olarak saklanır, geri okumada tip değişir ve yuvarlama riski
doğardı. `str(Decimal)` API yanıtının metin temsiliyle BİREBİR aynıdır: ekranda
`284800.00` görünen değer dosyada da `284800.00`dur.

## 🔴 Sıfır para hücresine "—" BASILMAZ (mockup'tan bilinçli sapma)

Mizan mockup'ı (satır 84-159) ve E8 sıfır tarafa `—` çizer. Dosyada `—`
basılsaydı Excel'deki hücre ile API'nin JSON'u AYNI satır için farklı metin
taşırdı ve `GENEL TOPLAM` satırı kolonundaki `—`lerin toplamı olmazdı. `—`
yalnız **`null`** alan içindir (`units/export.py` emsali) ve bu üç zarfta
`null` olabilen tek alan `detail_note`tur. Sıfır bir OLGUDUR, boşluk değildir.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.modules.accounting.models import ChartAccountType
from app.modules.accounting.reports_schemas import (
    TrialBalanceResponse,
    TrialBalanceRow,
    TrialBalanceTotals,
)
from app.modules.accounting.schemas import (
    ChartAccountListResponse,
    ChartAccountResponse,
    LedgerResponse,
    LedgerRow,
)

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

#: `units/export.py`nin sabiti — yalnız **`null`** alan için.
EMPTY_VALUE = "—"

#: Tanınmayan enum değerinin GÖRÜNÜR düşüşü (`payroll/export.py` emsali).
UNKNOWN_LABEL_PREFIX = "?"


def _label(sozluk: dict, anahtar: object) -> str:
    """Etiket erişiminin TEK yolu — **sözlük DOĞRUDAN İNDEKSLENMEZ.**

    `ChartAccountType` bir kez zaten genişledi (`equity`, MT-1/KK-1). Doğrudan
    indekslenseydi altıncı bir üye export'u sessizce `KeyError` → **500**
    yapardı. Erişim dayanıklıdır ama düşüş SESSİZ DEĞİLDİR: etiket eklemeyi
    unutan sonraki dilim hücrede `?yeni_deger` görür.
    """
    return sozluk.get(anahtar, f"{UNKNOWN_LABEL_PREFIX}{getattr(anahtar, 'value', anahtar)}")


def _money(value: object) -> str:
    """Para hücresi — yuvarlama/biçimlendirme YOK, `str(Decimal)` (FLOAT-YASAK)."""
    return str(value)


def _write_row(sheet: Worksheet, row: int, values: tuple[str | None, ...]) -> None:
    for column, value in enumerate(values, start=1):
        if value is None:
            # `null` alana DOKUNULMAZ: `""` geri okumada `None`a döner ama niyeti
            # anlatmaz, `0` ise olmayan bir olguyu uydururdu.
            continue
        sheet.cell(row=row, column=column).value = value


def _apply_layout(sheet: Worksheet, widths: tuple[int, ...], header_row: int) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[
            sheet.cell(row=header_row, column=index).column_letter
        ].width = width
    sheet.freeze_panes = f"A{header_row + 1}"


def _workbook(title: str) -> tuple[Workbook, Worksheet]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    return workbook, sheet


def _buffer(workbook: Workbook) -> BytesIO:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


# --------------------------------------------------------------------------- #
# 1 — MİZAN (`Muhasebe - Mizan.dc.html`)
# --------------------------------------------------------------------------- #

TRIAL_BALANCE_SHEET_TITLE = "Mizan"

#: Mockup satır 84-99: iki kimlik kolonu + ÜÇ pencerenin borç/alacak çifti.
#: Üst başlık satırı (`Açılış Bakiyesi` / `Dönem Hareketi` / `Kapanış Bakiyesi`)
#: birleştirilmiş hücrelerdir; dosyada BİRLEŞTİRME yoktur (biçimlendirme yasak)
#: ve pencere adı her kolonun kendi metnine katılır — aksi hâlde altı kolonun
#: üçü "Borç" adını paylaşır ve hangisinin hangi pencere olduğu kaybolurdu.
TRIAL_BALANCE_HEADERS: tuple[str, ...] = (
    "Hesap Kodu",
    "Hesap Adı",
    "Açılış Borç",
    "Açılış Alacak",
    "Dönem Borç",
    "Dönem Alacak",
    "Kapanış Borç",
    "Kapanış Alacak",
)

#: Mockup satır 184.
TRIAL_BALANCE_TOTAL_LABEL = "GENEL TOPLAM"

#: Mockup satır 57 / 61 kontrol banner'ının İKİ hâli. Banner ekranın ayrılmaz
#: parçasıdır: `is_balanced=false` iken mali tablolar güvenilir değildir ve
#: dosyayı e-postayla alan kişi bunu satırlardan çıkaramaz.
TRIAL_BALANCE_STATE_LABELS: dict[bool, str] = {
    True: "Mizan Dengede",
    False: "Mizan Dengede Değil",
}

#: Şerit etiketleri — `payroll/export.py`nin `INFO_LABELS` deseni.
TRIAL_BALANCE_INFO_LABELS: tuple[str, ...] = ("Dönem", "Denge")

#: Tablo başlığının satır numarası: şerit + bir boş ayraç satırı.
TRIAL_BALANCE_HEADER_ROW = len(TRIAL_BALANCE_INFO_LABELS) + 2

_TRIAL_BALANCE_WIDTHS = (14, 34, 16, 16, 16, 16, 16, 16)


def trial_balance_state_label(is_balanced: object) -> str:
    return _label(TRIAL_BALANCE_STATE_LABELS, is_balanced)


def trial_balance_period_label(year: int, month: int) -> str:
    """Mockup satır 46 `Ocak–Temmuz 2026` — dönem BİRİKİMLİ bir ARALIKTIR.

    Ay ADI yazılmaz: on iki Türkçe ay adından oluşan ikinci bir sözlük, hiçbir
    sunucu alanına karşılık gelmeyen bir icat olurdu. Aralık aynı olguyu
    sayıyla söyler ve `filename` ile aynı yazımı kullanır.
    """
    return f"01.{year}–{month:02d}.{year}"


def _trial_balance_row_cells(row: TrialBalanceRow) -> tuple[str, ...]:
    return (
        row.account_code,
        row.account_name,
        _money(row.opening_debit),
        _money(row.opening_credit),
        _money(row.period_debit),
        _money(row.period_credit),
        _money(row.closing_debit),
        _money(row.closing_credit),
    )


def _trial_balance_total_cells(totals: TrialBalanceTotals) -> tuple[str | None, ...]:
    """tfoot (mockup satır 183-191) — 🔴 satırlardan YENİDEN TOPLANMAZ.

    Zarfın kendi `totals`ı basılır; ikinci bir toplayıcı, ekranın tfoot'u ile
    dosyanınkini ayrıştırabilirdi.
    """
    return (
        TRIAL_BALANCE_TOTAL_LABEL,
        None,
        _money(totals.opening_debit),
        _money(totals.opening_credit),
        _money(totals.period_debit),
        _money(totals.period_credit),
        _money(totals.closing_debit),
        _money(totals.closing_credit),
    )


def build_trial_balance_workbook(report: TrialBalanceResponse) -> BytesIO:
    """Mizan zarfından xlsx üretir.

    Satırsız dönemde de GEÇERLİ bir dosya döner: şerit, başlıklar ve sıfırlı
    `GENEL TOPLAM` satırı basılır — uydurma bir hesap satırı üretilmez.
    """
    workbook, sheet = _workbook(TRIAL_BALANCE_SHEET_TITLE)

    serit = (
        trial_balance_period_label(report.year, report.month),
        trial_balance_state_label(report.is_balanced),
    )
    for index, (etiket, deger) in enumerate(
        zip(TRIAL_BALANCE_INFO_LABELS, serit, strict=True), start=1
    ):
        sheet.cell(row=index, column=1).value = etiket
        sheet.cell(row=index, column=2).value = deger

    _write_row(sheet, TRIAL_BALANCE_HEADER_ROW, TRIAL_BALANCE_HEADERS)

    row = TRIAL_BALANCE_HEADER_ROW
    for kayit in report.rows:
        row += 1
        _write_row(sheet, row, _trial_balance_row_cells(kayit))

    _write_row(sheet, row + 1, _trial_balance_total_cells(report.totals))
    _apply_layout(sheet, _TRIAL_BALANCE_WIDTHS, TRIAL_BALANCE_HEADER_ROW)
    return _buffer(workbook)


def trial_balance_filename(year: int, month: int) -> str:
    """`mizan-{yıl}-{ay}.xlsx` — iki dönemin dosyası birbirini EZMEMELİ."""
    return f"mizan-{year}-{month:02d}.xlsx"


# --------------------------------------------------------------------------- #
# 2 — HESAP PLANI (`Muhasebe - Hesap Planı.dc.html`)
# --------------------------------------------------------------------------- #

CHART_SHEET_TITLE = "Hesap Planı"

#: HP:59-63 — beş kolon, mockup SIRASIYLA.
CHART_HEADERS: tuple[str, ...] = ("Kod", "Hesap Adı", "Tür", "Bakiye", "Durum")

#: HP:60 `Tür` rozetleri — `models.ChartAccountType`in yorum satırlarıyla birebir.
ACCOUNT_TYPE_LABELS: dict[ChartAccountType, str] = {
    ChartAccountType.asset: "Aktif",
    ChartAccountType.liability: "Pasif",
    ChartAccountType.revenue: "Gelir",
    ChartAccountType.expense: "Gider",
    ChartAccountType.equity: "Özkaynak",
}

#: HP:62 `Durum` — mockup'ta METİN YOKTUR, renkli bir noktadır (satır 79).
#: 🔴 "Aktif"/"Pasif" YAZILAMAZ: o iki metin `Tür` kolonunun rozetleridir ve
#: R3'ün tam olarak uyardığı karışıklık (ikisi de Türkçe'de "aktif" okunur)
#: dosyada iki kolonu ayırt edilemez hâle getirirdi.
ACCOUNT_ACTIVE_LABELS: dict[bool, str] = {True: "Kullanımda", False: "Kullanım Dışı"}

_CHART_WIDTHS = (14, 38, 12, 18, 16)


def account_type_label(account_type: object) -> str:
    return _label(ACCOUNT_TYPE_LABELS, account_type)


def account_active_label(is_active: object) -> str:
    return _label(ACCOUNT_ACTIVE_LABELS, is_active)


def _chart_row_cells(account: ChartAccountResponse) -> tuple[str, ...]:
    """HP satırı. 🔴 `SINIF n` / grup bantları (HP:68-74) BASILMAZ: onlar
    istemcinin `class_code`/`level` alanlarından KURDUĞU sunum bantlarıdır,
    sunucudan gelen satırlar değildir — dosyaya konsaydı hesap sayısı satır
    sayısıyla uyuşmaz ve küme eşitliği yapısal olarak ölçülemez olurdu."""
    return (
        account.code,
        account.name,
        account_type_label(account.account_type),
        _money(account.balance),
        account_active_label(account.is_active),
    )


def build_chart_of_accounts_workbook(accounts: ChartAccountListResponse) -> BytesIO:
    """Hesap planı zarfından xlsx üretir; hesapsız kümede de geçerli dosya."""
    workbook, sheet = _workbook(CHART_SHEET_TITLE)
    _write_row(sheet, 1, CHART_HEADERS)

    row = 1
    for account in accounts.items:
        row += 1
        _write_row(sheet, row, _chart_row_cells(account))

    _apply_layout(sheet, _CHART_WIDTHS, 1)
    return _buffer(workbook)


def chart_of_accounts_filename() -> str:
    """`hesap-plani.xlsx` — kapsam TEKTİR (şirket geneli katalog, spec §3:
    proje/şantiye süzgeci YOKTUR), dolayısıyla ada gömülecek bir kapsam yoktur."""
    return "hesap-plani.xlsx"


# --------------------------------------------------------------------------- #
# 3 — YEVMİYE DEFTERİ (`Ekran 8 - Muhasebe.dc.html` / `Muhasebe - Profesyonel`)
# --------------------------------------------------------------------------- #

JOURNAL_SHEET_TITLE = "Yevmiye Defteri"

#: E8:101-106 — ALTI kolon. 🔴 `Muhasebe - Profesyonel.dc.html` (satır 118-124)
#: aynı tabloyu BEŞ kolonla çizer (`Bakiye` yoktur) ve iki mockup burada
#: ÇELİŞİR. E8 kazanır: `running_balance` bu ucun varlık sebebidir
#: (`ledger.py` modül docstring'i) ve dışarıda bırakılsaydı dosya ekranın
#: gösterdiği bir kolonu SESSİZCE kaybederdi. Fazla kolon veri kaybetmez,
#: eksik kolon kaybeder.
JOURNAL_HEADERS: tuple[str, ...] = (
    "Tarih",
    "Hesap Kodu",
    "Açıklama",
    "Borç",
    "Alacak",
    "Bakiye",
)

_JOURNAL_WIDTHS = (14, 14, 46, 16, 16, 18)


def journal_description(row: LedgerRow) -> str:
    """E8:113 hücresi İKİ SATIRLIDIR: açıklama + altında gri `detail_note`.

    Tek satıra indirgenseydi (`description` basılıp not atılsaydı) fişin ayırt
    edici notu — banka referansı, fiş numarası — dosyada kaybolurdu. `null` not
    hiç satır AÇMAZ; boş bir ikinci satır "not girilmiş ama boş" yalanı olurdu.
    """
    if row.detail_note is None:
        return row.description
    return f"{row.description}\n{row.detail_note}"


def _journal_row_cells(row: LedgerRow) -> tuple[str, ...]:
    return (
        # `date` → ISO (`str`). TR biçimi (E8:112 `17.07.2026`) burada elle
        # KURULMAZ: satır içi `strftime` biçim dizesi yasaktır ve `date`in
        # görüntüleme dönüşümü yoktur (saat dilimi yalnız `datetime`da anlamlı).
        str(row.entry_date),
        row.account_code,
        journal_description(row),
        _money(row.debit),
        _money(row.credit),
        # 🔴 `carried_balance` DAHİLDİR: `running_balance` mutlak bir seridir,
        # devri zaten içinde taşır (`ledger.py`). Ayrı bir devir satırı, aynı
        # sayıyı ikinci kez basıp toplanabilir sanısı yaratırdı.
        _money(row.running_balance),
    )


def build_journal_workbook(ledger: LedgerResponse) -> BytesIO:
    """Defter zarfından xlsx üretir; satırsız ayda da geçerli dosya döner."""
    workbook, sheet = _workbook(JOURNAL_SHEET_TITLE)
    _write_row(sheet, 1, JOURNAL_HEADERS)

    row = 1
    for kayit in ledger.items:
        row += 1
        _write_row(sheet, row, _journal_row_cells(kayit))

    _apply_layout(sheet, _JOURNAL_WIDTHS, 1)
    return _buffer(workbook)


def journal_filename(year: int, month: int) -> str:
    """`yevmiye-{yıl}-{ay}.xlsx` — pencere AYDIR, iki ay birbirini ezmemeli."""
    return f"yevmiye-{year}-{month:02d}.xlsx"


__all__ = [
    "ACCOUNT_ACTIVE_LABELS",
    "ACCOUNT_TYPE_LABELS",
    "CHART_HEADERS",
    "EMPTY_VALUE",
    "JOURNAL_HEADERS",
    "TRIAL_BALANCE_HEADERS",
    "TRIAL_BALANCE_HEADER_ROW",
    "TRIAL_BALANCE_INFO_LABELS",
    "TRIAL_BALANCE_STATE_LABELS",
    "TRIAL_BALANCE_TOTAL_LABEL",
    "UNKNOWN_LABEL_PREFIX",
    "XLSX_MEDIA_TYPE",
    "build_chart_of_accounts_workbook",
    "build_journal_workbook",
    "build_trial_balance_workbook",
    "chart_of_accounts_filename",
    "journal_filename",
    "trial_balance_filename",
]
