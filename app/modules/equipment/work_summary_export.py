"""Makine çalışma kaydı Excel çıktısı (EXPORT-XLSX) — mockup M3 "Excel İndir".

`units/export.py` · `payroll/export.py` · `audit/export.py` deseninin kardeşi:
saf SUNUM katmanı. `Request`/`Response` bilmez, DB'ye DOKUNMAZ, kapsam/yetki
kararı VERMEZ — yalnız `work_summary` servisinin ürettiği zarfı çalışma
kitabına çevirir.

## Neden ÖZET yüzeyi, ham iş kayıtları değil (ölçülmüş karar)

`GET /equipment/work-logs` satırları yalnız UUID taşır (`equipment_id`,
`site_id`, `operator_id`) — adları TAŞIMAZ. O yüzeyden bir Excel üretmek ya ham
UUID kolonları basardı (kullanıcı için okunmaz) ya da adları çözmek için ekranın
kullanmadığı İKİNCİ bir sorgu yolu açardı. `work-summary` zaten sayfalanmamıştır
ve `equipment_name` taşır; dosya doğrudan ondan basılır.

## 🔴 Hiçbir toplam BURADA hesaplanmaz — tfoot `totals`tan gelir (K15/K16)

Toplam satırı zarfın `totals` alanından OKUNUR; satırlar üzerinden YENİDEN
toplanmaz. Yeniden toplansaydı K16 sessizce çiğnenirdi: maliyeti BİLİNMEYEN
satır `null`dur ve toplama GİRMEZ; export kendi `sum`ını kursa o satırı ya 0
sayardı (uydurma) ya da bütün toplamı `null` yapardı. İkisi de ekranla çelişirdi.

Mockup'ın kendi tfoot'u (428 saat · ₺124.800 · %69) satırlarıyla TUTARSIZDIR
(692 · ₺144.200 · %57,7) ve KOPYALANMAZ — servis kararı (K15) aynen geçerlidir.

## FLOAT-YASAK (B4 dersi)

Her hücre AÇIKÇA `str` yazılır; `str(Decimal)` API yanıtının metniyle BİREBİR
aynıdır. `%` işareti hücreye GİRMEZ (sütun başlığı zaten "Kullanım %"): sayının
yanına konsaydı hücre metin olarak bile ekrandan farklı bir sayı temsili taşırdı.

## `null` alana DOKUNULMAZ (K16)

Kapasitesi tanımsız makinenin `usage_pct`i, ücreti tanımsız makinenin `cost`u
`null`dur ve hücreye HİÇ dokunulmaz. 0 yazmak "bedeli yok" yalanı olurdu.

## "Şantiye" sütunu YOKTUR — ölçülmüş sınır

Mockup ekipman adının ALTINDA şantiye ADInı gösterir; `WorkSummaryRow` yalnız
`site_id` (UUID) taşır. Ham UUID okunmaz, adı çözmek ikinci sorgu yolu açardı
(yukarıdaki gerekçe). Sütun bilinçli DIŞARIDA — "ekrandan az" meşru, "fazla"
değildir.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.modules.equipment.schemas import WorkSummaryResponse, WorkSummaryRow, WorkSummaryTotals

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

SHEET_TITLE = "Çalışma Kaydı"

#: M3 thead'i BİREBİR (sıra ve metin).
COLUMN_HEADERS: tuple[str, ...] = (
    "Ekipman",
    "Çalışma (Saat)",
    "Kullanım %",
    "Arıza (Saat)",
    "Maliyet",
)

#: M3 tfoot'unun ilk hücresi.
TOTAL_LABEL = "Toplam"

EMPTY_VALUE = "—"

_COLUMN_WIDTHS = (32, 16, 14, 14, 16)


def _text(value: object | None) -> str | None:
    """`null` ise **`None`** döner ve hücreye DOKUNULMAZ (K16)."""
    return None if value is None else str(value)


def _cells(row: WorkSummaryRow) -> tuple[str | None, ...]:
    return (
        row.equipment_name,
        str(row.hours),
        _text(row.usage_pct),
        str(row.breakdown_hours),
        _text(row.cost),
    )


def _total_cells(totals: WorkSummaryTotals) -> tuple[str | None, ...]:
    """🔴 Toplam satırı ZARFTAN okunur, satırlardan YENİDEN TOPLANMAZ (K15/K16)."""
    return (
        TOTAL_LABEL,
        str(totals.hours),
        _text(totals.usage_pct_avg),
        str(totals.breakdown_hours),
        str(totals.cost),
    )


def _write_row(sheet: Worksheet, row: int, values: tuple[str | None, ...]) -> None:
    for column, value in enumerate(values, start=1):
        if value is None:
            # `null` alana DOKUNULMAZ (K16).
            continue
        sheet.cell(row=row, column=column).value = value


def _apply_layout(sheet: Worksheet) -> None:
    for index, width in enumerate(_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width
    sheet.freeze_panes = "A2"


def build_work_summary_workbook(summary: WorkSummaryResponse) -> BytesIO:
    """Çalışma özeti zarfından xlsx üretir ve bellekteki tamponu döner.

    Kayıtsız ayda da GEÇERLİ bir dosya döner: başlık + sıfırlı toplam satırı;
    uydurma bir ekipman satırı üretilmez. Haftalık mini grafik (M3 219-243)
    dosyaya GİRMEZ — bir çizim yüzeyidir, tablonun kolonu değildir.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_TITLE

    _write_row(sheet, 1, COLUMN_HEADERS)
    row = 1
    for satir in summary.rows:
        row += 1
        _write_row(sheet, row, _cells(satir))
    _write_row(sheet, row + 1, _total_cells(summary.totals))
    _apply_layout(sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def filename(year: int, month: int) -> str:
    """`makine-calisma-{yıl}-{ay}.xlsx` — dönem dosya adında GÖRÜNÜR.

    İki ay aynı klasöre indiğinde dosyalar birbirini EZMEMELİDİR
    (`payroll.export.filename` ile aynı gerekçe).
    """
    return f"makine-calisma-{year}-{month:02d}.xlsx"
