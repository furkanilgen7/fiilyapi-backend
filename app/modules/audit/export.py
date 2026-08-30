"""Denetim gunlugu Excel ciktisi (plan Task 5).

Saf sunum katmani: `Request`/`Response` bilmez, DB'ye dokunmaz — yalnizca
repository satirlarini calisma kitabina cevirir.

Sutun basliklari mockup (`Ayarlar - Denetim Gunlugu.dc.html`) tablosuyla BIREBIR.

FLOAT-YASAK (B4 dersi): her hucre ACIKCA `str` yazilir. openpyxl bir hucreye
`datetime`/`int`/`Decimal` verildiginde onu sessizce sayi+bicim olarak saklar ve
cikti hucreleri metin olmaktan cikar; bu yuzden donusum burada, tek yerde yapilir.

SAAT DILIMI: `occurred_at` UTC saklanir; Zaman sutunu `Europe/Istanbul`'a cevrilerek
yazilir (bkz. `app/core/timezone.py`) — ayni kayit ekranda ve Excel'de AYNI saatte
gorunur.
"""

from collections.abc import Iterable, Sequence
from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.core.timezone import DISPLAY_TIMESTAMP_FORMAT, to_display
from app.modules.audit.models import AuditAction
from app.modules.audit.repository import AuditRow

SHEET_TITLE = "Denetim Günlüğü"

#: Mockup tablosunun basliklari — sira ve metin degistirilmez.
COLUMN_HEADERS: tuple[str, ...] = ("Zaman", "Kullanıcı", "İşlem", "Detay", "IP Adresi")

#: Mockup'in Turkce rozet etiketleri; ham enum degeri disa aktarilmaz.
ACTION_LABELS: dict[AuditAction, str] = {
    AuditAction.login: "Giriş",
    AuditAction.create: "Oluşturma",
    AuditAction.update: "Güncelleme",
    AuditAction.delete: "Silme",
    AuditAction.approve: "Onay",
    AuditAction.backup: "Yedekleme",
    AuditAction.ai_turn: "AI Turu",
}

#: Aktoru olmayan (otomatik) satirlarin kullanici sutunu.
SYSTEM_ACTOR_LABEL = "Sistem"

#: Mockup bos IP hucresinde uzun tire gosterir.
EMPTY_VALUE = "—"

#: Ad ve rolu tek hucrede birlestiren ayrac (mockup'in detay dilinde de kullanilir).
_ACTOR_SEPARATOR = " · "

_COLUMN_WIDTHS = (18, 26, 14, 60, 16)


def _actor_label(row: AuditRow) -> str:
    _, actor, role = row
    if actor is None:
        return SYSTEM_ACTOR_LABEL
    role_name = role.name if role is not None else ""
    return f"{actor.full_name}{_ACTOR_SEPARATOR}{role_name}" if role_name else actor.full_name


def _cells(row: AuditRow) -> tuple[str, ...]:
    """Tek denetim satirini mockup sutun sirasinda METIN hucrelere cevirir."""
    entry = row[0]
    return (
        # Ekranla ayni saat: `occurred_at` UTC saklanir, TR'ye cevrilerek yazilir.
        to_display(entry.occurred_at).strftime(DISPLAY_TIMESTAMP_FORMAT),
        _actor_label(row),
        ACTION_LABELS.get(entry.action, str(entry.action.value)),
        str(entry.detail),
        str(entry.ip_address) if entry.ip_address is not None else EMPTY_VALUE,
    )


def _apply_layout(sheet: Worksheet) -> None:
    for index, width in enumerate(_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width
    sheet.freeze_panes = "A2"


def _write_row(sheet: Worksheet, index: int, values: Sequence[str]) -> None:
    for column, value in enumerate(values, start=1):
        # Deger her zaman `str`; openpyxl'in tip tahminine alan birakilmaz.
        sheet.cell(row=index, column=column).value = value


def build_audit_workbook(rows: Iterable[AuditRow]) -> BytesIO:
    """Denetim satirlarindan xlsx calisma kitabi uretir ve bellekteki tamponu doner.

    Bos girdide bile gecerli bir dosya doner: yalnizca baslik satiri yazilir.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_TITLE

    _write_row(sheet, 1, COLUMN_HEADERS)
    for index, row in enumerate(rows, start=2):
        _write_row(sheet, index, _cells(row))
    _apply_layout(sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
