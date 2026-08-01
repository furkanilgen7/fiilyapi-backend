"""B9 — Excel ice aktarma: SAF cozumleme katmani (spec §6.4, §7.8).

Bu modul veritabanina DOKUNMAZ ve servis katmanindan HICBIR SEY ITHAL ETMEZ.
Boylece bagimlilik yonu tek yonludur (`service` → `importer`) ve baslik/hucre
mantigi bir HTTP istegi ya da DB olmadan test edilebilir. Alan kurallarindan
DB'ye ihtiyac duyanlar (`net > brut`, `owner_side` proje tipi, blokta mevcut
`unit_no`) BILEREK burada degil, servistedir — tekil `POST` ile ayni kod yolunu
kullanmalari icin.

DOSYA HICBIR YERE YAZILMAZ: girdi `bytes`'tir, `io.BytesIO` uzerinden okunur.
Bir dosya yolu ya da gecici dosya bu modulde YOKTUR ve eklenmemelidir (spec §7.8).
"""

import io
import re
from dataclasses import dataclass, replace
from decimal import ROUND_HALF_UP, Decimal, DecimalException

from openpyxl import load_workbook

from app.modules.units.models import UnitFacing, UnitKind, UnitOwnerSide

# Spec §7.8 sinir tablosu — sihirli sayi birakilmaz.
# 1000 satirlik bir `.xlsx` ~50 KB'tir; 2 MB fazlasiyla yeterlidir ve bellek
# saldirisini keser. KY'de 52, KKP'de 42 unite var; 1000 en buyuk gercekci
# projenin de ustundedir.
MAX_IMPORT_BYTES = 2 * 1024 * 1024
MAX_IMPORT_ROWS = 1000
# 60 hatali satirin tamamini govdeye dizmek yaniti okunamaz kilar; ilk 50 hata
# kullaniciya dosyayi duzeltmeye yeter, kalani ozetlenir (spec §7.8).
MAX_REPORTED_ERRORS = 50

# Spec §7.11 tablosundan BIREBIR. Ice aktarmaya ozgu metinler burada durur ki
# servis onlari ithal edebilsin; ters yon (importer → service) dongu olurdu.
IMPORT_BAD_TYPE = "Yalnızca .xlsx dosyası yüklenebilir"
IMPORT_TOO_LARGE = "Dosya çok büyük (en fazla 2 MB)"
IMPORT_TOO_MANY_ROWS = f"Dosyada en fazla {MAX_IMPORT_ROWS} satır olabilir"
IMPORT_MISSING_HEADERS = "Excel başlıkları eksik: {fields}"
IMPORT_ROW_ERRORS = "Dosya işlenemedi, {count} satırda hata var"
# P3.1 T11 satir metinleri (spec §6.5, §8.3). Spec bunlari `guards.py`'de
# listeliyor; COZUMLEME duzeyinde uretildikleri icin BURADA duruyorlar —
# `guards` DB'ye ve `projects.service`'e bagimlidir ve bu modulun saf kalmasi
# (DB'siz, servis-ithalsiz) modul docstring'inin acik sozudur. Metinler spec'ten
# BIREBIRDIR; yalniz dosya yeri farklidir.
IMPORT_ROW_GROSS_REQUIRED = "Brüt m² sıfır olamaz"
IMPORT_ROW_FLOOR_TOO_LONG = "Kat bilgisi en fazla 20 karakter olabilir"
# EI 173 — mockup'taki TEK uyari kurali. Uyari kumesi KAPALIDIR (spec §6.5):
# `min_sale_price` karsilastirmasi, m² basina fiyat sapmasi gibi kurallar
# ICAT EDILMEZ.
IMPORT_ROW_PRICE_BELOW_COST = "Fiyat maliyetin altında (₺{cost}) — kontrol edin"


def _lira(amount: Decimal) -> str:
    """Tutari BINLIK AYRACLI yazar: `860000.00` → `860.000` (koordinator karari).

    Ham `Decimal` gomulurse kullanici mesajda `₺860000.00` gorur ve rakamlari
    gozle saymak zorunda kalir. Mockup'taki `₺860K` KISALTMASI ise EKRAN
    gosterimidir: mesaj metnine gomulseydi frontend tutari bir daha kendi
    para birimi bicimlendiricisinden geciremezdi (metinden sayi geri
    okunamaz). Bu yuzden metne TAM sayi, binlik ayracli girer.

    Kurus BASAMAKLARI GOSTERILMEZ — bu bir UYARI ipucudur, muhasebe satiri
    degil; dosyadaki ham deger rapor satirinin `RowEcho`'sunda AYNEN durur.
    """
    return f"{amount:,.0f}".replace(",", ".")


# `units.floor` METINDIR (karar 4) ve sutun `String(20)`; sinir COZUMLEMEDE
# uygulanir ki kullanici DB'nin anlamsiz hatasini degil Turkce mesaji gorsun.
MAX_FLOOR_LENGTH = 20

_XLSX_SUFFIX = ".xlsx"
_MONEY = Decimal("0.01")

# Turkce `İ/ı` TUZAGI: `"İ".lower()` iki karakter uretir (`i` + U+0307 birlesik
# nokta), dolayisiyla ham `.lower()` ile "LİSTE FİYATI" hicbir zaman "liste
# fiyatı"na esitlenmez ve baslik eslestirmesi SESSIZCE basarisiz olur. Tum `i`
# turevleri TEK harfe katlanir; boylede ASCII yazan ("LISTE FIYATI") kullanici da
# ayni anahtara duser.
_LETTER_FOLD = str.maketrans({"İ": "i", "I": "i", "ı": "i", "i": "i"})


def normalize_header(raw: object) -> str:
    """Baslik/sozluk anahtari normalizasyonu: `İ/ı` katlama + kucultme + bosluk sadelestirme."""
    text = "" if raw is None else str(raw)
    return re.sub(r"\s+", " ", text.translate(_LETTER_FOLD).lower()).strip()


@dataclass(frozen=True)
class ImportColumn:
    label: str  # Kullaniciya gosterilen Turkce baslik (mockup etiketiyle birebir)
    field: str
    required: bool


# Spec §6.4 sutun tablosu (EI 85) — A `Blok` … L `Sahiplik`. Sira
# DEGISTIRILEMEZ: eksik baslik mesaji ve `.xlsx` sablonu bu sirayla uretilir.
#
# P3.1 T11: 9 → 12 sutun. `Kat`, `Cephe`, `Maliyet` YENI; `Tip` → `Oda Tipi` ve
# `Pay` → `Sahiplik` YENIDEN ADLANDIRILDI (eskiler `_HEADER_SYNONYMS` ile
# esanlamli kabul edilir). `Oda Tipi` ve `Brüt m²` ZORUNLU oldu (EI 161).
COLUMNS: tuple[ImportColumn, ...] = (
    ImportColumn("Blok", "block_name", True),
    ImportColumn("Kat", "floor", False),
    ImportColumn("Ünite No", "unit_no", True),
    ImportColumn("Tür", "unit_kind", True),
    ImportColumn("Oda Tipi", "layout", True),
    ImportColumn("Brüt m²", "gross_area_m2", True),
    ImportColumn("Net m²", "net_area_m2", False),
    ImportColumn("Cephe", "facing", False),
    ImportColumn("Liste Fiyatı", "list_price", False),
    ImportColumn("Rayiç Değer", "appraisal_value", False),
    # KARAR 10: okunur → uyariyi uretir → ATILIR. `ImportRow`'da karsiligi
    # YOKTUR ve `units`'te kolonu ACILMAZ (spec §4.5).
    ImportColumn("Maliyet", "cost", False),
    ImportColumn("Sahiplik", "owner_side", False),
)

# Geriye donuk uyum (spec §6.4): P3 CANLIDADIR ve kullanicinin elinde eski
# sablonla doldurulmus dosyalar olabilir. Esanlamli kabul etmemek sessiz bir
# "baslik eksik" 422'si uretirdi.
_HEADER_SYNONYMS: dict[str, tuple[str, ...]] = {
    "layout": ("Tip",),
    "owner_side": ("Pay",),
}

_DECIMAL_FIELDS = ("gross_area_m2", "net_area_m2", "list_price", "appraisal_value", "cost")

_KIND_BY_LABEL = {
    normalize_header("Daire"): UnitKind.apartment,
    normalize_header("Dükkan"): UnitKind.shop,
    # P3.1 §4.3 (UE 74): uc yeni deger. Enum genislemesi migration `c1d2e3f4a5b6`.
    normalize_header("Ofis"): UnitKind.office,
    normalize_header("Depo"): UnitKind.warehouse,
    normalize_header("Otopark"): UnitKind.parking,
}
_OWNER_SIDE_BY_LABEL = {
    normalize_header("BİZ"): UnitOwnerSide.contractor,
    normalize_header("ARSA"): UnitOwnerSide.landowner,
    # UE 95 etiketleri: kullanici FORMDA GORDUGU metni Excel'e yazacaktir.
    normalize_header("Yüklenici (Biz)"): UnitOwnerSide.contractor,
    normalize_header("Arsa Sahibi Payı"): UnitOwnerSide.landowner,
}
# Karar 7 (spec §4.2): mockup'ta gecen TAM OLARAK bes deger. Pusulanin kalan uc
# yonu ICAT EDILMEZ.
_FACING_BY_LABEL = {
    normalize_header("Güney"): UnitFacing.south,
    normalize_header("Güney-Batı"): UnitFacing.southwest,
    normalize_header("Doğu"): UnitFacing.east,
    normalize_header("Kuzey"): UnitFacing.north,
    normalize_header("Batı"): UnitFacing.west,
}

_KIND_CHOICES = "Daire, Dükkan, Ofis, Depo, Otopark"
_OWNER_SIDE_CHOICES = "BİZ, ARSA"
_FACING_CHOICES = "Güney, Güney-Batı, Doğu, Kuzey, Batı"


@dataclass(frozen=True)
class ImportRow:
    """Cozumlenmis TEK satir. Excel satir numarasi (`row`) korunur: hata raporu
    kullaniciyi dosyadaki satira gonderebilmelidir (baslik = 1, veri 2'den baslar)."""

    row: int
    block_name: str
    unit_no: str
    unit_kind: UnitKind
    layout: str | None
    gross_area_m2: Decimal | None
    net_area_m2: Decimal | None
    list_price: Decimal | None
    appraisal_value: Decimal | None
    owner_side: UnitOwnerSide | None
    # P3.1 T11 (spec §6.4): `floor` METINDIR ve DONUSTURULMEZ; `facing` bes
    # degerli sozlukten gelir. `cost` ALANI BILEREK YOKTUR (karar 10): maliyet
    # yalniz uyariyi uretmek icin okunur, hicbir yere sizmaz.
    floor: str | None = None
    facing: UnitFacing | None = None


@dataclass(frozen=True)
class RowError:
    """SATIR HATASI — satir YAZILMAZ (spec §6.5). Importer Pydantic'e bagimli
    degildir: bu tip `UnitImportRowReport.messages`'a servis katmaninda cevrilir."""

    row: int
    column: str | None
    message: str


@dataclass(frozen=True)
class RowWarning:
    """UYARI — satir GECERLIDIR, kullanici isterse yazilir (EI 192, spec §6.5).

    `RowError` ile AYNI sekilde olmasina ragmen AYRI bir tiptir: tek listede
    tasinsalardi bir uyarinin satiri sessizce atlatmasi (ya da bir hatanin
    yazilmasi) tek bir `if` hatasi kadar yakin olurdu — kismi aktarimin en
    pahali hata sinifi tam olarak budur (plan T12 risk notu).
    """

    row: int
    column: str | None
    message: str


@dataclass(frozen=True)
class RowEcho:
    """EI 118-125 sutunlarinin HAM yankisi — HATALI satirda da doludur.

    Rapor satiri kullanicinin dosyadaki satiri bulmasinin tek yoludur (EI 154
    hata satirini `C-6 · C · 2 · — · 0 · 1.258.600` diye basiyor); yalniz satir
    numarasi verilseydi 1000 satirlik bir dosyada arama yeniden kullanicinin
    isi olurdu.
    """

    row: int
    unit_no: str | None
    block_name: str | None
    floor: str | None
    layout: str | None
    gross_area_m2: Decimal | None
    list_price: Decimal | None


@dataclass(frozen=True)
class ParsedRow:
    """Cozumlenmis TEK satirin TAM sonucu: yanki + veri + hatalar + uyarilar.

    `data is None` ⇔ satir HATALIDIR. Bu esdegerlik kismi aktarimin cekirdegidir
    (spec §6.1): yalniz `data`'si dolu satirlar yazilabilir, dolayisiyla "hatali
    satir yazildi" hatasi TIP duzeyinde imkânsizlasir.
    """

    row: int
    echo: RowEcho
    data: ImportRow | None
    errors: list[RowError]
    warnings: list[RowWarning]


class ImportFileError(Exception):
    """Dosyanin TAMAMINI reddeden hata (tip, boyut, satir sayisi, eksik baslik).

    Satir bazli hatalardan ayridir: bunlar `errors` listesi uretmez, dosya hic
    islenmez.
    """


def parse_kind(raw: object) -> UnitKind:
    kind = _KIND_BY_LABEL.get(normalize_header(raw))
    if kind is None:
        raise ValueError(f"Tür tanınmıyor ({_KIND_CHOICES})")
    return kind


def parse_owner_side(raw: object) -> UnitOwnerSide | None:
    """Bos hucre `None` doner: pay noterden SONRA girilir (KKP 78, spec §5.3)."""
    key = normalize_header(raw)
    if not key:
        return None
    side = _OWNER_SIDE_BY_LABEL.get(key)
    if side is None:
        raise ValueError(f"Sahiplik tanınmıyor ({_OWNER_SIDE_CHOICES})")
    return side


def parse_facing(raw: object) -> UnitFacing | None:
    """Bos hucre `None` doner (cephe zorunlu degil); tanınmayan deger SATIR HATASI."""
    key = normalize_header(raw)
    if not key:
        return None
    facing = _FACING_BY_LABEL.get(key)
    if facing is None:
        raise ValueError(f"Cephe tanınmıyor ({_FACING_CHOICES})")
    return facing


def _text(raw: object) -> str:
    """Excel sayisal hucresi metin alanina duserse (`Ünite No` = 12) `12.0` OLMAZ."""
    if raw is None:
        return ""
    if isinstance(raw, float) and raw.is_integer():
        return str(int(raw))
    return str(raw).strip()


def parse_decimal(raw: object, label: str) -> Decimal | None:
    """Turkce sayi yazimi desteklenir: `1.234,56` → `1234.56`.

    Excel sayisal hucreleri zaten `int`/`float` gelir; metin hucreleri kullanicinin
    elle yazdigi hâldedir ve binlik/ondalik ayraci Turkce klavyede ters yerlestirilir.
    `float` ARA DEGERI KULLANILMAZ — para ve alan sutunlari `Decimal`'dir.
    """
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None
    if isinstance(raw, bool):
        raise ValueError(f"{label} sayıya çevrilemedi")
    text = str(raw).strip().replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        value = Decimal(text)
    except (DecimalException, ValueError) as exc:
        raise ValueError(f"{label} sayıya çevrilemedi") from exc
    if value < 0:
        raise ValueError(f"{label} negatif olamaz")
    return value.quantize(_MONEY, rounding=ROUND_HALF_UP)


def ensure_xlsx(filename: str | None) -> None:
    """`.xls` / `.csv` REDDEDILIR (spec §7.8): `openpyxl` yalniz `.xlsx` okur ve
    yanlis tipe anlamsiz bir cozumleme hatasi vermek yerine acik mesaj verilir."""
    if not (filename or "").lower().endswith(_XLSX_SUFFIX):
        raise ImportFileError(IMPORT_BAD_TYPE)


def ensure_size(size: int | None) -> None:
    """Boyut IKI KEZ kontrol edilir (plan B9): akis basinda istemcinin bildirdigi
    uzunlukla, sonra GERCEKTEN okunan `bytes` uzunluguyla. Istemci basligina
    guvenilmez; `None` gelmesi kontrolu atlatmaz, yalnizca ikinci kontrole birakir."""
    if size is not None and size > MAX_IMPORT_BYTES:
        raise ImportFileError(IMPORT_TOO_LARGE)


def _header_index(sheet) -> dict[str, int]:
    """Ilk satir basliktir. Beklenmeyen ek sutunlar YOK SAYILIR (spec §7.8).

    Kanonik baslik once denenir, YOKSA esanlamlisi (spec §6.4 geriye donuk
    uyum): kullanicinin elindeki P3 sablonu `Tip`/`Pay` tasiyor.
    """
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None) or ()
    positions = {normalize_header(cell): index for index, cell in enumerate(header_row)}
    index: dict[str, int] = {}
    for column in COLUMNS:
        for label in (column.label, *_HEADER_SYNONYMS.get(column.field, ())):
            key = normalize_header(label)
            if key in positions:
                index[column.field] = positions[key]
                break
    missing = [column.label for column in COLUMNS if column.required and column.field not in index]
    if missing:
        raise ImportFileError(IMPORT_MISSING_HEADERS.format(fields=", ".join(missing)))
    return index


def _cell(row: tuple, index: dict[str, int], field: str) -> object:
    position = index.get(field)
    if position is None or position >= len(row):
        return None
    return row[position]


def _label(field: str) -> str:
    return next(column.label for column in COLUMNS if column.field == field)


def _optional_text(value: object) -> str | None:
    return value if isinstance(value, str) else None


def _optional_decimal(value: object) -> Decimal | None:
    return value if isinstance(value, Decimal) else None


def _parse_row(number: int, row: tuple, index: dict[str, int]) -> ParsedRow:
    """Bir satirin TUM hatalari toplanir — ilk hatada durulmaz.

    Kullanici 48 satirlik bir dosyayi hata basina bir kez yuklemek zorunda
    kalmamalidir (spec §7.8 gerekcesi) ve EI 161 bir satirda IKI mesaj gosteriyor.

    `Maliyet` YEREL bir degiskendir (karar 10): yalniz EI 173 uyarisini uretir,
    donen `ImportRow`'a girmez ve hicbir sutuna yazilmaz.
    """
    errors: list[RowError] = []
    warnings: list[RowWarning] = []
    values: dict[str, object] = {}

    for field in ("block_name", "floor", "unit_no", "layout"):
        label = _label(field)
        text = _text(_cell(row, index, field))
        required = next(column.required for column in COLUMNS if column.field == field)
        if required and not text:
            errors.append(RowError(number, label, f"{label} boş olamaz"))
        # KARAR 4: `Kat` METINDIR — sozluk YOKTUR, tek kural uzunluktur.
        if field == "floor" and len(text) > MAX_FLOOR_LENGTH:
            errors.append(RowError(number, label, IMPORT_ROW_FLOOR_TOO_LONG))
        values[field] = text or None

    try:
        values["unit_kind"] = parse_kind(_cell(row, index, "unit_kind"))
    except ValueError as exc:
        errors.append(RowError(number, "Tür", str(exc)))

    for field in _DECIMAL_FIELDS:
        label = _label(field)
        try:
            values[field] = parse_decimal(_cell(row, index, field), label)
        except ValueError as exc:
            errors.append(RowError(number, label, str(exc)))
    # EI 161 "Brüt m² sıfır olamaz": bos DA sifir DA hatadir (spec §6.5).
    if values.get("gross_area_m2") in (None, Decimal(0)):
        errors.append(RowError(number, _label("gross_area_m2"), IMPORT_ROW_GROSS_REQUIRED))

    try:
        values["facing"] = parse_facing(_cell(row, index, "facing"))
    except ValueError as exc:
        errors.append(RowError(number, "Cephe", str(exc)))

    try:
        values["owner_side"] = parse_owner_side(_cell(row, index, "owner_side"))
    except ValueError as exc:
        errors.append(RowError(number, "Sahiplik", str(exc)))

    cost = values.pop("cost", None)
    list_price = values.get("list_price")
    if isinstance(cost, Decimal) and isinstance(list_price, Decimal) and list_price < cost:
        warnings.append(
            RowWarning(
                number, _label("list_price"), IMPORT_ROW_PRICE_BELOW_COST.format(cost=_lira(cost))
            )
        )

    # EI 118-125: rapor satiri dosyadaki degerleri HATALI satirda da tasir —
    # kullanici satiri ancak boyle bulur. Bu yuzden yanki `errors`'tan BAGIMSIZ
    # kurulur ve `ImportRow`'un varligina bagli degildir.
    echo = RowEcho(
        row=number,
        unit_no=_optional_text(values.get("unit_no")),
        block_name=_optional_text(values.get("block_name")),
        floor=_optional_text(values.get("floor")),
        layout=_optional_text(values.get("layout")),
        gross_area_m2=_optional_decimal(values.get("gross_area_m2")),
        list_price=_optional_decimal(values.get("list_price")),
    )
    if errors:
        # Satirin durumu TEKTIR (EI 119): hatali satir ayrica "uyarili" olmaz.
        return ParsedRow(row=number, echo=echo, data=None, errors=errors, warnings=[])
    return ParsedRow(
        row=number,
        echo=echo,
        data=ImportRow(
            row=number,
            block_name=str(values["block_name"]),
            unit_no=str(values["unit_no"]),
            unit_kind=values["unit_kind"],  # type: ignore[arg-type]
            layout=values["layout"],  # type: ignore[arg-type]
            gross_area_m2=values["gross_area_m2"],  # type: ignore[arg-type]
            net_area_m2=values["net_area_m2"],  # type: ignore[arg-type]
            list_price=values["list_price"],  # type: ignore[arg-type]
            appraisal_value=values["appraisal_value"],  # type: ignore[arg-type]
            owner_side=values["owner_side"],  # type: ignore[arg-type]
            floor=values["floor"],  # type: ignore[arg-type]
            facing=values["facing"],  # type: ignore[arg-type]
        ),
        errors=[],
        warnings=warnings,
    )


def _mark_duplicates(parsed_rows: list[ParsedRow]) -> list[ParsedRow]:
    """Dosya ICINDE ayni `(Blok, Ünite No)` ikilisi — DB'ye hic gitmeden yakalanir.

    Hata IKINCI satira yazilir ve o satirin `data`'si DUSURULUR: kismi aktarimda
    "hatali satir yazilmaz" garantisi `data is None` esdegerligiyle taşınır, ayri
    bir hata listesiyle degil — iki kaynak olsaydi biri sessizce atlanabilirdi.
    """
    seen: set[tuple[str, str]] = set()
    marked: list[ParsedRow] = []
    for parsed in parsed_rows:
        if parsed.data is None:
            marked.append(parsed)
            continue
        # Blok adi NORMALLESTIRILEREK eslesir (servis de blogu boyle bulur), ama
        # `unit_no` HARFI HARFINE karsilastirilir: `uq_units_block_no` tam
        # esitliktir, "A1" ile "a1" DB'de iki ayri unitedir.
        key = (normalize_header(parsed.data.block_name), parsed.data.unit_no)
        if key in seen:
            marked.append(
                replace(
                    parsed,
                    data=None,
                    errors=[
                        RowError(
                            parsed.row, "Ünite No", "Bu ünite numarası dosyada birden çok kez var"
                        )
                    ],
                    warnings=[],
                )
            )
            continue
        seen.add(key)
        marked.append(parsed)
    return marked


def parse_units_file(content: bytes) -> list[ParsedRow]:
    """`bytes` → SATIR SATIR cozumleme sonucu (`ParsedRow` listesi).

    Dosyanin tamamini reddeden durumlar `ImportFileError` firlatir; satir bazli
    hatalar DONDURULUR, cunku cagiran onlari DB kaynakli hatalarla BIRLESTIRIP
    tek raporda sunar (spec §6.3).

    Uyarilar (EI 173) hatalardan AYRI listede doner: hata satiri YAZILMAZ, uyari
    satiri kullanici isterse YAZILIR (EI 192) — ikisini tek listede tasimak
    kismi aktarimin en pahali hata sinifini (yanlis satirin yazilmasi) davet
    ederdi.
    """
    if len(content) > MAX_IMPORT_BYTES:
        raise ImportFileError(IMPORT_TOO_LARGE)
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl cesitli istisna tipleri firlatir
        raise ImportFileError(IMPORT_BAD_TYPE) from exc
    try:
        sheet = workbook.active
        index = _header_index(sheet)
        parsed_rows: list[ParsedRow] = []
        processed = 0
        # Satir numarasi Excel'in kendi numarasidir: baslik 1, veri 2'den baslar.
        for number, raw in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None or str(cell).strip() == "" for cell in raw):
                continue  # Tamamen bos satir (Excel'in kuyruk satirlari) sessizce atlanir
            processed += 1
            # Sinir OKUMA SIRASINDA uygulanir: 1001. satirda durulur, once tum
            # dosyayi bellege alip sonra saymak sinirin amacini bozardi.
            if processed > MAX_IMPORT_ROWS:
                raise ImportFileError(IMPORT_TOO_MANY_ROWS)
            parsed_rows.append(_parse_row(number, raw, index))
    finally:
        # `read_only` modunda acik kalan dosya tanitici birakilmaz.
        workbook.close()
    return _mark_duplicates(parsed_rows)
