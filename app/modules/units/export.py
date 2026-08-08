"""Paylasim tablosu Excel ciktisi (P9 T4, spec §5 / S3) — mockup KKP 24 "Excel".

`boq/export.py`, `audit/export.py` ve `timesheet/export.py` deseninin kardesi:
saf SUNUM katmani. `Request`/`Response` bilmez, DB'ye DOKUNMAZ, kapsam/yetki
karari VERMEZ — yalnizca `UnitListResponse` zarfini (liste ucunun urettigi
zarfin AYNISI) calisma kitabina cevirir.

## Hicbir deger BURADA hesaplanmaz

Etiket (`label`), taraf, hissedar/alici adi ve satis durumu zarftan HAZIR gelir;
zarf da `service.list_units` ile uretilir. Excel kendi sorgusunu/hesabini
kursaydi ekran ile dosya zamanla ayrisir ve hangisinin dogru oldugu tartisilir
hale gelirdi (`timesheet/export.py` ile birebir ayni gerekce: bir kere kur, iki
kere bas).

## FLOAT-YASAK (B4 dersi, `audit/export.py` emsali)

Her hucre ACIKCA `str` yazilir: openpyxl bir hucreye `Decimal`/`int` verildiginde
onu sessizce sayi+bicim olarak saklar ve geri okumada tip degisir. `Decimal`ler
API yanitinin metin temsiliyle AYNI sekilde basilir (`str(Decimal)`), boylece
ekranda "148.00" gorunen deger dosyada da "148.00"dur — yeniden yuvarlama YOK.

## Iki BIRLESIM sutunu mockup'tandir, icat degildir

- KKP 91 "Hissedar / Alici": ARSA satirinda `shareholder_name`, BIZ satirinda
  `buyer_name` (spec §4.3). Taraf ATANMAMISSA hissedar zaten `None`dur
  (guards §4.2) ve satirin tek gercek tarafi varsa aliciysa o basilir.
- KKP 92 "Satis Durumu": ARSA satiri "Arsa Sahibinde" basar. Bu deger
  `UnitSalesStatus` kumesine GIRMEZ, `owner_side` turevidir (`models.py` §110)
  ve zarfin kendi hesapladigi `is_landowner_share` alanindan okunur.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.modules.units.models import UnitOwnerSide, UnitSalesStatus
from app.modules.units.schemas import UnitListResponse, UnitResponse

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

SHEET_TITLE = "Paylaşım Tablosu"

#: KKP 86-92 — sira ve metin mockup'la BIREBIR (spec §5).
COLUMN_HEADERS: tuple[str, ...] = (
    "Ünite",
    "Tip",
    "m²",
    "Rayiç Değer",
    "Sahip",
    "Hissedar / Alıcı",
    "Satış Durumu",
)

#: KKP 100/109 rozetleri. Metinler `importer._OWNER_SIDE_BY_LABEL` ile ayni:
#: disa aktarilan dosya, ice aktarma cozumleyicisinin okuyabildigi metni tasir.
OWNER_SIDE_LABELS: dict[UnitOwnerSide, str] = {
    UnitOwnerSide.contractor: "BİZ",
    UnitOwnerSide.landowner: "ARSA",
}

#: KKP 92 rozetleri. "Satışta" KKP'nin yazimidir; UE 94'un "Satışta (Boş)"
#: parantezi FORM etiketidir ve tabloya girmez.
SALES_STATUS_LABELS: dict[UnitSalesStatus, str] = {
    UnitSalesStatus.listed: "Satışta",
    UnitSalesStatus.reserved: "Rezerve",
    UnitSalesStatus.sold: "Satıldı",
    UnitSalesStatus.closed: "Satışa Kapalı",
}

#: KKP 119'un "—" hucresi (`audit/export.py` emsali).
EMPTY_VALUE = "—"

#: KKP 92'nin ARSA satirindaki sabit metni; `UnitSalesStatus` kumesine girmez.
LANDOWNER_STATUS_LABEL = "Arsa Sahibinde"

_COLUMN_WIDTHS = (18, 10, 10, 16, 10, 26, 16)


def _text(value: object | None) -> str:
    """`Decimal`/`str` degeri hucreye yazilabilir metne cevirir; bos ise "—".

    Yuvarlama/bicimlendirme YAPILMAZ: `str(Decimal)` API yanitiyla ayni metni
    verir (boq §5.3 karari).
    """
    return EMPTY_VALUE if value is None else str(value)


def _party(unit: UnitResponse) -> str:
    """KKP 91 birlesimi: ARSA → hissedar, BIZ → alici.

    Tarafi ATANMAMIS unitede hissedar zaten olamaz (spec §4.2 kapisi), bu yuzden
    o satirda da alici dali calisir — varsa gercek alici gizlenmez, yoksa "—".
    """
    if unit.owner_side is UnitOwnerSide.landowner:
        return _text(unit.shareholder_name)
    return _text(unit.buyer_name)


def _sales_status(unit: UnitResponse) -> str:
    """KKP 92: ARSA satiri her zaman "Arsa Sahibinde" (vitrin durumu okunmaz)."""
    if unit.is_landowner_share:
        return LANDOWNER_STATUS_LABEL
    if unit.sales_status is None:
        return EMPTY_VALUE
    return SALES_STATUS_LABELS[unit.sales_status]


def _unit_row(unit: UnitResponse) -> tuple[str, ...]:
    return (
        # KKP 96 "A · Daire 1" — etiket zarfin TUREV alanidir, burada kurulmaz.
        unit.label,
        _text(unit.layout),
        _text(unit.gross_area_m2),
        _text(unit.appraisal_value),
        EMPTY_VALUE if unit.owner_side is None else OWNER_SIDE_LABELS[unit.owner_side],
        _party(unit),
        _sales_status(unit),
    )


def _write_row(sheet: Worksheet, row: int, values: tuple[str, ...]) -> None:
    for column, value in enumerate(values, start=1):
        # Deger her zaman `str`; openpyxl'in tip tahminine alan birakilmaz.
        sheet.cell(row=row, column=column).value = value


def _apply_layout(sheet: Worksheet) -> None:
    for index, width in enumerate(_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width
    sheet.freeze_panes = "A2"


def build_units_workbook(units: UnitListResponse) -> BytesIO:
    """Unite zarfindan xlsx calisma kitabi uretir ve bellekteki tamponu doner.

    Blok basligi satiri YOKTUR: KKP tablosu duz bir listedir ve blok adi zaten
    her satirin `Ünite` hucresindedir. Unitesi olmayan projede de GECERLI bir
    dosya doner — yalniz baslik satiri.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_TITLE

    _write_row(sheet, 1, COLUMN_HEADERS)

    row = 1
    for group in units.blocks:
        for unit in group.units:
            row += 1
            _write_row(sheet, row, _unit_row(unit))

    _apply_layout(sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def filename(project_code: str) -> str:
    """`paylasim-tablosu-{proje kodu}.xlsx` — iki projenin dosyasi ayni klasorde
    birbirini ezmemeli (`timesheet.export.filename` ile ayni gerekce)."""
    return f"paylasim-tablosu-{project_code}.xlsx"
