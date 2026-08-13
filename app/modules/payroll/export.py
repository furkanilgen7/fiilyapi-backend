"""Bordro Excel çıktısı (İK-3 T5, spec §5) — mockup BY 55 "Excel" düğmesi.

`timesheet/export.py`nin kardeşi ve onun deseninin AYNISI: saf SUNUM katmanı.
`Request`/`Response` bilmez, DB'ye dokunmaz, kapsam/yetki kararı VERMEZ — yalnız
T3'ün ürettiği dönem detay zarfını çalışma kitabına çevirir.

## Hiçbir toplam BURADA hesaplanmaz

Satır tutarları `compute.py`den, dört kartın ve toplam satırının sayıları
`summary.py`den HESAPLANMIŞ olarak gelir. Excel kendi toplamını kursaydı ekran
ile dosya zamanla ayrışır ve hangisinin doğru olduğu tartışılırdı.

## FLOAT-YASAK (B4 dersi, `audit/export.py` emsali)

Her hücre AÇIKÇA `str` yazılır: openpyxl bir hücreye `Decimal`/`int`/`date`
verildiğinde onu sessizce sayı+biçim olarak saklar ve geri okumada tip değişir.

## 🔴 `null` alana DOKUNULMAZ — 0 BASILMAZ (S4)

Hesaplanamamış satırın brütü/kesintisi/neti/bölüşümü `null`dur ve hücreye HİÇ
dokunulmaz (`""` yazmak openpyxl'de geri okumada `None`'a döner ama niyeti
anlatmaz — boq §5.3 dersi). 0 yazmak "ödenecek bir şey yok" yalanı olurdu ve
Excel'i açan kişi eksik veriyi göremezdi. Gün ise `null` OLABİLİR (serbest
meslekte, S7/BY 254 "—") ama hesaplanamamış satırda YAZILIR: gün puantajdan
OKUNAN bir olgudur (`compute._uncomputed`).
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.modules.payroll.models import PayrollLineStatus
from app.modules.payroll.schemas import (
    PayrollLineResponse,
    PayrollPeriodDetailResponse,
    PayrollSummaryResponse,
)
from app.modules.site_diary.models import WorkerSource

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

SHEET_TITLE = "Bordro"

#: Başlık şeridi: BY 52 (ay seçici) · BY 61 (durum banner'ı) · BY 63 ("Son
#: ödeme") + BY 69-93'ün DÖRT KARTI. Kartlar dosyada da durur çünkü BY'yi
#: gören kullanıcı Excel'i o kartlarla birlikte okur.
INFO_LABELS: tuple[str, ...] = (
    "Dönem",
    "Durum",
    "Son Ödeme",
    "Toplam Net Ödenecek",
    "Banka Transferi",
    "Elden (Nakit)",
    "İşverene Toplam Maliyet",
)

#: BY 110-118 — dokuz sütun, mockup SIRASIYLA.
COLUMN_HEADERS: tuple[str, ...] = (
    "Personel",
    "Tür",
    "Gün",
    "Brüt",
    "Kesinti",
    "Net",
    "Banka",
    "Elden",
    "Durum",
)

#: BY 298 "TOPLAM (48 çalışan)".
TOTAL_LABEL_PREFIX = "TOPLAM"

#: BY 127 / 175 / 243 / 271 bölüm başlıkları — mockup metniyle BİREBİR.
#: `general` bordro tipi DEĞİLDİR (spec §4) ama satırı varsa GİZLENMEZ
#: (`service.SECTION_ORDER` aynı kararı taşır): görünmeyen satır sessizce
#: kaybolmuş demektir.
SECTION_LABELS: dict[WorkerSource, str] = {
    WorkerSource.company: "ŞİRKET KADROSU — SGK 4a",
    WorkerSource.subcontractor: "TAŞERON İŞÇİSİ — SGK Taşeron",
    WorkerSource.freelance: "SERBEST MESLEK — Serbest Makbuz · %20 Stopaj",
    WorkerSource.intern: "STAJYER — Staj ücreti",
    WorkerSource.general: "GENEL İŞÇİ",
}

#: BY 150/170 rozetleri (`timesheet/export.py` ile AYNI metinler — iki dosyada
#: iki farklı rozet metni kullanıcıya iki farklı kavram öğretirdi).
SOURCE_LABELS: dict[WorkerSource, str] = {
    WorkerSource.company: "Şirket",
    WorkerSource.subcontractor: "Taşeron",
    WorkerSource.general: "Genel",
    WorkerSource.freelance: "Serbest",
    WorkerSource.intern: "Stajyer",
}

#: BY 148 "Beklemede". Diğer dört durumun metni mockup'ta YOKTUR (tablo tek bir
#: anı gösteriyor) ve ekranın kendi sözlüğünden gelemez — dosya sunucudan iner.
#: Metinler durumun ANLAMINI söyler, ham enum değeri BASILMAZ.
STATUS_LABELS: dict[PayrollLineStatus, str] = {
    PayrollLineStatus.uncomputed: "Hesaplanamadı",
    PayrollLineStatus.pending: "Beklemede",
    PayrollLineStatus.approved: "Onaylandı",
    PayrollLineStatus.paid: "Ödendi",
    PayrollLineStatus.excluded: "Bordrodan ödenmez",
}

#: Tanınmayan enum değerinin GÖRÜNÜR düşüşü (`timesheet/export.py` emsali).
UNKNOWN_LABEL_PREFIX = "?"

EMPTY_VALUE = "—"

#: Tablo başlık satırının numarası: şerit + bir boş ayraç satırı.
HEADER_ROW = len(INFO_LABELS) + 2

_COLUMN_WIDTHS = (28, 12, 8, 14, 14, 14, 14, 14, 20)

#: Toplam satırında net/banka/elden hangi sütuna düşer (BY 299-301).
_NET_COLUMN = COLUMN_HEADERS.index("Net")
_BANK_COLUMN = COLUMN_HEADERS.index("Banka")
_CASH_COLUMN = COLUMN_HEADERS.index("Elden")


def _label(sozluk: dict, anahtar: object) -> str:
    """Etiket erişiminin TEK yolu — **sözlük DOĞRUDAN İNDEKSLENMEZ.**

    🔴 T2 dersi: `SOZLUK[enum]` yazımı, paylaşılan bir enum genişlediğinde
    (İK-3 `worker_source`a iki değer ekledi) export'u sessizce `KeyError` → 500
    yapıyordu. Erişim dayanıklıdır ama düşüş SESSİZ DEĞİLDİR (WORKFLOW §3):
    etiket eklemeyi unutan sonraki dilim hücrede `?yeni_deger` görür.
    """
    return sozluk.get(anahtar, f"{UNKNOWN_LABEL_PREFIX}{getattr(anahtar, 'value', anahtar)}")


def status_label(status: object) -> str:
    return _label(STATUS_LABELS, status)


def section_label(source: object) -> str:
    return _label(SECTION_LABELS, source)


def source_label(source: object) -> str:
    return _label(SOURCE_LABELS, source)


def _money(value: object | None) -> str | None:
    """Para hücresi — `null` ise **`None`** döner ve hücreye DOKUNULMAZ (S4)."""
    return None if value is None else str(value)


def _period_label(detail: PayrollPeriodDetailResponse) -> str:
    return f"{detail.month:02d}.{detail.year}"


def _info_values(detail: PayrollPeriodDetailResponse) -> tuple[str, ...]:
    ozet: PayrollSummaryResponse = detail.summary
    return (
        _period_label(detail),
        detail.status.value,
        # Sunucu tarih ÜRETMEZ (T4b): tarih yoksa "—" basılır, uydurulmaz.
        str(detail.payment_due_date) if detail.payment_due_date else EMPTY_VALUE,
        str(ozet.net_total),
        str(ozet.bank_total),
        str(ozet.cash_total),
        str(ozet.total_employer_cost),
    )


def _line_cells(line: PayrollLineResponse) -> tuple[str | None, ...]:
    """BY 133-148'in bir satırı. `null` alanlar `None` kalır (S4)."""
    return (
        line.personnel_name,
        source_label(line.personnel_source),
        None if line.days is None else str(line.days),
        _money(line.gross_amount),
        _money(line.deduction_amount),
        _money(line.net_amount),
        _money(line.bank_amount),
        _money(line.cash_amount),
        status_label(line.status),
    )


def _write_row(sheet: Worksheet, row: int, values: tuple[str | None, ...]) -> None:
    for column, value in enumerate(values, start=1):
        if value is None:
            # `null` alana DOKUNULMAZ (docstring: boş niyeti böyle anlatılır).
            continue
        sheet.cell(row=row, column=column).value = value


def _write_info(sheet: Worksheet, detail: PayrollPeriodDetailResponse) -> None:
    for index, (label, value) in enumerate(
        zip(INFO_LABELS, _info_values(detail), strict=True), start=1
    ):
        sheet.cell(row=index, column=1).value = label
        sheet.cell(row=index, column=2).value = value


def _write_total(sheet: Worksheet, row: int, ozet: PayrollSummaryResponse) -> None:
    """BY 298-301 — etiketteki sayı ile tutarların TABANI AYNI DEĞİLDİR.

    Etiket dönemin TÜM satırlarını sayar (BY 298 "TOPLAM (48 çalışan)" =
    tfoot 12+29+5+2, taşeron DAHİL); net/banka/elden ise ÖDEME tabanıdır
    (`summary.py`: `excluded` taşeron ve `uncomputed` satır GİRMEZ, K2/S4).
    Tek tabana indirgenseydi ya sayı ya tutar yalan söylerdi.
    """
    hucreler: list[str | None] = [None] * len(COLUMN_HEADERS)
    hucreler[0] = f"{TOTAL_LABEL_PREFIX} ({ozet.line_count} çalışan)"
    hucreler[_NET_COLUMN] = str(ozet.net_total)
    hucreler[_BANK_COLUMN] = str(ozet.bank_total)
    hucreler[_CASH_COLUMN] = str(ozet.cash_total)
    _write_row(sheet, row, tuple(hucreler))


def _apply_layout(sheet: Worksheet) -> None:
    for index, width in enumerate(_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[
            sheet.cell(row=HEADER_ROW, column=index).column_letter
        ].width = width
    sheet.freeze_panes = sheet.cell(row=HEADER_ROW + 1, column=2).coordinate


def build_payroll_workbook(detail: PayrollPeriodDetailResponse) -> BytesIO:
    """Dönem detay zarfından xlsx üretir ve bellekteki tamponu döner.

    Satırsız dönemde de GEÇERLİ bir dosya döner: başlıklar ve sıfırlı toplam
    satırı basılır, uydurma bir kişi satırı üretilmez.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_TITLE

    _write_info(sheet, detail)
    _write_row(sheet, HEADER_ROW, COLUMN_HEADERS)

    row = HEADER_ROW
    for section in detail.sections:
        row += 1
        sheet.cell(row=row, column=1).value = section_label(section.personnel_source)
        for line in section.lines:
            row += 1
            _write_row(sheet, row, _line_cells(line))

    _write_total(sheet, row + 1, detail.summary)
    _apply_layout(sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def filename(year: int, month: int) -> str:
    """`bordro-{yıl}-{ay}.xlsx` — dönem dosya adında GÖRÜNÜR.

    İki ay aynı klasöre indiğinde dosyalar birbirini ezmemelidir.
    """
    return f"bordro-{year}-{month:02d}.xlsx"
