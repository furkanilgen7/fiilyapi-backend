"""Bordro GEÇMİŞİ Excel çıktısı (EXPORT-XLSX) — mockup BG "Excel İndir".

🔴 `payroll/export.py` ile KARIŞTIRILMAZ: o, TEK bir dönemin DETAYINI
(personel satırları) basar; bu modül DÖNEM LİSTESİNİ (BG tablosu) basar. İki
ayrı ekran, iki ayrı zarf, iki ayrı dosya — tek modülde birleştirilselerdi
hangi sütun kümesinin hangi ekrana ait olduğu kaybolurdu.

`units/export.py` · `audit/export.py` deseninin kardeşi: saf SUNUM katmanı.
`Request`/`Response` bilmez, DB'ye DOKUNMAZ, kapsam/yetki kararı VERMEZ —
yalnız `PayrollPeriodListRow` satırlarını çalışma kitabına çevirir.

## Hiçbir toplam BURADA hesaplanmaz

Satırların brütü/SGK'sı/neti/maliyeti `summary.build_period_summary`den
HESAPLANMIŞ gelir. BG'nin tfoot'u ("2026 Toplam (7 Ay) · Ort. 45 · …") dosyaya
BASILMAZ: liste zarfı böyle bir yıl toplamı TAŞIMAZ ve burada toplansaydı Excel
ekranın hiçbir yerinde olmayan İKİNCİ bir hesap kaynağı olurdu (K15 gerekçesi).

## FLOAT-YASAK (B4 dersi, `audit/export.py` emsali)

Her hücre AÇIKÇA `str` yazılır; `str(Decimal)` API yanıtının metniyle BİREBİR
aynıdır — yeniden yuvarlama YOK, kuruş aritmetiği YOK.

## `null` alana DOKUNULMAZ

Ödeme tarihi girilmemiş dönemde hücreye HİÇ dokunulmaz: sunucu tarih ÜRETMEZ
(`payroll/export.py` T4b kuralı).
"""

from collections.abc import Sequence
from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.modules.payroll.models import PayrollPeriodStatus
from app.modules.payroll.schemas import PayrollPeriodListRow

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

SHEET_TITLE = "Bordro Geçmişi"

#: BG thead'i BİREBİR (sıra ve metin). "Detay" bir BAĞLANTIDIR, veri değildir.
COLUMN_HEADERS: tuple[str, ...] = (
    "Dönem",
    "Çalışan",
    "Brüt Maaş",
    "SGK İşveren",
    "Net Ödenen",
    "Toplam Maliyet",
    "Ödeme Tarihi",
    "Durum",
)

#: BG durum rozeti. Mockup yalnız İKİ metni gösterir ("Bekliyor" + alt satır
#: "Ödeme bekliyor" · "Ödendi"); dosya sunucudan iner ve DÖRT durumun hepsini
#: karşılamak zorundadır. Metinler durumun ANLAMINI söyler, ham enum
#: değeri BASILMAZ.
PERIOD_STATUS_LABELS: dict[PayrollPeriodStatus, str] = {
    PayrollPeriodStatus.draft: "Taslak",
    PayrollPeriodStatus.pending_approval: "Onay Bekliyor",
    PayrollPeriodStatus.approved: "Ödeme Bekliyor",
    PayrollPeriodStatus.paid: "Ödendi",
}

#: Tanınmayan enum değerinin GÖRÜNÜR düşüşü (`payroll/export.py` emsali).
UNKNOWN_LABEL_PREFIX = "?"

EMPTY_VALUE = "—"

XLSX_FILENAME = "bordro-gecmisi.xlsx"

_COLUMN_WIDTHS = (12, 10, 16, 16, 16, 18, 14, 16)


def _label(sozluk: dict, anahtar: object) -> str:
    """Etiket erişiminin TEK yolu — **sözlük DOĞRUDAN İNDEKSLENMEZ.**

    Geçiş zinciri bir gün bir durum daha eklerse (`payroll/export.py` T2 dersi)
    `SOZLUK[enum]` yazımı export'u sessizce 500 yapardı. Düşüş dayanıklıdır ama
    SESSİZ DEĞİLDİR: hücrede `?yeni_deger` görünür.
    """
    return sozluk.get(anahtar, f"{UNKNOWN_LABEL_PREFIX}{getattr(anahtar, 'value', anahtar)}")


def status_label(status: object) -> str:
    return _label(PERIOD_STATUS_LABELS, status)


def period_label(row: PayrollPeriodListRow) -> str:
    """`07.2026` — dönem etiketinin kanonik biçimi (`payroll/export.py` ile aynı)."""
    return f"{row.month:02d}.{row.year}"


def _cells(row: PayrollPeriodListRow) -> tuple[str | None, ...]:
    return (
        period_label(row),
        str(row.personnel_count),
        str(row.gross_total),
        str(row.sgk_employer_total),
        str(row.net_total),
        str(row.total_cost),
        # `date` → ISO metin; sunucu tarih UYDURMAZ, yoksa hücreye dokunulmaz.
        None if row.payment_due_date is None else str(row.payment_due_date),
        status_label(row.status),
    )


def _write_row(sheet: Worksheet, row: int, values: tuple[str | None, ...]) -> None:
    for column, value in enumerate(values, start=1):
        if value is None:
            # `null` alana DOKUNULMAZ.
            continue
        sheet.cell(row=row, column=column).value = value


def _apply_layout(sheet: Worksheet) -> None:
    for index, width in enumerate(_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width
    sheet.freeze_panes = "A2"


def build_period_history_workbook(rows: Sequence[PayrollPeriodListRow]) -> BytesIO:
    """Dönem satırlarından xlsx üretir ve bellekteki tamponu döner.

    Hiç dönem açılmamışken de GEÇERLİ bir dosya döner — yalnız başlık satırı.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_TITLE

    _write_row(sheet, 1, COLUMN_HEADERS)
    for index, row in enumerate(rows, start=2):
        _write_row(sheet, index, _cells(row))
    _apply_layout(sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def filename() -> str:
    """`bordro-gecmisi.xlsx` — dönem DETAYININ `bordro-{yıl}-{ay}.xlsx` adıyla
    çakışmaz; iki dosya aynı klasöre indiğinde birbirini EZMEZ."""
    return XLSX_FILENAME
